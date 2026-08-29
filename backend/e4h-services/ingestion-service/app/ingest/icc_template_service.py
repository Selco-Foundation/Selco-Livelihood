"""Assembles the IC Report template workbook the Project Manager downloads.

The workbook itself comes out of filestore untouched -- it is the blank template seeded per
Solution. All this module does is append the read-only reference sheet, which is the "made
any updates if needed" half of the e4h filestore pattern.
"""
from typing import Any, Dict, List

from openpyxl.styles import Font, Protection
from openpyxl.worksheet.worksheet import Worksheet

from app.core.logging import AppLogger

logger = AppLogger().get_logger()

SITES_SHEET = "Sites"
SITES_HEADERS = ("End User Site", "Site Id", "State", "District", "Block", "Status")


def append_sites_sheet(workbook, sites: List[Dict[str, Any]]) -> Worksheet:
    """Add a read-only sheet listing the sites this template covers.

    Reference only -- none of it is read back on upload. It earns its place by telling the
    Project Manager which sites a Solution's single template applies to, and by flagging any
    site already published in a sibling plan so the problem surfaces here rather than at
    Vendor Assignment.

    Replaces the sheet if it already exists, so a re-download does not accumulate copies.
    """
    if SITES_SHEET in workbook.sheetnames:
        del workbook[SITES_SHEET]

    sheet = workbook.create_sheet(SITES_SHEET)
    for column, header in enumerate(SITES_HEADERS, start=1):
        cell = sheet.cell(row=1, column=column, value=header)
        cell.font = Font(bold=True)

    for row_index, site in enumerate(sites, start=2):
        values = (
            site.get("name") or "",
            site.get("facility_id") or "",
            site.get("state") or "",
            site.get("district") or "",
            site.get("block") or "",
            site.get("status") or "",
        )
        for column, value in enumerate(values, start=1):
            sheet.cell(row=row_index, column=column, value=value)

    # Every cell locked: this sheet is context, not input. Sheet protection has to be switched
    # on for cell-level locking to take effect at all in Excel.
    for row in sheet.iter_rows(min_row=1, max_row=max(sheet.max_row, 1),
                               max_col=len(SITES_HEADERS)):
        for cell in row:
            cell.protection = Protection(locked=True)
    sheet.protection.sheet = True

    for column, header in enumerate(SITES_HEADERS, start=1):
        longest = max([len(str(header))] + [
            len(str(sheet.cell(row=r, column=column).value or ""))
            for r in range(2, sheet.max_row + 1)
        ])
        sheet.column_dimensions[sheet.cell(row=1, column=column).column_letter].width = \
            min(max(longest + 2, 12), 45)

    logger.info(f"Appended {SITES_SHEET} sheet with {len(sites)} site(s)")
    return sheet
