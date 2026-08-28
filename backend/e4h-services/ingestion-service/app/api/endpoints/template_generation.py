from datetime import datetime
from typing import Optional, List
import psycopg2

import pandas as pd
from PIL import ImageDraw, Image, ImageFont
from fastapi import APIRouter, Form, HTTPException, Depends, Body
from fastapi.responses import FileResponse
from fastapi import BackgroundTasks
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Protection, PatternFill
from openpyxl.utils import get_column_letter

from app.core.logging import AppLogger
from app.core.tenant import LIVELIHOOD_TENANT_ID
from app.decorators.rbac_validator import get_authorized_request_info
from app.ingest.facility_template_service import FacilityTemplateService
from app.ingest.asset_template_service import AssetTemplateService
from app.ingest.project_service import ProjectService
from app.schemas.boundary import Boundary, flatten_boundaries
from app.utils.amc_scheduler_service_client import AMCSchedulerServiceClient
from app.utils.convertor import request_info_from_json
from app.utils.excel_utils import add_dropdowns_to_excel, autofit_columns, lock_prefilled_rows_in_excel, \
    lock_excel_columns
from app.utils.facility_service_client import FacilityServiceClient
from app.utils.fieldplan_activity_service_client import FieldPlanActivityServiceClient
from app.utils.fieldplan_service_client import FieldPlanServiceClient
from app.utils.file_utils import create_temp_file, cleanup_temp_file
from app.utils.mdms_client import MDMSClient
from app.utils.project_service_client import ProjectServiceClient
from app.utils.field_plan_locks import build_project_lock_map, lock_status_label, solution_names_by_code
from app.utils.solution_eligibility import build_solution_options_by_row, clear_solution_column_dropdown
from app.utils.state_sunshine_hours_repository import fetch_state_sunshine_hours
from app.utils.vendor_registry_client import VendorRegistryClient
import os, tempfile, zipfile, qrcode, shutil

router = APIRouter()
logger = AppLogger().get_logger()

from dotenv import load_dotenv

load_dotenv()
mdms_url = os.getenv("MDMS_URL")
project_service_url = os.getenv("PROJECT_SERVICE_URL")
facility_service_url = os.getenv("FACILITY_SERVICE_URL")
fieldPlan_service_url = os.getenv("FIELDPLAN_SERVICE_URL")
fieldPlan_activity_service_url = os.getenv("FIELDPLAN_ACTIVITY_SERVICE_URL")
amc_scheduler_service_url = os.getenv("AMC_SCHEDULER_SERVICE_URL")
vendor_service_url = os.getenv("VENDOR_SERVICE_URL")
DEFAULT_AMC_ASSET_TYPES = ["INVERTER", "PANEL", "BATTERY"]
DB_CONFIG = {
    "host": os.getenv("DB_HOST"),
    "port": int(os.getenv("DB_PORT", 5432)),
    "database": os.getenv("DB_NAME"),
    "user": os.getenv("DB_USER"),
    "password": os.getenv("DB_PASSWORD")
}

@router.post('/facilityIngestionTemplateWithData',
            summary='Generate facility ingestion template Excel file with schema, already present data and boundary codes',
            response_description="Returns Excel template with facility schema, facility data and boundary codes")
async def get_facility_ingestion_template_with_data(
        background_tasks: BackgroundTasks,
        facility_service: FacilityTemplateService = Depends(),
    payload: dict = Body(..., description="Payload object")
):
    logger.trace("Starting facility ingestion template generation with data")
    request_info = request_info_from_json(payload.get("RequestInfo", {}))
    boundary_data = payload.get("boundary_data", {})
    project_id = payload.get("project_id")
    logger.info(f"Generating facility ingestion template: project_id={project_id}, boundaries={len(boundary_data) if boundary_data else 0}")
    
    mdms_client = MDMSClient(mdms_url)
    try:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"facility_ingestion_template_{timestamp}.xlsx"
        output_file_path = create_temp_file(suffix=".xlsx")
        logger.debug(f"Created temporary file: {output_file_path}")
        
        try:
            logger.info("Fetching facility schema from MDMS")
            facility_schema = mdms_client.get_column_definitions_with_metadata(request_info, 'data-ingestion.FacilityIngestionSchema')
            logger.debug(f"Retrieved facility schema with {len(facility_schema) if facility_schema else 0} columns")
            
            logger.info("Flattening boundary data")
            boundary_list: List[Boundary] = flatten_boundaries(boundary_data)
            logger.debug(f"Flattened {len(boundary_list)} boundaries")
        except Exception as e:
            logger.error(f"Error fetching data from external services: {e}", exc_info=True)
            cleanup_temp_file(output_file_path)
            raise HTTPException(status_code=502, detail=f"External service error: {str(e)}")

        logger.info("Fetching facilities by boundary codes")
        all_facilities = []
        if facility_service_url and boundary_list:
            facility_client = FacilityServiceClient(facility_service_url)

            # Deduplicate boundaries by code to avoid redundant calls
            unique_boundaries = {}
            for boundary in boundary_list:
                if boundary.code and boundary.code not in unique_boundaries:
                    unique_boundaries[boundary.code] = boundary
            unique_boundary_list = list(unique_boundaries.values())

            logger.info(f"Total unique boundary codes for facility search: {len(unique_boundary_list)}")

            # Use bulk facility search by boundary codes to reduce number of API calls
            boundary_codes = [b.code for b in unique_boundary_list if b.code]
            try:
                if boundary_codes:
                    bulk_result = facility_client.bulk_search_facility_with_boundary(
                        request_info=request_info,
                        tenant_ids=[LIVELIHOOD_TENANT_ID],
                        boundary_codes=boundary_codes,
                        limit=max(len(boundary_codes) * 50, 50),
                        send_non_paginated_response=True,
                    )
                    facilities = bulk_result.get("facilities", []) or []
                    all_facilities.extend(facilities)
                    logger.info(
                        f"Fetched {len(facilities)} facilities from bulk facility search for "
                        f"{len(boundary_codes)} boundary codes"
                    )
            except Exception as e:
                logger.error(
                    f"Error fetching facilities via bulk boundary facility search: {e}",
                    exc_info=True,
                )

        # Fetch project-linked facilities if project_id is provided
        logger.info(f"Fetching project-linked facilities: project_id={project_id}")
        project_linked_facility_ids = set()
        project_facilities_data = []
        if project_id and project_service_url:
            try:
                project_client = ProjectServiceClient(project_service_url)
                project_facilities_response = project_client.search_project_facility(request_info, project_id)
                project_facilities = project_facilities_response.get("ProjectFacilities", [])
                project_linked_facility_ids = {pf.get("facilityId") for pf in project_facilities if pf.get("facilityId")}
                logger.info(f"Found {len(project_linked_facility_ids)} facilities linked to project {project_id}")
                logger.debug(f"Project facility IDs: {list(project_linked_facility_ids)[:10]}...")  # Log first 10 IDs

                # Optimization: avoid redundant facility-service calls for facilities
                # that are already present from boundary-based search, and fetch the
                # remaining project facilities via bulk search.
                existing_boundary_facility_ids = {f.get("facility_id") for f in all_facilities}
                project_facilities_to_fetch = [
                    pf for pf in project_facilities
                    if pf.get("facilityId") and pf.get("facilityId") not in existing_boundary_facility_ids
                ]

                logger.info(
                    f"Project facilities total={len(project_facilities)}, "
                    f"already_in_boundaries={len(existing_boundary_facility_ids)}, "
                    f"to_fetch_from_facility_service={len(project_facilities_to_fetch)}"
                )

                if facility_service_url and project_facilities_to_fetch:
                    facility_ids_to_fetch = [
                        pf.get("facilityId")
                        for pf in project_facilities_to_fetch
                        if pf.get("facilityId")
                    ]
                    try:
                        if facility_ids_to_fetch:
                            bulk_result = facility_client.bulk_search_facility(
                                request_info=request_info,
                                tenant_ids=[LIVELIHOOD_TENANT_ID],
                                facility_ids=facility_ids_to_fetch,
                                limit=max(len(facility_ids_to_fetch), 50),
                                send_non_paginated_response=True,
                            )
                            facilities = bulk_result.get("facilities", []) or []
                            project_facilities_data.extend(facilities)
                            logger.info(
                                f"Fetched {len(facilities)} facilities from bulk facility search for project {project_id}"
                            )
                    except Exception as e:
                        logger.error(
                            f"Error fetching project facilities via bulk facility search: {e}",
                            exc_info=True,
                        )

            except Exception as e:
                logger.error(f"Error fetching project facilities: {e}")
                # Continue without project facility data if there's an error

        # Combine boundary facilities with project facilities (avoid duplicates)
        # Only include project facilities that belong to the current boundary codes
        existing_facility_ids = {f.get('facility_id') for f in all_facilities}
        valid_boundary_codes = {boundary.code for boundary in boundary_list}
        
        for pf_facility in project_facilities_data:
            facility_id = pf_facility.get('facility_id')
            facility_boundary_code = pf_facility.get('boundary_code') or pf_facility.get('boundaryCode')
            
            # Only add if not already present and belongs to current boundary codes
            if (facility_id not in existing_facility_ids and 
                facility_boundary_code in valid_boundary_codes):
                all_facilities.append(pf_facility)
                logger.info(f"Added project facility {facility_id} to template (boundary: {facility_boundary_code})")
            elif facility_boundary_code not in valid_boundary_codes:
                logger.info(f"Skipped project facility {facility_id} - boundary code {facility_boundary_code} not in current boundary list")

        logger.info(f"Total facilities in template: {len(all_facilities)} (boundary: {len(existing_facility_ids)}, project: {len(project_facilities_data)})")

        # Mark facilities as included in project if they are already linked
        logger.info("Marking facilities with project inclusion status")
        if project_id:
            linked_count = 0
            for facility in all_facilities:
                facility_id = facility.get("facility_id")
                if facility_id in project_linked_facility_ids:
                    facility["include_in_project"] = "Yes"
                    linked_count += 1
                    logger.trace(f"Facility {facility_id} is linked to project - marking as Yes")
                else:
                    facility["include_in_project"] = "No"
                    logger.trace(f"Facility {facility_id} is NOT linked to project - marking as No")
            logger.debug(f"Marked {linked_count} facilities as linked to project")
        else:
            # If no project_id provided, set all facilities to "No"
            logger.debug("No project_id provided - marking all facilities as No")
            for facility in all_facilities:
                facility["include_in_project"] = "No"

        try:
            logger.info("Generating template file with facility data")
            facility_service.generate_template_file_with_data(
                output_path=output_file_path,
                facility_schema=facility_schema,
                boundary_list=boundary_list,
                facility_data=all_facilities,
                type="project",
                extra_append_rows=200,
                optimize_for_performance=True
            )
            logger.info(f"Successfully created facility ingestion template: {output_filename}")
            logger.debug(f"Template file path: {output_file_path}")
        except Exception as e:
            logger.error(f"Error generating template file: {e}", exc_info=True)
            cleanup_temp_file(output_file_path)
            raise HTTPException(status_code=500, detail=f"Template generation error: {str(e)}")
        background_tasks.add_task(cleanup_temp_file, output_file_path)
        logger.info(f"Template generation completed successfully: {output_filename}")
        return FileResponse(
            path=output_file_path,
            filename=output_filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        logger.error(f"Unhandled error in get_facility_ingestion_template: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"An unexpected error occurred: {str(e)}")


@router.post('/boundaryIngestionTemplate',
             summary='Generate empty boundary ingestion template Excel file',
             response_description="Returns an empty Excel template for boundary ingestion")
async def get_boundary_ingestion_template(
        background_tasks: BackgroundTasks,
        payload: dict = Body(..., description="Payload object")
):
    """
    Generate an empty boundary ingestion template with the required columns.
    Sheet name: 'Boundary Data'
    Columns: Country, State, District, Block
    """
    output_file_path = None
    request_info = request_info_from_json(payload.get("RequestInfo", {}))

    try:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"boundary_ingestion_template_{timestamp}.xlsx"
        output_file_path = create_temp_file(suffix=".xlsx")

        # Create an empty DataFrame with the expected boundary columns
        df = pd.DataFrame(columns=["Country", "State", "District", "Block"])
        df.to_excel(output_file_path, sheet_name="Boundary Data", index=False)

        logger.info(f"Successfully created boundary ingestion template at {output_file_path}")

        background_tasks.add_task(cleanup_temp_file, output_file_path)
        return FileResponse(
            path=output_file_path,
            filename=output_filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        logger.error(f"Unhandled error in get_boundary_ingestion_template: {e}")
        if output_file_path:
            cleanup_temp_file(output_file_path)
        raise HTTPException(status_code=500, detail=f"An unexpected error occurred: {str(e)}")


@router.post('/fieldplanFacilityIngestionTemplate',
            summary='Generate facility ingestion template Excel file with schema, already present data and boundary codes',
            response_description="Returns Excel template with facility schema, facility data and boundary codes")
async def get_facility_ingestion_template_with_data(
        background_tasks: BackgroundTasks,
        facility_service: FacilityTemplateService = Depends(),
    payload: dict = Body(..., description="Payload object")
):
    request_info = request_info_from_json(payload.get("RequestInfo", {}))
    # RequestInfo = request_info_from_json(payload.get("RequestInfo", {}))
    boundary_data = payload.get("boundary_data", {})
    fieldplan_id = payload.get("fieldplan_id")
    project_id = payload.get("project_id")
    # Sent by the caller rather than read back off the plan: screen 1 already chose it, and
    # this keeps the endpoint read-only.
    plan_sector = payload.get("sector")
    mdms_client = MDMSClient(mdms_url)
    fieldplan_activity_client = FieldPlanActivityServiceClient(fieldPlan_activity_service_url)
    try:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"facility_ingestion_template_{timestamp}.xlsx"
        output_file_path = create_temp_file(suffix=".xlsx")
        try:
            facility_schema = mdms_client.get_column_definitions_with_metadata(request_info, 'data-ingestion.InstallationScopeIngestionSchema')
            boundary_list: List[Boundary] = flatten_boundaries(boundary_data)
        except Exception as e:
            logger.error(f"Error fetching data from external services: {e}")
            cleanup_temp_file(output_file_path)
            raise HTTPException(status_code=502, detail=f"External service error: {str(e)}")

        # Get all project-linked facilities
        project_client = ProjectServiceClient(project_service_url)
        project_facilities_response = project_client.search_project_facility(request_info, project_id)
        project_facilities = project_facilities_response.get("ProjectFacilities", [])
        project_linked_facility_ids = {pf.get("facilityId") for pf in project_facilities if pf.get("facilityId")}
        logger.info(f"Found {len(project_linked_facility_ids)} facilities currently linked to project {project_id}")

        all_facilities = []
        facility_client = FacilityServiceClient(facility_service_url) if facility_service_url else None
        if facility_client and project_linked_facility_ids and boundary_list:
            boundary_codes = [b.code for b in boundary_list if b.code]
            try:
                boundary_bulk_result = facility_client.bulk_search_facility_with_boundary(
                    request_info=request_info,
                    tenant_ids=[LIVELIHOOD_TENANT_ID],
                    boundary_codes=boundary_codes,
                    limit=max(len(boundary_codes) * 50, 50),
                    send_non_paginated_response=True,
                )
                boundary_facilities = boundary_bulk_result.get("facilities", []) or []
                all_facilities = [
                    f for f in boundary_facilities
                    if f.get("facility_id") in project_linked_facility_ids
                ]
            except Exception as e:
                logger.error(f"Error fetching boundary facilities in bulk: {e}", exc_info=True)

        # Fetch fieldplan-linked facilities if fieldplan_id is provided
        fieldplan_linked_facility_ids = set()
        fieldplan_facilities_data = []
        if fieldplan_id and fieldPlan_service_url:
            try:
                fieldplan_client = FieldPlanServiceClient(fieldPlan_service_url)
                fieldplan_facilities_response = fieldplan_client.search_fieldplan_facility(request_info, fieldplan_id)
                fieldplan_facilities = fieldplan_facilities_response.get("FieldPlanFacilities", [])
                fieldplan_linked_facility_ids = {pf.get("facilityId") for pf in fieldplan_facilities if
                                               pf.get("facilityId")}
                logger.info(
                    f"Found {len(fieldplan_linked_facility_ids)} facilities linked to fieldplan {fieldplan_id}")

                # Fetch all fieldplan-linked facility details in one bulk call
                if facility_client and fieldplan_linked_facility_ids:
                    facility_ids = list(fieldplan_linked_facility_ids)
                    try:
                        facilities_bulk_result = facility_client.bulk_search_facility(
                            request_info=request_info,
                            tenant_ids=[LIVELIHOOD_TENANT_ID],
                            facility_ids=facility_ids,
                            limit=max(len(facility_ids), 50),
                            send_non_paginated_response=True,
                        )
                        fieldplan_facilities_data.extend(facilities_bulk_result.get("facilities", []) or [])
                    except Exception as e:
                        logger.error(f"Error bulk fetching fieldplan facilities: {e}")

            except Exception as e:
                logger.error(f"Error fetching fieldplan facilities: {e}")
                # Continue without fieldplan facility data if there's an error

        # Combine boundary facilities with fieldplan facilities (avoid duplicates)
        # Only include fieldplan facilities that belong to the current boundary codes
        existing_facility_ids = {f.get('facility_id') for f in all_facilities}
        valid_boundary_codes = {boundary.code for boundary in boundary_list}

        for pf_facility in fieldplan_facilities_data:
            facility_id = pf_facility.get('facility_id')
            facility_boundary_code = pf_facility.get('boundary_code') or pf_facility.get('boundaryCode')

            # Only add if not already present and belongs to current boundary codes
            if (facility_id not in existing_facility_ids and facility_id in project_linked_facility_ids and facility_boundary_code in valid_boundary_codes):
                all_facilities.append(pf_facility)
                logger.info(
                    f"Added fieldplan facility {facility_id} to template (boundary: {facility_boundary_code})")
            elif (facility_id not in project_linked_facility_ids): # In case the facility is no longer mapped to project, so unlink the facility
                fieldPlan_facility_data = next(
                    (pf for pf in fieldplan_facilities if pf.get("facilityId") == facility_id),
                    None)
                fieldplan_client.unlink_fieldplan_facility(
                    request_info=request_info,
                    fieldplan_id=fieldplan_id,
                    facility_id=facility_id,
                    fieldplan_facility_data=fieldPlan_facility_data
                )

                facilities_activity_response = fieldplan_activity_client.search_facility_activity(
                    request_info, fieldplan_id, facility_id)
                facilities_activity = facilities_activity_response.get("FacilityActivities", [])
                facility_activity_ids = list({fa.get("activityFacility").get("id") for fa in facilities_activity if
                                              fa.get("activityFacility").get("id")})
                fieldplan_activity_client.delete_facility_activity(request_info=request_info,
                                                                   facility_activity_id=facility_activity_ids)
            elif facility_boundary_code not in valid_boundary_codes:
                logger.info(
                    f"Skipped fieldplan facility {facility_id} - boundary code {facility_boundary_code} not in current boundary list")

        logger.info(
            f"Total facilities in template: {len(all_facilities)} (boundary: {len(existing_facility_ids)}, fieldplan: {len(fieldplan_facilities_data)})")

        # Mark facilities as included in fieldplan if they are already linked
        if fieldplan_id:
            for facility in all_facilities:
                facility_id = facility.get("facility_id")
                if facility_id in fieldplan_linked_facility_ids:
                    facility["include_in_fieldplan"] = "Yes"
                    logger.info(f"Facility {facility_id} is linked to fieldplan - marking as Yes")
                else:

                    facility["include_in_fieldplan"] = "No"
                    logger.info(f"Facility {facility_id} is NOT linked to fieldplan - marking as No")
        else:
            # If no fieldplan_id provided, set all facilities to "No"
            for facility in all_facilities:
                facility["include_in_fieldplan"] = "No"
                logger.info(f"No fieldplan_id provided - marking facility {facility.get('facility_id')} as No")

        # One sector per plan, chosen on screen 1 and sent with this request, so the Sector
        # column is identical on every row and Solution options vary only by the site's
        # state (FR-01). Sites tagged with any other sector are out of this plan's scope.
        if plan_sector:
            wanted_sector = str(plan_sector).strip().casefold()
            before_count = len(all_facilities)
            all_facilities = [
                f for f in all_facilities
                if str(f.get("facility_type") or "").strip().casefold() == wanted_sector
            ]
            logger.info(
                f"Filtered facilities by sector '{plan_sector}': {len(all_facilities)} of {before_count}")
        else:
            logger.warning("No sector supplied; skipping sector filter and Solution dropdowns")

        solutions = []
        solution_options_by_row = {}
        if plan_sector:
            try:
                solutions = mdms_client.fetch_installation_solutions(request_info)
                solution_options_by_row = build_solution_options_by_row(
                    facilities=all_facilities,
                    solutions=solutions,
                    plan_sector=plan_sector,
                    sunshine_hours_by_state=fetch_state_sunshine_hours(),
                )
            except Exception as e:
                logger.error(f"Error resolving eligible solutions: {e}", exc_info=True)

        # Sites already under installation -- in this plan or a sibling plan in the same
        # project -- are shown as-is and frozen, so the PM can see they are spoken for but
        # cannot re-scope them. The lock is project-scoped, hence the project-wide lookup.
        lock_map = {}
        if project_id and fieldPlan_service_url:
            lock_map = build_project_lock_map(
                FieldPlanServiceClient(fieldPlan_service_url), request_info, project_id, fieldplan_id
            )

        solution_name_by_code = solution_names_by_code(solutions)
        freeze_row_positions = []
        lock_status_by_row = {}
        for position, facility in enumerate(all_facilities):
            lock = lock_map.get(facility.get("facility_id"))
            lock_status_by_row[position] = lock_status_label(lock)
            if lock is None:
                continue
            freeze_row_positions.append(position)
            # Show the locking plan's choice, including a sibling plan's, so the row explains itself.
            facility["include_in_fieldplan"] = "Yes"
            facility["locked_solution_name"] = solution_name_by_code.get(lock.solution_id, "")
            # A frozen value must not be offered as changeable.
            solution_options_by_row.pop(position, None)

        # Exactly one Solution per site, filtered per row by that site's state.
        facility_schema = clear_solution_column_dropdown(facility_schema)

        try:
            facility_service.generate_template_file_with_data(
                output_path=output_file_path,
                facility_schema=facility_schema,
                boundary_list=boundary_list,
                facility_data=all_facilities,
                type="fieldplan",
                extra_append_rows=0,
                optimize_for_performance=True,
                constant_column_values={"Sector": plan_sector or ""},
                row_specific_dropdowns={"Solution": solution_options_by_row},
                per_row_column_values={
                    "Solution": {p: all_facilities[p].get("locked_solution_name", "") for p in freeze_row_positions},
                    "Lock Status": lock_status_by_row,
                },
                freeze_columns=["Included in Field Plan", "Solution"],
                freeze_row_positions=freeze_row_positions,
            )
            logger.info(f"Successfully created facility ingestion template at {output_file_path}")
        except Exception as e:
            logger.error(f"Error generating template file: {e}")
            cleanup_temp_file(output_file_path)
            raise HTTPException(status_code=500, detail=f"Template generation error: {str(e)}")
        background_tasks.add_task(cleanup_temp_file, output_file_path)
        return FileResponse(
            path=output_file_path,
            filename=output_filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        logger.error(f"Unhandled error in get_facility_ingestion_template: {e}")
        raise HTTPException(status_code=500, detail=f"An unexpected error occurred: {str(e)}")


@router.post('/facilityIngestion',
            summary='Generate facility ingestion template Excel file with schema and boundary codes',
            response_description="Returns Excel template with facility schema and boundary codes")
async def get_facility_ingestion_template(
        facility_service: FacilityTemplateService = Depends(),
        request_info: str = Form(default="")
):
    request_info = request_info_from_json(request_info)
    mdms_client = MDMSClient(mdms_url)
    try:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"end_user_ingestion_template_{timestamp}.xlsx"
        output_file_path = create_temp_file(suffix=".xlsx")
        try:
            facility_schema = mdms_client.get_column_definitions_with_metadata(request_info, 'data-ingestion.FacilityIngestionSchemaWithoutBoundaryCode')
            boundary_data = facility_service.get_all_boundaries(request_info)
        except Exception as e:
            logger.error(f"Error fetching data from external services: {e}")
            cleanup_temp_file(output_file_path)
            raise HTTPException(status_code=502, detail=f"External service error: {str(e)}")

        try:
            facility_service.generate_template_file(
                output_path=output_file_path,
                facility_schema=facility_schema,
                boundary_data=boundary_data
            )
            logger.info(f"Successfully created facility ingestion template at {output_file_path}")
        except Exception as e:
            logger.error(f"Error generating template file: {e}")
            cleanup_temp_file(output_file_path)
            raise HTTPException(status_code=500, detail=f"Template generation error: {str(e)}")

        return FileResponse(
            path=output_file_path,
            filename=output_filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        logger.error(f"Unhandled error in get_facility_ingestion_template: {e}")
        raise HTTPException(status_code=500, detail=f"An unexpected error occurred: {str(e)}")

@router.post('/assetIngestion',
            summary='Generate asset ingestion template Excel file from the asset schema',
            response_description="Returns Excel template with asset schema columns")
async def get_asset_ingestion_template(request_info: str = Form(default="")):
    request_info = request_info_from_json(request_info)
    mdms_client = MDMSClient(mdms_url)
    asset_service = AssetTemplateService()
    try:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"asset_ingestion_template_{timestamp}.xlsx"
        output_file_path = create_temp_file(suffix=".xlsx")
        try:
            asset_schema = mdms_client.get_column_definitions_with_metadata(
                request_info, 'data-ingestion.AssetIngestionSchema')
        except Exception as e:
            logger.error(f"Error fetching asset schema from MDMS: {e}")
            cleanup_temp_file(output_file_path)
            raise HTTPException(status_code=502, detail=f"External service error: {str(e)}")

        facility_data = []
        if facility_service_url:
            try:
                facility_client = FacilityServiceClient(facility_service_url)
                bulk_result = facility_client.bulk_search_facility_with_boundary(
                    request_info=request_info,
                    tenant_ids=[LIVELIHOOD_TENANT_ID],
                    limit=10000,
                    send_non_paginated_response=True,
                )
                facility_data = bulk_result.get("facilities", []) or []
            except Exception as e:
                logger.error(f"Error fetching facility data for asset template: {e}")

        vendor_records = []
        if vendor_service_url:
            try:
                vendor_client = VendorRegistryClient(vendor_service_url)
                vendor_records = vendor_client.get_all_vendor_codes(request_info)
            except Exception as e:
                logger.error(f"Error fetching vendor codes for asset template: {e}")

        try:
            asset_service.generate_asset_template_file(
                output_path=output_file_path,
                asset_schema=asset_schema,
                facility_data=facility_data,
                vendor_records=vendor_records,
            )
            logger.info(f"Successfully created asset ingestion template at {output_file_path}")
        except Exception as e:
            logger.error(f"Error generating asset template file: {e}")
            cleanup_temp_file(output_file_path)
            raise HTTPException(status_code=500, detail=f"Template generation error: {str(e)}")

        return FileResponse(
            path=output_file_path,
            filename=output_filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Unhandled error in get_asset_ingestion_template: {e}")
        raise HTTPException(status_code=500, detail=f"An unexpected error occurred: {str(e)}")

@router.post('/facilityWithStaff',
            summary='Generate facility ingestion template with staff Excel file',
            response_description="Returns Excel template with facility schema")
async def get_facility_ingestion_template_with_staff(
        parent_id: str = Form(..., description="Parent project ID"),
        request_info: str = Form(..., description="Serialized RequestInfo JSON")
):
    temp_dir = tempfile.gettempdir()
    ts = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    output_filename = f"facility_staff_template_{parent_id}_{ts}.xlsx"
    output_file_path = os.path.join(temp_dir, output_filename)

    try:
        request_info = request_info_from_json(request_info)
        get_authorized_request_info(request_info)

        project_service = ProjectService()
        facilities = project_service.get_facilities(request_info, parent_id, "Staff")
        facility_template_service = FacilityTemplateService()

        try:
            original_df = pd.DataFrame(facilities)
            df = facility_template_service.add_supervisor_columns_to_dataframe(original_df)

            with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Facilities_Staff')
                worksheet = writer.sheets['Facilities_Staff']
                for i, col in enumerate(df.columns):
                    column_width = max(df[col].astype(str).map(len).max(), len(col)) + 2
                    worksheet.column_dimensions[get_column_letter(i + 1)].width = column_width

            dropdowns_map = {'Role (Mandatory) ?': ['Staff', 'Field Planner']}
            add_dropdowns_to_excel(
                file_path=output_file_path,
                sheet_name="Facilities_Staff",
                dropdowns=dropdowns_map
            )

        except Exception as e:
            logger.error(f"Error generating template file: {e}")
            cleanup_temp_file(output_file_path)
            raise HTTPException(status_code=500, detail=f"Template generation error: {str(e)}")

        return FileResponse(
            path=output_file_path,
            filename=output_filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        logger.error(f"Unhandled error in get_facility_ingestion_template_with_staff: {e}")
        cleanup_temp_file(output_file_path)
        raise HTTPException(status_code=500, detail=f"An unexpected error occurred: {str(e)}")

@router.post('/facilityWithSupervisors',
            summary='Generate facility ingestion template with supervisors Excel file',
            response_description="Returns Excel template with facility schema")
async def get_facility_ingestion_template_with_supervisors(
        parent_id: str = Form(..., description="Parent project ID"),
        request_info: str = Form(..., description="Serialized RequestInfo JSON")
):
    temp_dir = tempfile.gettempdir()
    ts = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    output_filename = f"facility_supervisors_template_{parent_id}_{ts}.xlsx"
    output_file_path = os.path.join(temp_dir, output_filename)

    try:
        request_info = request_info_from_json(request_info)
        get_authorized_request_info(request_info)

        project_service = ProjectService()
        facilities = project_service.get_facilities(request_info, parent_id, "Supervisor")
        facility_template_service = FacilityTemplateService()

        try:
            original_df = pd.DataFrame(facilities)
            df = facility_template_service.add_supervisor_columns_to_dataframe(original_df)

            with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Facilities_Supervisors')
                worksheet = writer.sheets['Facilities_Supervisors']
                for i, col in enumerate(df.columns):
                    column_width = max(df[col].astype(str).map(len).max(), len(col)) + 2
                    worksheet.column_dimensions[get_column_letter(i + 1)].width = column_width

            dropdowns_map = {'Role (Mandatory) ?': ['Supervisor', 'Field Planner']}
            add_dropdowns_to_excel(
                file_path=output_file_path,
                sheet_name="Facilities_Supervisors",
                dropdowns=dropdowns_map
            )

        except Exception as e:
            logger.error(f"Error generating template file: {e}")
            cleanup_temp_file(output_file_path)
            raise HTTPException(status_code=500, detail=f"Template generation error: {str(e)}")

        return FileResponse(
            path=output_file_path,
            filename=output_filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        logger.error(f"Unhandled error in get_facility_ingestion_template_with_supervisors: {e}")
        cleanup_temp_file(output_file_path)
        raise HTTPException(status_code=500, detail=f"An unexpected error occurred: {str(e)}")

@router.get('/facilitySelection',
            summary='Generate facility selection template Excel file',
            response_description="Returns Excel template with facility data")
async def get_facility_selection_template(
        facility_service: FacilityTemplateService = Depends(),
        parent_project_id: Optional[str] = Form(default=None),
        boundary_codes: str = Form(...),
        request_info: str = Form(default="")
):
    request_info = request_info_from_json(request_info)
    get_authorized_request_info(request_info)
    mdms_client = MDMSClient(mdms_url)

    boundary_code_list: List[str] = [code.strip() for code in boundary_codes.split(",") if code.strip()]

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_filename = f"facility_selection_template_{timestamp}.xlsx"
    output_file_path = create_temp_file(suffix=".xlsx")

    boundary_facilities = []
    project_facilities = []

    facility_client = None

    try:
        facility_selection_schema = mdms_client.fetch_facility_selection_schema(request_info=request_info)
    except Exception as e:
        logger.error(f"Error fetching data from external services: {e}")
        cleanup_temp_file(output_file_path)
        raise HTTPException(status_code=502, detail=f"External service error: {str(e)}")

    if facility_service_url:
        facility_client = FacilityServiceClient(facility_service_url)
        for boundary_code in boundary_code_list:
            try:
                results = facility_client.search_facility(tenant_id=LIVELIHOOD_TENANT_ID, boundary_code=boundary_code)
                boundary_facilities.extend(results.get('facilities', []))
            except Exception as e:
                logger.error(f"Error fetching boundary facilities for boundary code {boundary_code}: {e}", exc_info=True)

    if project_service_url and parent_project_id:
        project_client = ProjectServiceClient(project_service_url)
        try:
            pf_response = project_client.search_project_facility(
                request_info=request_info,
                project_id=parent_project_id
            )
            raw_project_facilities = pf_response.get("ProjectFacilities", [])
            if raw_project_facilities and facility_client:
                for pf in raw_project_facilities:
                    facility_id = pf.get("facilityId")
                    if facility_id and any(f.get('facility_id') == facility_id for f in boundary_facilities):
                        try:
                            facility_data = facility_client.search_facility(tenant_id=LIVELIHOOD_TENANT_ID, facility_id=facility_id)
                            if facility_data:
                                project_facilities.extend(facility_data.get('facilities', []))
                        except Exception as e:
                            logger.error(f"Error fetching facility {facility_id}: {e}", exc_info=True)
        except Exception as e:
            logger.error(f"Error fetching project facilities for project {parent_project_id}: {e}", exc_info=True)

    # Intersect by facility_id
    if parent_project_id:
        intersected_facilities = project_facilities
    else:
        intersected_facilities = boundary_facilities

    try:
        facility_service.generate_selection_template_file(
            output_path=output_file_path,
            facility_selection_schema=facility_selection_schema,
            facility_data=intersected_facilities
        )
        logger.info(f"Successfully created facility ingestion template at {output_file_path}")
    except Exception as e:
        logger.error(f"Error generating template file: {e}")
        cleanup_temp_file(output_file_path)
        raise HTTPException(status_code=500, detail=f"Template generation error: {str(e)}")

    return FileResponse(
        path=output_file_path,
        filename=output_filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@router.post('/facilityQRGeneration', summary='Generate QR for facility',
             response_description="Returns zip with QR codes")
async def get_facility_QR_for_autologin(
        request_info: str = Form(default="")
):
    request_info_obj = request_info_from_json(request_info)

    try:
        mdms_client = MDMSClient(mdms_url=mdms_url)
        mdms_content = mdms_client.get_tenant_mapping(request_info_obj, ["as", "gj", "ml", "mn", "mz", "nl", "or", "pg", "sk"])


        base_url = "https://saura-emitra-uat.selcofoundation.org"
        password = "Health@2026"

        temp_dir = tempfile.mkdtemp()

        for tenant_id, health_facility_data in mdms_content.items():

            state = health_facility_data.get("address")
            district = health_facility_data.get("city", {}).get("districtName")
            block = health_facility_data.get("city", {}).get("blockCode")
            facility_name = health_facility_data.get("name")

            if not all([state, district, block, facility_name]):
                logger.warning(f"Skipping tenant {tenant_id}: Missing state/district/block/facility_name")
                continue

            qr_folder = os.path.join(temp_dir, state, district, block, facility_name)
            os.makedirs(qr_folder, exist_ok=True)

            conn = psycopg2.connect(**DB_CONFIG)
            with conn.cursor() as cursor:
                sql = "SELECT code FROM eg_hrms_employee WHERE tenantid = %s"
                cursor.execute(sql, (health_facility_data["code"],))
                rows = cursor.fetchall()

            if not rows:
                logger.warning(f"Skipping tenant {tenant_id} ({facility_name}): No HRMS employee code found for tenantid {health_facility_data['code']}")
                continue

            username = rows[0][0]


            if state=='Karnataka':
                url_state_name = 'digit-ui'
            else:
                url_state_name = state.lower()

            login_url = f"{base_url}/{url_state_name}/employee/user/login?tenantid={tenant_id}&username={username}&passwd={password}"

            qr = qrcode.make(login_url).convert("RGB")

            # Create a new image (taller) to hold QR and text
            width, height = qr.size
            font_size = 30
            padding = 10

            try:
                font = ImageFont.truetype("DejaVuSans-Bold.ttf", font_size)
            except:
                font = ImageFont.load_default()

            # Create a new image with extra space for text
            new_height = height + font_size + 2 * padding
            combined = Image.new("RGB", (width, new_height), "white")
            combined.paste(qr, (0, 0))

            # Draw the facility name
            draw = ImageDraw.Draw(combined)
            text = facility_name
            bbox = draw.textbbox((0, 0), text, font=font)
            text_width = bbox[2] - bbox[0]
            text_position = ((width - text_width) // 2, height + padding)
            draw.text(text_position, text, font=font, fill="black")

            # Save the combined image
            qr_filename = f"{username}.png"
            img_path = os.path.join(qr_folder, qr_filename)
            combined.save(img_path)

        # Create ZIP
        zip_path = os.path.join(tempfile.gettempdir(), "facility_qr_codes.zip")
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zipf:
            for root, _, files in os.walk(temp_dir):
                for file in files:
                    abs_file = os.path.join(root, file)
                    rel_path = os.path.relpath(abs_file, temp_dir)
                    zipf.write(abs_file, arcname=rel_path)

        shutil.rmtree(temp_dir)

        return FileResponse(
            path=zip_path,
            filename="facility_qr_codes.zip",
            media_type="application/zip"
        )
    except Exception as e:
        return {"status": "error", "message": str(e)}


@router.post(
    '/livelihoodFacilityQRGeneration',
    summary='Generate Livelihood facility QR codes for OTP login',
    response_description='ZIP of PNG QR codes (one per facility)',
)
async def get_livelihood_facility_qr_for_otp_login(
        payload: dict = Body(
            ...,
            description=(
                "RequestInfo + optional filters. "
                "QR encodes facilityId for OTP login "
                "(end user is facility-level; assets are linked to the facility)."
            ),
        ),
):
    """
    Generate printable QR codes for Livelihood facility end-user OTP login.

    Each QR encodes:
      {baseUrl}/employee/user/login?tenantId=livelihood&facilityId={facilityId}
    """
    request_info = request_info_from_json(payload.get("RequestInfo", {}))
    tenant_id = payload.get("tenantId") or LIVELIHOOD_TENANT_ID
    base_url = (payload.get("baseUrl") or os.getenv(
        "LIVELIHOOD_UI_BASE_URL",
        "https://setu4livelihood-dev.selcofoundation.org/livelihood-ui/",
    )).rstrip("/")
    boundary_code = payload.get("boundaryCode")
    facility_ids = payload.get("facilityIds") or []
    if isinstance(facility_ids, str):
        facility_ids = [facility_ids]

    if not facility_service_url:
        raise HTTPException(status_code=500, detail="FACILITY_SERVICE_URL is not configured")

    facility_client = FacilityServiceClient(facility_service_url)
    temp_dir = tempfile.mkdtemp()
    generated = 0
    skipped = 0

    try:
        facilities = []
        if facility_ids:
            for fid in facility_ids:
                if not fid:
                    continue
                result = facility_client.search_facility(tenant_id=tenant_id, facility_id=str(fid).strip())
                facilities.extend(result.get("facilities") or [])
        else:
            result = facility_client.search_facility(tenant_id=tenant_id, boundary_code=boundary_code)
            facilities = result.get("facilities") or []

        if not facilities:
            raise HTTPException(status_code=404, detail="No facilities found for the given filters")

        # Deduplicate by facility_id
        seen = set()
        unique_facilities = []
        for facility in facilities:
            fid = facility.get("facility_id") or facility.get("facilityId")
            if not fid or fid in seen:
                continue
            seen.add(fid)
            unique_facilities.append(facility)

        for facility in unique_facilities:
            facility_id = facility.get("facility_id") or facility.get("facilityId")
            facility_name = (
                facility.get("facility_name")
                or facility.get("facilityName")
                or facility_id
            )
            phone = facility.get("facility_poc_phone") or facility.get("facilityPocPhone") or ""
            boundary = facility.get("boundaryCode") or facility.get("boundary_code") or "unknown"

            # Folder: boundary / facility name (sanitized)
            safe_boundary = _sanitize_path_segment(boundary)
            safe_name = _sanitize_path_segment(facility_name)
            qr_folder = os.path.join(temp_dir, safe_boundary, safe_name)
            os.makedirs(qr_folder, exist_ok=True)

            login_url = (
                f"{base_url}/employee/user/login"
                f"?tenantId={tenant_id}&facilityId={facility_id}"
            )

            qr = qrcode.make(login_url).convert("RGB")
            width, height = qr.size
            font_size = 28
            padding = 10
            try:
                font = ImageFont.truetype("DejaVuSans-Bold.ttf", font_size)
            except Exception:
                font = ImageFont.load_default()

            label = facility_name
            if phone:
                label = f"{facility_name} | {phone}"

            bbox = ImageDraw.Draw(Image.new("RGB", (1, 1))).textbbox((0, 0), label, font=font)
            text_width = bbox[2] - bbox[0]
            text_height = bbox[3] - bbox[1]
            canvas_width = max(width, text_width + 2 * padding)
            new_height = height + text_height + 2 * padding
            combined = Image.new("RGB", (canvas_width, new_height), "white")
            combined.paste(qr, ((canvas_width - width) // 2, 0))
            draw = ImageDraw.Draw(combined)
            draw.text(
                ((canvas_width - text_width) // 2, height + padding),
                label,
                font=font,
                fill="black",
            )

            qr_filename = f"{_sanitize_path_segment(facility_id)}.png"
            combined.save(os.path.join(qr_folder, qr_filename))
            generated += 1

            # Also write a small sidecar with the encoded URL for ops/debug
            with open(os.path.join(qr_folder, f"{_sanitize_path_segment(facility_id)}.url.txt"), "w", encoding="utf-8") as f:
                f.write(login_url)

        if generated == 0:
            raise HTTPException(status_code=404, detail="No QR codes generated")

        zip_path = os.path.join(tempfile.gettempdir(), f"livelihood_facility_qr_codes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip")
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zipf:
            for root, _, files in os.walk(temp_dir):
                for file in files:
                    abs_file = os.path.join(root, file)
                    rel_path = os.path.relpath(abs_file, temp_dir)
                    zipf.write(abs_file, arcname=rel_path)

        logger.info(
            "Livelihood facility QR generation done | tenantId=%s generated=%s skipped=%s requestMsgId=%s",
            tenant_id,
            generated,
            skipped,
            getattr(request_info, "msg_id", None) if request_info else None,
        )

        return FileResponse(
            path=zip_path,
            filename="livelihood_facility_qr_codes.zip",
            media_type="application/zip",
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error("Livelihood facility QR generation failed: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)


def _sanitize_path_segment(value: str) -> str:
    if not value:
        return "unknown"
    cleaned = "".join(c if c.isalnum() or c in ("-", "_", ".") else "_" for c in str(value).strip())
    return cleaned[:120] or "unknown"


@router.post('/amcConfigurationTemplate',
            summary='Generate AMC configuration ingestion template',
            response_description="Returns Excel template with facility asset metadata for AMC configurations")
async def get_amc_configuration_template(
        background_tasks: BackgroundTasks,
        payload: dict = Body(..., description="Payload containing RequestInfo, boundary_data")
):
    request_info = request_info_from_json(payload.get("RequestInfo", {}))

    boundary_data = payload.get("boundary_data", {})
    project_id = payload.get("project_id")  

    if not boundary_data:
        raise HTTPException(status_code=400, detail="boundary_data is required")

    if not facility_service_url:
        raise HTTPException(status_code=500, detail="Facility service is not configured")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_filename = f"amc_configuration_template_{timestamp}.xlsx"
    output_file_path = create_temp_file(suffix=".xlsx")

    sheet_name = "amc-configurations"
    columns = [
        "Facility Id",
        "NIN/HFR ID",
        "BoundaryCode",
        "Health Facility Name",
        "AMC-Frequency",
        "AMC-Duration"
    ]

    try:
        # Flatten boundaries from boundary_data
        boundary_list: List[Boundary] = flatten_boundaries(boundary_data)

        # Get all project-linked facilities if project_id is provided
        project_linked_facility_ids = set()
        if project_id and project_service_url:
            try:
                project_client = ProjectServiceClient(project_service_url)
                project_facilities_response = project_client.search_project_facility(request_info, project_id)
                project_facilities = project_facilities_response.get("ProjectFacilities", [])
                project_linked_facility_ids = {pf.get("facilityId") for pf in project_facilities if pf.get("facilityId")}
                logger.info(f"Found {len(project_linked_facility_ids)} facilities currently linked to project {project_id}")
            except Exception as e:
                logger.error(f"Error fetching project facilities: {e}")
                # Continue without project facility filtering if there's an error

        # Fetch facilities by boundary codes, then optionally filter by project-linked facility ids.
        # This avoids an expensive boundary x facility nested API call pattern.
        facility_client = FacilityServiceClient(facility_service_url)
        all_facilities = []
        seen_facility_keys = set()

        def get_facility_dedup_key(facility):
            facility_id = facility.get("facility_id")
            if facility_id:
                return facility_id
            return (
                facility.get("nin_id"),
                facility.get("hfr_id"),
                facility.get("facility_name"),
                facility.get("boundary_code") or facility.get("boundaryCode")
            )

        def add_unique_facilities(facilities):
            for facility in facilities:
                dedup_key = get_facility_dedup_key(facility)
                if dedup_key in seen_facility_keys:
                    continue
                seen_facility_keys.add(dedup_key)
                all_facilities.append(facility)

        boundary_codes = [b.code for b in boundary_list if b.code]
        if boundary_codes:
            try:
                bulk_result = facility_client.bulk_search_facility_with_boundary(
                    request_info=request_info,
                    tenant_ids=[LIVELIHOOD_TENANT_ID],
                    boundary_codes=boundary_codes,
                    limit=max(len(boundary_codes) * 50, 50),
                    send_non_paginated_response=True,
                )
                facilities = bulk_result.get("facilities", []) or []
                if project_id and project_linked_facility_ids:
                    facilities = [
                        f for f in facilities
                        if f.get("facility_id") in project_linked_facility_ids
                    ]
                add_unique_facilities(facilities)
            except Exception as e:
                logger.error(f"Error fetching boundary facilities in bulk for AMC template: {e}", exc_info=True)

        logger.info(
            f"Total facilities in AMC template: {len(all_facilities)} "
            f"(raw: {len(all_facilities)}, project_id: {project_id}, boundaries: {len(boundary_list)})"
        )

        # Initialize AMC scheduler client and prefetch existing configs for this project once
        amc_client = None
        existing_amc_by_facility = {}
        if amc_scheduler_service_url and project_id:
            amc_client = AMCSchedulerServiceClient(amc_scheduler_service_url)
            try:
                all_existing_configs_resp = amc_client.search_amc_configurations(
                    request_info,
                    project_id=project_id
                )
                all_existing_configs = all_existing_configs_resp.get("AmcConfigurations", [])
                for config in all_existing_configs:
                    facility_id = config.get("facilityId")
                    if facility_id and facility_id not in existing_amc_by_facility:
                        existing_amc_by_facility[facility_id] = config
                logger.info(
                    f"Fetched {len(existing_amc_by_facility)} existing AMC configurations for project {project_id}"
                )
            except Exception as e:
                logger.warning(f"Error fetching existing AMC configurations for project {project_id}: {e}")

        # Create rows for AMC configuration template - one row per facility
        # Asset types ["INVERTER","PANEL","BATTERY"] will be used as default for each configuration during processing
        rows = []
        rows_with_existing_amc = []  # Track row indices that have existing AMC configurations

        def convert_frequency_to_display(frequency_months):
            """Convert frequency in months to display format"""
            if frequency_months == 6:
                return "Every 6 Months"
            elif frequency_months == 12:
                return "Every 1 Year"
            return ""

        def convert_duration_to_display(duration_months):
            """Convert duration in months to display format"""
            if duration_months == 12:
                return "1 Year"
            elif duration_months == 36:
                return "3 Years"
            elif duration_months == 60:
                return "5 Years"
            return ""

        for idx, facility in enumerate(all_facilities):
            # facility_details = facility.get("facility_details", {}) or {}
            nin_id = facility.get("nin_id", "")
            hfr_id = facility.get("hfr_id", "")
            # Use NIN ID if available, otherwise HFR ID, otherwise empty
            nin_hfr_id = nin_id if nin_id else (hfr_id if hfr_id else "")
            facility_name = facility.get("facility_name", "")
            boundary_code = facility.get("boundary_code") or facility.get("boundaryCode", "")
            facility_id = facility.get("facility_id", "")

            # Initialize row with empty values
            frequency_value = ""
            duration_value = ""

            # Read existing AMC configuration from prefetched map
            if amc_client and project_id and facility_id:
                try:
                    existing_configs = amc_client.search_amc_configurations(
                        request_info,
                        facility_id=facility_id,
                        project_id=project_id
                    )
                    # Check if any configurations exist
                    configs = existing_configs.get("AmcConfigurations", [])
                    if configs:
                        # Use the first configuration found
                        existing_config = configs[0]
                        frequency_months = existing_config.get("frequency")
                        duration_months = existing_config.get("duration")

                        frequency_value = convert_frequency_to_display(frequency_months)
                        duration_value = convert_duration_to_display(duration_months)
                        rows_with_existing_amc.append(idx)
                        logger.info(f"Found existing AMC config for facility {facility_id}: frequency={frequency_value}, duration={duration_value}")
                except Exception as e:
                    logger.warning(f"Error checking existing AMC config for facility {facility_id}: {e}")
                    # Continue without existing config data

            # Create one row per facility (asset types are handled internally during processing)
            rows.append({
                "Facility Id": facility_id,
                "NIN/HFR ID": nin_hfr_id,
                "BoundaryCode": boundary_code,
                "Health Facility Name": facility_name,
                "AMC-Frequency": frequency_value,
                "AMC-Duration": duration_value
            })

        # Create DataFrame and write to Excel
        df = pd.DataFrame(rows, columns=columns)
        with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name=sheet_name)

            # Add Boundary Data Sheet
            boundary_records = []
            for boundary in boundary_list:
                boundary_records.append({
                    "Country": boundary.get("country", ""),
                    "State": boundary.get("state", ""),
                    "District": boundary.get("district", ""),
                    "Block": boundary.get("block", ""),
                    "BoundaryCode": boundary.get("code", "")
                })
            if not boundary_records:
                boundary_records.append({
                    "Country": "",
                    "State": "",
                    "District": "",
                    "Block": "",
                    "BoundaryCode": ""
                })

            df_boundary = pd.DataFrame(boundary_records)
            df_boundary.to_excel(writer, index=False, sheet_name="BoundaryCodes")

        # Lock the boundary sheet
        lock_excel_columns(
            file_path=output_file_path,
            sheet_name="BoundaryCodes",
            column_headers_to_unlock=[]
        )

        # Add dropdowns for amc-frequency and amc-duration
        dropdowns_map = {
            "AMC-Frequency": ["Every 6 Months", "Every 1 Year"],
            "AMC-Duration": ["1 Year", "3 Years", "5 Years"],
        }

        allow_blank_map = {
            "AMC-Frequency": True,
            "AMC-Duration": True
        }

        add_dropdowns_to_excel(
            file_path=output_file_path,
            sheet_name=sheet_name,
            dropdowns=dropdowns_map,
            allow_blank_map=allow_blank_map,
            max_extra_rows=500
        )

        # Lock prefilled rows except editable columns
        # BoundaryCode is pre-filled and should be locked
        lock_prefilled_rows_in_excel(
            file_path=output_file_path,
            sheet_name=sheet_name,
            editable_columns=["AMC-Frequency", "AMC-Duration"],
            total_rows=len(rows),
            total_columns=len(columns),
            always_locked_columns=["BoundaryCode"],
            extra_append_rows=500
        )

        # Lock AMC-Frequency and AMC-Duration for rows with existing AMC configurations
        if rows_with_existing_amc:
            wb = load_workbook(output_file_path)
            ws = wb[sheet_name]
            grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

            header_row = [cell.value for cell in ws[1]]
            frequency_col_idx = None
            duration_col_idx = None

            for idx, header in enumerate(header_row, start=1):
                if header == "AMC-Frequency":
                    frequency_col_idx = idx
                elif header == "AMC-Duration":
                    duration_col_idx = idx

            # Lock cells for rows with existing AMC configurations
            # Note: rows_with_existing_amc contains 0-based indices, Excel rows are 1-based
            # Header is row 1, so data starts at row 2
            for row_idx_0based in rows_with_existing_amc:
                excel_row = row_idx_0based + 2  # +2 because header is row 1, and 0-based to 1-based

                if frequency_col_idx:
                    cell = ws.cell(row=excel_row, column=frequency_col_idx)
                    cell.protection = Protection(locked=True)
                    cell.fill = grey_fill

                if duration_col_idx:
                    cell = ws.cell(row=excel_row, column=duration_col_idx)
                    cell.protection = Protection(locked=True)
                    cell.fill = grey_fill

            # Re-enable sheet protection
            ws.protection.sheet = True
            ws.protection.enable()
            wb.save(output_file_path)
            logger.info(f"Locked AMC fields for {len(rows_with_existing_amc)} rows with existing AMC configurations")

        # Autofit columns
        autofit_columns(
            file_path=output_file_path,
            sheet_name=sheet_name
        )
        autofit_columns(
            file_path=output_file_path,
            sheet_name="BoundaryCodes"
        )

        background_tasks.add_task(cleanup_temp_file, output_file_path)
        return FileResponse(
            path=output_file_path,
            filename=output_filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except HTTPException:
        cleanup_temp_file(output_file_path)
        raise
    except Exception as e:
        logger.error(f"Unhandled error in get_amc_configuration_template: {e}")
        cleanup_temp_file(output_file_path)
        raise HTTPException(status_code=500, detail=f"An unexpected error occurred: {str(e)}")
