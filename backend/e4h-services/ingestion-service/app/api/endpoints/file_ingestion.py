import io
import json
import os
import tempfile
import time
from datetime import datetime, timedelta
import uuid
from typing import Optional, Dict, List, Set

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Protection, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

from app.utils.amc_scheduler_service_client import AMCSchedulerServiceClient
from app.utils.excel_utils import (
    FACILITY_IDENTIFIER_COLUMNS,
    autofit_columns,
    normalize_excel_integer_columns,
    prepare_dataframe_for_excel_export,
)
from app.utils.facility_validator import (
    project_facility_validation,
    facility_validation,
    collect_hfr_nin_errors_for_row,
    collect_anganwadi_poc_username_errors_for_row,
    validate_installation_scope_solutions,
    find_site_id_column,
    SITE_ID_COLUMNS,
)
from app.utils.state_sunshine_hours_repository import fetch_state_sunshine_hours
from app.utils.field_plan_locks import build_project_lock_map, solution_codes_by_name, \
    solution_names_by_code, PLAN_STATUS_PUBLISHED
from app.utils.icc_template_parser import annotate_worksheet, first_data_sheet, parse_worksheet, \
    to_sections, validate_line_items
from fastapi import APIRouter, File, Form, UploadFile, HTTPException, BackgroundTasks, Depends
from fastapi.responses import FileResponse
import psycopg2
from starlette.responses import JSONResponse, StreamingResponse
import requests

from app.core.logging import AppLogger
from app.core.tenant import LIVELIHOOD_TENANT_ID
from app.decorators.rbac_validator import get_authorized_request_info
from app.ingest.excel_data_writer import ExcelDataWriter
from app.processor.factory.boundary_data_processor_factory import BoundaryDataProcessorFactory
from app.processor.factory.vendor_data_processor_factory import VendorDataProcessorFactory
from app.schemas.request_info import RequestInfo
from app.producer.producer import Producer
from app.utils.convertor import request_info_from_json, create_vendor_request, create_facility_payload, \
    get_project_creation_payload, check_role_mismatch_for_user_type, get_user_creation_payload_staff, \
    get_user_creation_payload_supervisors, \
    get_staff_creation_payload, create_project_payload, get_installation_spoc_creation_payload, \
    get_staff_search_payload, create_update_payload, get_incident_request_info, \
    resolve_boundary_codes_for_dataframe, create_asset_payload
from app.utils.asset_validator import asset_validation
from app.utils.asset_service_client import AssetServiceClient
from app.utils.boundary_service_client import BoundaryServiceClient
from app.utils.facility_service_client import FacilityServiceClient
from app.utils.fieldplan_activity_service_client import FieldPlanActivityServiceClient
from app.utils.fieldplan_service_client import FieldPlanServiceClient
from app.utils.file_utils import cleanup_temp_file, create_temp_file
from app.utils.im_service_client import IMServiceClient
from app.utils.mdms_client import MDMSClient
from app.utils.organization_service_client import OrganizationServiceClient
from app.utils.project_service_client import ProjectServiceClient
from app.utils.hrms_service_client import HRMSServiceClient
from app.utils.vendor_registry_client import VendorRegistryClient

router = APIRouter()
logger = AppLogger().get_logger()

from dotenv import load_dotenv
from collections import defaultdict


async def _save_upload_to_temp_file(upload_file: UploadFile, suffix: str = ".xlsx", chunk_size: int = 1024 * 1024):
    """
    Persist an UploadFile to disk in chunks to avoid loading
    large uploads entirely into memory.
    """
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
    total_bytes = 0
    try:
        while True:
            chunk = await upload_file.read(chunk_size)
            if not chunk:
                break
            temp_file.write(chunk)
            total_bytes += len(chunk)
    finally:
        temp_file.close()
    return temp_file, total_bytes


load_dotenv()
mdms_url = os.getenv("MDMS_URL")
org_service_url = os.getenv("VENDOR_SERVICE_URL")
project_service_url = os.getenv("PROJECT_SERVICE_URL")
fieldPlan_service_url = os.getenv("FIELDPLAN_SERVICE_URL")
fieldPlan_activity_service_url = os.getenv("FIELDPLAN_ACTIVITY_SERVICE_URL")
facility_service_url = os.getenv("FACILITY_SERVICE_URL")
asset_service_url = os.getenv("ASSET_SERVICE_URL")
hrms_service_url = os.getenv("HRMS_SERVICE_URL")
im_services_url = os.getenv("IM_SERVICES_URL")
amc_scheduler_service_url = os.getenv("AMC_SCHEDULER_SERVICE_URL")
localization_service_url = os.getenv("LOCALIZATION_SERVICE_URL")
boundary_service_url = os.getenv("BOUNDARY_SERVICE_URL")
DEFAULT_AMC_ASSET_TYPES = ["INVERTER", "PANEL", "BATTERY"]
BULK_INGEST_CHUNK_SIZE = 200

# The bulk link endpoint answers 202 as soon as the message is queued; the row is written
# two async hops later (bulk topic -> consumer -> save-fieldplan-facility-topic ->
# egov-persister). These bound how long we wait to confirm it before saying so. ~5s.
SCOPE_LINK_CONFIRM_ATTEMPTS = 10
SCOPE_LINK_CONFIRM_INTERVAL_SECONDS = 0.5
AMC_CONFIGURATION_BULK_CHUNK_SIZE = 400
ENVIRONMENT = os.getenv("ENVIRONMENT", "uat").lower()
base_path = os.path.dirname(os.path.abspath(__file__))
config_path = os.path.abspath(os.path.join(base_path, "..", "..", "config"))

with open(os.path.join(config_path, "tenant_creator_mapping.json"), 'r') as f:
    TENANT_CREATOR_MAPPING = json.load(f).get(ENVIRONMENT, {})

with open(os.path.join(config_path, "user_profiles.json"), 'r') as f:
    USER_PROFILE = json.load(f).get(ENVIRONMENT, {})

DB_CONFIG = {
    "host": os.getenv("DB_HOST"),
    "port": int(os.getenv("DB_PORT", 5432)),
    "database": os.getenv("DB_NAME"),
    "user": os.getenv("DB_USER"),
    "password": os.getenv("DB_PASSWORD")
}

@router.post('/vendors',
             summary='Upload and process vendor Excel file with multiple sheets',
             response_description="Returns processed Excel file with validation results")
async def upload_vendors_excel_sheet(
        vendor_file: UploadFile = File(description="Excel file containing vendor data and boundary codes"),
        vendor_sheet_name: str = Form(default="Vendor Input", description="Name of the sheet containing vendor data"),
        boundary_sheet_name: str = Form(default="Boundary Code",
                                        description="Name of the sheet containing boundary codes"),
        request_info: str = Form(default="")
):
    logger.trace("Starting vendor Excel file upload and processing")
    input_temp_file = None
    output_temp_file = None
    request_info = request_info_from_json(request_info)
    #get_authorized_request_info(request_info)
    logger.info(f"Processing vendor file: vendor_sheet={vendor_sheet_name}, boundary_sheet={boundary_sheet_name}")

    try:
        logger.debug("Creating temporary files for vendor processing")
        input_temp_file, uploaded_size = await _save_upload_to_temp_file(vendor_file, suffix=".xlsx")
        vendor_file_path = input_temp_file.name
        logger.debug(f"Saved uploaded file to: {vendor_file_path}, size: {uploaded_size} bytes")

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"vendor_validation_results_{timestamp}.xlsx"
        output_temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        output_temp_file.close()
        output_file_path = output_temp_file.name

        logger.info("Creating vendor data processor")
        processor = VendorDataProcessorFactory.create_processor(
            file_path=vendor_file_path,
            vendor_sheet=vendor_sheet_name,
            boundary_sheet=boundary_sheet_name,
            mdms_url=mdms_url,
            request_info=request_info
        )
        logger.info("Processing vendor data")
        tuple_vendors = processor.process_data()
        vendors = tuple_vendors[0]
        vendor_df = tuple_vendors[1]
        logger.info(f"Vendor processing completed: {len(vendors)} valid vendors, {len(vendor_df)} total rows")

        if org_service_url and vendors:
            logger.info(f"Creating {len(vendors)} vendors in organization service")
            org_client = OrganizationServiceClient(org_service_url)

            success_count = 0
            for index, vendor in enumerate(vendors):
                logger.trace(f"Creating vendor {index + 1}/{len(vendors)}: {vendor.vendor_name}")
                vendor_payload = create_vendor_request(request_info, vendor)

                try:
                    org_data = org_client.create_vendor(vendor_payload)
                    if org_data and org_data.get("organisations"):
                        vendor_df.at[index, "status"] = "success"
                        vendor_df.at[index, "error"] = None
                        vendor_id = org_data["organisations"][0].get("id")
                        vendor_df.at[index, "vendor_id"] = vendor_id
                        success_count += 1
                        logger.debug(f"Vendor created successfully: {vendor.vendor_name}, id={vendor_id}")
                    else:
                        logger.warning(f"Failed to create vendor: {vendor.vendor_name} - no organization data returned")
                except Exception as e:
                    logger.error(f"Error creating vendor {vendor.vendor_name} in org service: {e}", exc_info=True)
            logger.info(f"Vendor creation completed: {success_count}/{len(vendors)} successful")

        logger.info("Writing processed vendor data to Excel file")
        with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
            vendor_df.to_excel(writer, sheet_name="Vendor Output", index=False)
            boundary_df = pd.read_excel(vendor_file_path, sheet_name=boundary_sheet_name)
            boundary_df.to_excel(writer, sheet_name=boundary_sheet_name, index=False)
        logger.info(f"Vendor processing completed successfully: {output_filename}")

        return FileResponse(
            path=output_file_path,
            filename=output_filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        logger.error(f"Error processing vendor data: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to process vendor data: {str(e)}")

    finally:
        if input_temp_file and os.path.exists(input_temp_file.name):
            os.unlink(input_temp_file.name)


@router.post('/boundaries',
             summary='Upload and process boundary Excel file',
             response_description="Returns processed Excel file with validation results")
async def upload_boundaries_excel_sheet(
        boundary_file: UploadFile = File(description="Excel file containing boundary data"),
        boundary_sheet_name: str = Form(default="Boundary Data",
                                        description="Name of the sheet containing boundary data"),
        request_info: str = Form(default="")
):
    logger.trace("Starting boundary Excel file upload and processing")
    input_temp_file = None
    output_temp_file = None
    request_info = request_info_from_json(request_info)
    logger.info(f"Processing boundary file: boundary_sheet={boundary_sheet_name}")

    try:
        logger.debug("Creating temporary files for boundary processing")
        input_temp_file, uploaded_size = await _save_upload_to_temp_file(boundary_file, suffix=".xlsx")
        boundary_file_path = input_temp_file.name
        logger.debug(f"Saved uploaded file to: {boundary_file_path}, size: {uploaded_size} bytes")

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"boundary_validation_results_{timestamp}.xlsx"
        output_temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        output_temp_file.close()
        output_file_path = output_temp_file.name

        with open(boundary_file_path, 'rb') as src, open(output_file_path, 'wb') as dst:
            dst.write(src.read())

        logger.info("Creating boundary data processor")
        processor = BoundaryDataProcessorFactory.create_processor(
            file_path=output_file_path,
            boundary_sheet=boundary_sheet_name,
            mdms_url=mdms_url,
            request_info=request_info
        )
        logger.info("Processing boundary data")
        boundary_df = processor.process_data()
        logger.info(f"Boundary processing completed: {len(boundary_df)} boundaries processed")

        writer = ExcelDataWriter(output_file_path, output_sheet="Boundary Data")
        writer.write_data(boundary_df)

        error_count = int(
            boundary_df["status"].astype(str).str.strip().str.lower().eq("fail").sum()
        )

        response = FileResponse(
            path=output_file_path,
            filename=output_filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response.headers["X-Error-Count"] = str(error_count)
        return response

    except Exception as e:
        logger.error(f"Error processing boundary data: {e}")
        raise HTTPException(
            status_code=500,
            detail="Failed to process boundary data"
        ) from e


    finally:
        if input_temp_file and os.path.exists(input_temp_file.name):
            os.unlink(input_temp_file.name)


@router.post('/addFacilitiesValidateData',
             summary='Validate add bulk facility Excel file before processing',
             response_description='Returns validation report Excel with PASSED/FAILED rows')
async def validate_facilities_excel_sheet(
        background_tasks: BackgroundTasks,
        facility_file: UploadFile = File(..., description="Excel file containing facility data"),
        facility_sheet_name: str = Form(default="EndUserIngestionTemplate",
                                        description="Name of the sheet containing facility data"),
        boundary_sheet_name: str = Form(default="BlockBoundaryCodes",
                                        description="Name of the sheet containing boundary data"),
        request_info: str = Form(default="")
):
    temp_input_file = None
    request_info_obj = request_info_from_json(request_info)
    mdms_client = MDMSClient(mdms_url)
    facility_client = FacilityServiceClient(facility_service_url)

    try:
        # Save uploaded Excel to a temp file
        temp_input_file, _ = await _save_upload_to_temp_file(facility_file, suffix=".xlsx")

        # Load workbook to preserve everything
        wb = load_workbook(temp_input_file.name)

        # ----------------- Read Boundary Sheet ----------------- #
        if boundary_sheet_name not in wb.sheetnames:
            raise HTTPException(status_code=400, detail=f"Boundary sheet '{boundary_sheet_name}' not found")

        boundary_data_df = pd.read_excel(temp_input_file.name, sheet_name=boundary_sheet_name)

        # ----------------- Read Facility Sheet ----------------- #
        if facility_sheet_name not in wb.sheetnames:
            raise HTTPException(status_code=400, detail=f"Facility sheet '{facility_sheet_name}' not found")

        df = pd.read_excel(temp_input_file.name, sheet_name=facility_sheet_name)
        df.columns = [str(c).strip() for c in df.columns]
        df = df.loc[:, ~df.columns.str.startswith('Unnamed')]
        df = normalize_excel_integer_columns(df, force_columns=FACILITY_IDENTIFIER_COLUMNS)

        # ----------------- Read Facility Column ----------------- #
        facility_id_column = 'End user Id'
        if facility_id_column not in df.columns:
            raise HTTPException(status_code=400, detail=f"'End user Id' column in '{facility_sheet_name}' not found")

        # Ensure status/error columns exist
        if 'status' not in df.columns:
            df['status'] = ''
        if 'error' not in df.columns:
            df['error'] = ''

        # Resolve boundary codes from State/District/Block via localization
        df = resolve_boundary_codes_for_dataframe(
            df, localization_service_url,
            boundary_code_column='Boundary Code (Mandatory)',
            logger=logger,
        )

        # ----------------- Verify resolved codes via Boundary Service ----------------- #
        resolved_codes = df.loc[
            df['Boundary Code (Mandatory)'].astype(str).str.strip() != '',
            'Boundary Code (Mandatory)'
        ].unique().tolist()

        if resolved_codes and boundary_service_url:
            boundary_client = BoundaryServiceClient(boundary_service_url)
            existing_codes = set()
            chunk_size = 50
            for i in range(0, len(resolved_codes), chunk_size):
                chunk = resolved_codes[i:i + chunk_size]
                try:
                    response_data = boundary_client.search_boundaries(
                        request_info=request_info_obj,
                        tenant_id=LIVELIHOOD_TENANT_ID,
                        codes=chunk,
                    )
                    if response_data and "Boundary" in response_data:
                        for b in response_data["Boundary"]:
                            existing_codes.add(b["code"])
                except Exception as e:
                    logger.error(f"Error verifying boundary codes: {e}", exc_info=True)

            missing_codes = set(resolved_codes) - existing_codes
            if missing_codes:
                logger.warning(f"Boundary codes not found in boundary service: {missing_codes}")
                for index, row in df.iterrows():
                    code = str(row.get('Boundary Code (Mandatory)', '') or '').strip()
                    if code in missing_codes:
                        state_val = str(row.get('State (Mandatory)', '') or '').strip()
                        district_val = str(row.get('District (Mandatory)', '') or '').strip()
                        block_val = str(row.get('Block (Mandatory)', '') or '').strip()
                        df.at[index, 'Boundary Code (Mandatory)'] = ''
                        df.at[index, 'status'] = 'FAILED'
                        df.at[index, 'error'] = f"Boundary code for State '{state_val}' District '{district_val}' Block '{block_val}' not found"

        # ----------------- Run Validation ----------------- #
        validation_errors = facility_validation(
            df,
            mdms_client,
            request_info_obj,
            facility_client,
            boundary_data_df,
            'data-ingestion.FacilityIngestionSchemaWithoutBoundaryCode'
        )

        # Mark rows based on validation results, preserving earlier boundary errors
        error_count = 0
        for i, errs in enumerate(validation_errors):
            existing_status = str(df.at[i, 'status']).strip().upper() if pd.notna(df.at[i, 'status']) else ''
            existing_error = str(df.at[i, 'error']).strip() if pd.notna(df.at[i, 'error']) else ''

            if existing_status == 'FAILED':
                if errs:
                    df.at[i, 'error'] = existing_error + "; " + "; ".join(dict.fromkeys(errs))
                error_count += 1
            elif errs:
                df.at[i, 'status'] = 'FAILED'
                df.at[i, 'error'] = "; ".join(dict.fromkeys(errs))
                error_count += 1
            else:
                df.at[i, 'status'] = 'PASSED'
                df.at[i, 'error'] = ''

        # ----------------- Update Facility Sheet In-Place ----------------- #
        ws = wb[facility_sheet_name]
        header_values = [cell.value for cell in ws[1]]

        # Add columns in same order as DataFrame: status, error, Boundary Code
        for col_name in ["status", "error", "Boundary Code (Mandatory)"]:
            if col_name not in header_values:
                new_col_idx = len(header_values) + 1
                cell = ws.cell(row=1, column=new_col_idx, value=col_name)
                cell.font = Font(bold=True)
                header_values.append(col_name)

        grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        # Write data rows back (without header row)
        export_df = prepare_dataframe_for_excel_export(df)
        for r_idx, row in enumerate(dataframe_to_rows(export_df, index=False, header=False), start=2):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)

                # force lock for status/error columns
                # if ws.cell(1, c_idx).value in ["status", "error"]:
                #     cell.protection = Protection(locked=True)
                #     cell.fill = grey_fill

        # Ensure sheet protection is ON
        # ws.protection.sheet = True
        # ws.protection.enable()

        # ----------------- Save to new temp file ----------------- #
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_temp_file_path = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx").name
        wb.save(output_temp_file_path)

        autofit_columns(output_temp_file_path, facility_sheet_name, auto_fit=True)

        background_tasks.add_task(cleanup_temp_file, output_temp_file_path)

        response = FileResponse(
            path=output_temp_file_path,
            filename=f"facility_validation_results_{timestamp}.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response.headers["X-Error-Count"] = str(error_count)

        return response

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Validation failed: {str(e)}")
    finally:
        if temp_input_file and os.path.exists(temp_input_file.name):
            os.unlink(temp_input_file.name)

@router.post('/facilities',
             summary='Upload and process facility Excel file',
             response_description='Returns processed Excel file with validations results')
async def upload_facilities_excel_sheet(
        facility_file: UploadFile = File(description="Excel file containing facility data"),
        facility_sheet_name: str = Form(default="EndUserIngestionTemplate",
                                        description="Name of the sheet containing facility data"),
        request_info: str = Form(default=""),
        are_facilities_onm_ready: bool = Form(description="FieldPlan ID")
):
    input_temp_file = None
    output_temp_file = None
    request_info = request_info_from_json(request_info)
    mdms_client = MDMSClient(mdms_url)

    try:
        input_temp_file, _ = await _save_upload_to_temp_file(facility_file, suffix=".xlsx")
        facility_file_path = input_temp_file.name

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"facility_ingestion_results_{timestamp}.xlsx"
        output_temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        output_temp_file.close()
        output_file_path = output_temp_file.name

        with open(facility_file_path, 'rb') as src, open(output_file_path, 'wb') as dst:
            dst.write(src.read())

        df = pd.read_excel(facility_file_path, sheet_name=facility_sheet_name)
        df.columns = [str(c).strip() for c in df.columns]
        df = df.loc[:, ~df.columns.str.startswith('Unnamed')]
        df = normalize_excel_integer_columns(df, force_columns=FACILITY_IDENTIFIER_COLUMNS)

        if 'status' not in df.columns:
            df['status'] = ''
        if 'error' not in df.columns:
            df['error'] = ''
        df['status'] = df['status'].fillna('').astype(str)
        df['error'] = df['error'].fillna('').astype(str)

        # Resolve boundary codes from State/District/Block via localization
        df = resolve_boundary_codes_for_dataframe(
            df, localization_service_url,
            boundary_code_column='Boundary Code (Mandatory)',
            logger=logger,
        )

        if facility_service_url and not df.empty:
            facility_client = FacilityServiceClient(facility_service_url)
            facility_schema = mdms_client.get_column_definitions_with_metadata(request_info,'data-ingestion.FacilityIngestionSchemaWithoutBoundaryCode')
            for index, row in df[df['status'] != 'success'].iterrows():
                try:
                    facility_data_payload = create_facility_payload(
                        request_info,
                        row,
                        are_facilities_onm_ready,
                        facility_schema,
                    )
                    response = facility_client.create_facility(facility_data_payload)
                    if response.status_code in (200, 201):
                        df.at[index, 'status'] = 'success'
                        df.at[index, 'error'] = ''
                    elif response.status_code == 400:
                        error_data = response.json()
                        error_message = error_data.get('Errors', [{}])[0].get('message', 'Unknown error')
                        df.at[index, 'status'] = 'failed'
                        df.at[index, 'error'] = error_message
                    else:
                        df.at[index, 'status'] = 'failed'
                        df.at[index, 'error'] = f'{response.status_code}: {response.text}'
                except Exception as e:
                    df.at[index, 'status'] = 'failed'
                    df.at[index, 'error'] = f'Exception: {str(e)}'

        writer = ExcelDataWriter(output_file_path, output_sheet=facility_sheet_name)
        writer.write_data(df)

        return FileResponse(
            path=output_file_path,
            filename=output_filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        logger.error(f"Error processing facility data: {e}")
        raise HTTPException(
            status_code=500,
            detail=f"Failed to process facility data: {str(e)}"
        )
    finally:
        if input_temp_file and os.path.exists(input_temp_file.name):
            os.unlink(input_temp_file.name)

@router.post('/addAssetsValidateData',
             summary='Validate bulk asset Excel file before processing',
             response_description='Returns validation report Excel with PASSED/FAILED rows')
async def validate_assets_excel_sheet(
        background_tasks: BackgroundTasks,
        asset_file: UploadFile = File(..., description="Excel file containing asset data"),
        asset_sheet_name: str = Form(default="AssetIngestionTemplate",
                                     description="Name of the sheet containing asset data"),
        request_info: str = Form(default="")
):
    temp_input_file = None
    request_info_obj = request_info_from_json(request_info)
    mdms_client = MDMSClient(mdms_url)
    try:
        temp_input_file, _ = await _save_upload_to_temp_file(asset_file, suffix=".xlsx")
        wb = load_workbook(temp_input_file.name)

        if asset_sheet_name not in wb.sheetnames:
            raise HTTPException(status_code=400, detail=f"Asset sheet '{asset_sheet_name}' not found")

        df = pd.read_excel(temp_input_file.name, sheet_name=asset_sheet_name)
        df.columns = [str(c).strip() for c in df.columns]
        df = df.loc[:, ~df.columns.str.startswith('Unnamed')]
        df = normalize_excel_integer_columns(df)

        if 'status' not in df.columns:
            df['status'] = ''
        if 'error' not in df.columns:
            df['error'] = ''

        validation_errors = asset_validation(
            df, mdms_client, request_info_obj, 'data-ingestion.AssetIngestionSchema'
        )

        error_count = 0
        for i, errs in enumerate(validation_errors):
            if errs:
                df.at[i, 'status'] = 'FAILED'
                df.at[i, 'error'] = "; ".join(dict.fromkeys(errs))
                error_count += 1
            else:
                df.at[i, 'status'] = 'PASSED'
                df.at[i, 'error'] = ''

        # Update the asset sheet in-place
        ws = wb[asset_sheet_name]
        header_values = [cell.value for cell in ws[1]]
        for col_name in ["status", "error"]:
            if col_name not in header_values:
                new_col_idx = len(header_values) + 1
                cell = ws.cell(row=1, column=new_col_idx, value=col_name)
                cell.font = Font(bold=True)
                header_values.append(col_name)

        export_df = prepare_dataframe_for_excel_export(df)
        for r_idx, row in enumerate(dataframe_to_rows(export_df, index=False, header=False), start=2):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_temp_file_path = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx").name
        wb.save(output_temp_file_path)
        autofit_columns(output_temp_file_path, asset_sheet_name, auto_fit=True)
        background_tasks.add_task(cleanup_temp_file, output_temp_file_path)

        response = FileResponse(
            path=output_temp_file_path,
            filename=f"asset_validation_results_{timestamp}.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response.headers["X-Error-Count"] = str(error_count)
        return response

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Validation failed: {str(e)}")
    finally:
        if temp_input_file and os.path.exists(temp_input_file.name):
            os.unlink(temp_input_file.name)


@router.post('/assets',
             summary='Upload and create assets from an Excel file',
             response_description='Returns processed Excel with creation results')
async def upload_assets_excel_sheet(
        asset_file: UploadFile = File(description="Excel file containing asset data"),
        asset_sheet_name: str = Form(default="AssetIngestionTemplate",
                                     description="Name of the sheet containing asset data"),
        request_info: str = Form(default="")
):
    input_temp_file = None
    output_temp_file = None
    request_info_obj = request_info_from_json(request_info)
    mdms_client = MDMSClient(mdms_url)

    try:
        input_temp_file, _ = await _save_upload_to_temp_file(asset_file, suffix=".xlsx")
        asset_file_path = input_temp_file.name

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"asset_ingestion_results_{timestamp}.xlsx"
        output_temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        output_temp_file.close()
        output_file_path = output_temp_file.name

        df = pd.read_excel(asset_file_path, sheet_name=asset_sheet_name)
        df.columns = [str(c).strip() for c in df.columns]
        df = df.loc[:, ~df.columns.str.startswith('Unnamed')]
        df = normalize_excel_integer_columns(df)

        if 'status' not in df.columns:
            df['status'] = ''
        if 'error' not in df.columns:
            df['error'] = ''
        df['status'] = df['status'].fillna('').astype(str)
        df['error'] = df['error'].fillna('').astype(str)

        asset_schema = mdms_client.get_column_definitions_with_metadata(
            request_info_obj, 'data-ingestion.AssetIngestionSchema')

        if asset_service_url and not df.empty:
            asset_client = AssetServiceClient(asset_service_url)
            vendor_lookup = {}
            if org_service_url:
                vendor_lookup = VendorRegistryClient(org_service_url).get_vendor_code_lookup(request_info_obj)
            for index, row in df[df['status'] != 'success'].iterrows():
                try:
                    asset_payload = create_asset_payload(request_info_obj, row, asset_schema, vendor_lookup)
                    response = asset_client.create_asset(asset_payload)
                    if response.status_code in (200, 201):
                        df.at[index, 'status'] = 'success'
                        df.at[index, 'error'] = ''
                    elif response.status_code == 400:
                        error_data = response.json()
                        first_error = error_data.get('Errors', [{}])[0]
                        error_message = first_error.get('message') or first_error.get('code') or 'Unknown error'
                        df.at[index, 'status'] = 'failed'
                        df.at[index, 'error'] = error_message
                    else:
                        df.at[index, 'status'] = 'failed'
                        df.at[index, 'error'] = f'{response.status_code}: {response.text}'
                except Exception as e:
                    df.at[index, 'status'] = 'failed'
                    df.at[index, 'error'] = f'Exception: {str(e)}'

        # Write a fresh single-sheet result file (the asset template has only one visible
        # sheet, so the in-place ExcelDataWriter path would leave a hidden-only workbook).
        export_df = prepare_dataframe_for_excel_export(df)
        with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
            export_df.to_excel(writer, sheet_name=asset_sheet_name, index=False)

        return FileResponse(
            path=output_file_path,
            filename=output_filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        logger.error(f"Error processing asset data: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to process asset data: {str(e)}")
    finally:
        if input_temp_file and os.path.exists(input_temp_file.name):
            os.unlink(input_temp_file.name)


@router.post('/workStreamWithFacilities',
             summary='Upload and process workstream with facilities excel file.',
             response_description='Returns processed Excel file with validation results')
async def upload_facilities_with_workstream(
        project_id_with_type_field_plan: str = Form(default="Project id of the project with type field plan"),
        request_info: str = Form(default=""),
        installation_spoc_user_name:str = Form(default=""),
        installation_spoc_user_mobile_number:str = Form(default=""),
        installation_spoc_user_email:str = Form(default="")
)->JSONResponse:
    request_info = request_info_from_json(request_info)
    #get_authorized_request_info(request_info)

    try:
        # Fetch project of type Field Plan using project_id
        if project_service_url and hrms_service_url:
            project_client = ProjectServiceClient(project_service_url)
            hrms_client = HRMSServiceClient(hrms_service_url)
            field_plan_project = project_client.search_project(request_info, project_id_with_type_field_plan)
            project = field_plan_project["Project"][0]
            if not project:
                raise Exception("Field plan id is not correct.")
            field_plan_project_facilities = project_client.search_project_facility(request_info,
                                                                                   project_id_with_type_field_plan)
            work_stream_creation_payload = get_project_creation_payload(
                request_info,
                project["project"]['name'] + "_work_stream",
                "Work Stream",
                project_id_with_type_field_plan,
                project["project"]["startDate"],
                project["project"]["endDate"],
                "Installation"
            )
            work_stream_creation_response = json.loads(project_client.create_project(work_stream_creation_payload).text)
            work_stream = work_stream_creation_response['Project'][0]

            # Link work stream project to facilities
            facilities = field_plan_project_facilities["ProjectFacilities"]
            for facility in facilities:
                project_client.create_project_facility(request_info,
                    work_stream["id"],
                    facility["facilityId"]
                )
            # Create a installation spoc user
            installation_spoc_creation_payload = get_installation_spoc_creation_payload(request_info, installation_spoc_user_name, installation_spoc_user_mobile_number,
                                                   installation_spoc_user_email)
            user_creation_response = json.loads(hrms_client.create_user(installation_spoc_creation_payload).text)
            user = user_creation_response['Employees'][0]
            staff_creation_payload = get_staff_creation_payload(request_info, user["uuid"], work_stream["id"])
            staff_creation_response = json.loads(project_client.create_project_staff(staff_creation_payload).text)
            staff = staff_creation_response['ProjectStaff']
            return JSONResponse(
                status_code=200,
                content={"staff": staff, "user": user, "work_stream": work_stream}
            )
        return JSONResponse(
            status_code=500,
            content="Connection failed with project service and hrms service."
        )

    except Exception as e:
        logger.error(f"Error processing facility data: {e}")
        raise HTTPException(
            status_code=500,
            detail=f"Failed to process facility data: {str(e)}"
        )


@router.post('/facilityWithStaff',
             summary='Upload and process facility with staff Excel file',
             response_description="Returns processed Excel file with validation results")
async def upload_facility_with_staff_excel_sheet(
        facility_with_staff: UploadFile = File(
            description="Excel file containing facility with staff data"),
        facility_sheet: str = Form(default="Facilities_Staff",
                                    description="Name of the sheet containing facility data"),
        request_info: str = Form(default=""),
        work_stream_project_id:str = Form(default="")
):
    input_temp_file = None
    output_temp_file = None
    request_info = request_info_from_json(request_info)
    #get_authorized_request_info(request_info)

    try:
        # Create input temporary file
        input_temp_file, _ = await _save_upload_to_temp_file(facility_with_staff, suffix=".xlsx")
        facility_with_staff_file_path = input_temp_file.name

        # Create output file with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"facility_with_staff_ingestion_results_{timestamp}.xlsx"
        output_temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        output_temp_file.close()
        output_file_path = output_temp_file.name

        # Copy input to output file first
        with open(facility_with_staff_file_path, 'rb') as src, open(output_file_path, 'wb') as dst:
            dst.write(src.read())

        # Read the Excel file
        df = pd.read_excel(facility_with_staff_file_path, sheet_name=facility_sheet)

        # Add status and error columns if they don't exist
        if 'status' not in df.columns:
            df['status'] = ''
        if 'error' not in df.columns:
            df['error'] = ''

        # Process each row if services are available
        if project_service_url and not df.empty:
            project_client = ProjectServiceClient(project_service_url)
            hrms_client = HRMSServiceClient(hrms_service_url)
            work_stream_project = project_client.search_project(request_info, work_stream_project_id)
            work_stream = work_stream_project["Project"][0]
            for index, row in df.iterrows():
                if row.get("status", "") != "success":
                    try:
                        # Create project of type facility
                        facility_creation_payload = get_project_creation_payload(request_info, row.get('Health Centre Name (Mandatory)', ''), "Facility",
                                                                                 work_stream_project_id, work_stream["startDate"],work_stream["endDate"],"")
                        facility_creation_response = project_client.create_project(facility_creation_payload)
                        facility = json.loads(facility_creation_response.text)
                        if facility_creation_response.status_code in [200, 201, 202]:
                            df.at[index, 'status'] = 'success'
                            # Check if user already exists and validate roles
                            user_search_payload = get_user_creation_payload_staff(request_info, row)
                            existing_user_response = hrms_client.search_user(user_search_payload)
                            existing_user = None
                            if existing_user_response.status_code == 200:
                                response_data = existing_user_response.json()
                                employees = response_data.get("Employees", [])
                                if employees:
                                    existing_user = employees[0]

                            if existing_user:
                                # Check for role mismatch
                                role_check = check_role_mismatch_for_user_type(existing_user, "staff")
                                if role_check["has_mismatch"]:
                                    df.at[index, 'status'] = 'error'
                                    df.at[index, 'error'] = f"Role mismatch detected: {role_check['mismatch_details']}. Current roles: {', '.join(role_check['current_roles'])}. Expected roles: {', '.join(role_check['expected_roles'])}"
                                    continue
                                else:
                                    # Use existing user
                                    user_uuid = existing_user.get("uuid")
                                    df.at[index, 'status'] = 'success'
                            else:
                                # Create new user
                                user_creation_payload = get_user_creation_payload_staff(request_info, row)
                                user_creation_response = hrms_client.create_user(user_creation_payload)
                                user = json.loads(user_creation_response.text)
                                if user_creation_response.status_code in [200, 201, 202]:
                                    user_uuid = user["Employees"][0]["uuid"]
                                    df.at[index, 'status'] = 'success'
                                else:
                                    df.at[index, 'status'] = 'failed'
                                    df.at[index, 'error'] = f"User Creation Error: {user_creation_response.status_code} - {user.get('Errors', [{}])[0].get('message', 'Unknown error')}"
                                    continue

                            # Validate user_uuid before staff creation
                            if not user_uuid:
                                df.at[index, 'status'] = 'failed'
                                df.at[index, 'error'] = "User UUID is required for staff creation but was not obtained"
                                continue

                            # Create staff
                            staff_creation_payload = get_staff_creation_payload(request_info, user_uuid, facility["Project"][0]["id"])
                            staff_creation_response = project_client.create_project_staff(staff_creation_payload)
                            if staff_creation_response.status_code in [200, 201, 202]:
                                df.at[index,'status'] = 'success'
                                df.at[index, 'error'] = ''
                            else:
                                df.at[index, 'status'] = 'failed'
                                df.at[index, 'error'] = f"Staff Creation Error: {staff_creation_response.status_code} - {staff_creation_response.text}"
                        else:
                            df.at[index, 'status'] = 'failed'
                            df.at[index, 'error'] = f"Facility Creation Error: {facility_creation_response.status_code} - {facility_creation_response.text}"
                    except Exception as e:
                        df.at[index, 'status'] = 'failed'
                        df.at[index, 'error'] = f"Processing Error: {str(e)}"

        # Write data to the same sheet name that was read
        with pd.ExcelWriter(output_file_path, engine='openpyxl', mode='a') as writer:
            # Remove the existing sheet if it exists
            if facility_sheet in writer.book.sheetnames:
                idx = writer.book.sheetnames.index(facility_sheet)
                writer.book.remove(writer.book.worksheets[idx])
            # Write data to the sheet
            df.to_excel(writer, sheet_name=facility_sheet, index=False)

        return FileResponse(
            path=output_file_path,
            filename=output_filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        logger.error(f"Error processing facility data: {e}")
        raise HTTPException(
            status_code=500,
            detail=f"Failed to process facility data: {str(e)}"
        )
    finally:
        if input_temp_file and os.path.exists(input_temp_file.name):
            os.unlink(input_temp_file.name)



@router.post('/facilityWithSupervisors',
             summary='Upload and process facility with supervisors Excel file',
             response_description="Returns processed Excel file with validation results")
async def upload_facility_with_supervisors_excel_sheet(
        facility_with_supervisors: UploadFile = File(
            description="Excel file containing facility with supervisors data"),
        facility_sheet: str = Form(default="Facilities_Supervisors",
                                    description="Name of the sheet containing facility data"),
        request_info: str = Form(default=""),
        work_stream_project_id:str = Form(default="")
):
    input_temp_file = None
    output_temp_file = None
    request_info = request_info_from_json(request_info)
    #get_authorized_request_info(request_info)

    try:
        # Create input temporary file
        input_temp_file, _ = await _save_upload_to_temp_file(facility_with_supervisors, suffix=".xlsx")
        facility_with_supervisors_file_path = input_temp_file.name

        # Create output file with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"facility_with_supervisor_ingestion_results_{timestamp}.xlsx"
        output_temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        output_temp_file.close()
        output_file_path = output_temp_file.name

        # Copy input to output file first
        with open(facility_with_supervisors_file_path, 'rb') as src, open(output_file_path, 'wb') as dst:
            dst.write(src.read())

        # Read the Excel file
        df = pd.read_excel(facility_with_supervisors_file_path, sheet_name=facility_sheet)

        # Add status and error columns if they don't exist
        if 'status' not in df.columns:
            df['status'] = ''
        if 'error' not in df.columns:
            df['error'] = ''

        # Process each row if services are available
        if project_service_url and not df.empty:
            project_client = ProjectServiceClient(project_service_url)
            hrms_client = HRMSServiceClient(hrms_service_url)
            work_stream_project = project_client.search_project(request_info, work_stream_project_id)
            work_stream = work_stream_project["Project"][0]
            for index, row in df.iterrows():
                if row.get("status", "") != "success":
                    try:
                        existing_facility = project_client.search_project_facility(request_info, work_stream_project_id)
                        facility_list = existing_facility.get('ProjectFacilities', [])

                        facility_created = False
                        if facility_list:
                            facility = facility_list[0]
                            facility_created = False  # existing, not newly created
                        else:
                            # Create facility if not found
                            facility_creation_payload = get_project_creation_payload(request_info, row.get('Health Centre Name (Mandatory)', ''), "Facility",
                                                                                 work_stream_project_id, work_stream["startDate"],work_stream["endDate"],"")
                            facility_creation_response = project_client.create_project(facility_creation_payload)
                            if facility_creation_response.status_code not in [200, 201, 202]:
                                df.at[index, 'status'] = 'failed'
                                df.at[index, 'error'] = (
                                    f"Facility Creation Error: {facility_creation_response.status_code} - {facility_creation_response.text}"
                                )
                                continue

                            facility = json.loads(facility_creation_response.text)
                            facility_created = True

                        # 🧠 Correctly extract project ID based on the structure
                        if facility_created:
                            project_id = facility["Project"][0]["id"]
                        else:
                            project_id = facility["projectId"]
                        # Check if user already exists and validate roles
                        user_search_payload = get_user_creation_payload_supervisors(request_info, row)
                        existing_user_response = hrms_client.search_user(user_search_payload)
                        existing_user = None
                        if existing_user_response.status_code == 200:
                            response_data = existing_user_response.json()
                            employees = response_data.get("Employees", [])
                            if employees:
                                existing_user = employees[0]

                        if existing_user:
                            # Check for role mismatch
                            role_check = check_role_mismatch_for_user_type(existing_user, "supervisor")
                            if role_check["has_mismatch"]:
                                df.at[index, 'status'] = 'error'
                                df.at[index, 'error'] = f"Role mismatch detected: {role_check['mismatch_details']}. Current roles: {', '.join(role_check['current_roles'])}. Expected roles: {', '.join(role_check['expected_roles'])}"
                                continue
                            else:
                                # Use existing user
                                user_uuid = existing_user.get("uuid")
                                df.at[index, 'status'] = 'success'
                        else:
                            # Create new user
                            user_creation_payload = get_user_creation_payload_supervisors(request_info, row)
                            user_creation_response = hrms_client.create_user(user_creation_payload)
                            user = json.loads(user_creation_response.text)
                            if user_creation_response.status_code in [200, 201, 202]:
                                user_uuid = user["Employees"][0]["uuid"]
                                df.at[index, 'status'] = 'success'
                            else:
                                df.at[index, 'status'] = 'failed'
                                df.at[index, 'error'] = f"User Creation Error: {user_creation_response.status_code} - {user.get('Errors', [{}])[0].get('message', 'Unknown error')}"
                                continue

                        # Create staff
                        staff_creation_payload = get_staff_creation_payload(request_info, user_uuid, project_id)
                        staff_creation_response = project_client.create_project_staff(staff_creation_payload)
                        if staff_creation_response.status_code in [200, 201, 202]:
                            df.at[index,'status'] = 'success'
                            df.at[index, 'error'] = ''
                        else:
                            df.at[index, 'status'] = 'failed'
                            df.at[index, 'error'] = f"Staff Creation Error: {staff_creation_response.status_code} - {staff_creation_response.text}"
                    except Exception as e:
                        df.at[index, 'status'] = 'failed'
                        df.at[index, 'error'] = f"Processing Error: {str(e)}"

        # Write data to the same sheet name that was read
        with pd.ExcelWriter(output_file_path, engine='openpyxl', mode='a') as writer:
            # Remove the existing sheet if it exists
            if facility_sheet in writer.book.sheetnames:
                idx = writer.book.sheetnames.index(facility_sheet)
                writer.book.remove(writer.book.worksheets[idx])
            # Write data to the sheet
            df.to_excel(writer, sheet_name=facility_sheet, index=False)

        return FileResponse(
            path=output_file_path,
            filename=output_filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        logger.error(f"Error processing facility data: {e}")
        raise HTTPException(
            status_code=500,
            detail=f"Failed to process facility data: {str(e)}"
        )
    finally:
        if input_temp_file and os.path.exists(input_temp_file.name):
            os.unlink(input_temp_file.name)

@router.post('/facilityWithSupervisorUpdateWorkflowState',
             summary='Upload and process facility with supervisors Excel file',
             response_description="Returns processed Excel file with validation results")
async def upload_facility_with_supervisors_workflow_state_excel_sheet(
        facility_with_supervisors: UploadFile = File(
            description="Excel file containing facility with supervisors data"),
        facility_sheet: str = Form(default="Facilities_Supervisors",
                                    description="Name of the sheet containing facility data"),
        request_info: str = Form(default=""),
        work_stream_project_id:str = Form(default="")
):
    input_temp_file = None
    output_temp_file = None
    request_info = request_info_from_json(request_info)
    #get_authorized_request_info(request_info)

    try:
        # Create input temporary file
        input_temp_file, _ = await _save_upload_to_temp_file(facility_with_supervisors, suffix=".xlsx")
        facility_with_supervisors_file_path = input_temp_file.name

        # Create output file with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"facility_with_supervisor_ingestion_results_{timestamp}.xlsx"
        output_temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        output_temp_file.close()
        output_file_path = output_temp_file.name

        # Copy input to output file first
        with open(facility_with_supervisors_file_path, 'rb') as src, open(output_file_path, 'wb') as dst:
            dst.write(src.read())

        # Read the Excel file
        df = pd.read_excel(facility_with_supervisors_file_path, sheet_name=facility_sheet)

        # Add status and error columns if they don't exist
        if 'status' not in df.columns:
            df['status'] = ''
        if 'error' not in df.columns:
            df['error'] = ''

        # Process each row if services are available
        if project_service_url and not df.empty:
            project_client = ProjectServiceClient(project_service_url)
            hrms_client = HRMSServiceClient(hrms_service_url)
            work_stream_project = project_client.search_project(request_info, work_stream_project_id)
            work_stream = work_stream_project["Project"][0]
            for index, row in df.iterrows():
                if row.get("status", "") != "success":
                    try:
                        existing_facility = project_client.search_project_facility(request_info, work_stream_project_id)
                        facility_list = existing_facility.get('ProjectFacilities', [])

                        facility_created = False
                        if facility_list:
                            facility = facility_list[0]
                            facility_created = False  # existing, not newly created
                        else:
                            # Create facility if not found
                            facility_creation_payload = get_project_creation_payload(request_info, row.get('Health Centre Name (Mandatory)', ''), "Facility",
                                                                                 work_stream_project_id, work_stream["startDate"],work_stream["endDate"],"")
                            facility_creation_response = project_client.create_project(facility_creation_payload)
                            if facility_creation_response.status_code not in [200, 201, 202]:
                                df.at[index, 'status'] = 'failed'
                                df.at[index, 'error'] = (
                                    f"Facility Creation Error: {facility_creation_response.status_code} - {facility_creation_response.text}"
                                )
                                continue

                            facility = json.loads(facility_creation_response.text)
                            facility_created = True

                        # 🧠 Correctly extract project ID based on the structure
                        if facility_created:
                            project_id = facility["Project"][0]["id"]
                        else:
                            project_id = facility["projectId"]
                        # Check if user already exists and validate roles
                        # Determine user type based on Role column
                        user_type = "supervisor"  # default
                        if 'Role' in df.columns:
                            role_value = df.at[index, 'Role']
                            if role_value and str(role_value).strip().lower() == 'supervisor':
                                user_type = "supervisor"
                            else:
                                user_type = "staff"

                        # Create search payload based on user type
                        if user_type == "supervisor":
                            user_search_payload = get_user_creation_payload_supervisors(request_info, row)
                        else:
                            user_search_payload = get_user_creation_payload_staff(request_info, row)

                        existing_user_response = hrms_client.search_user(user_search_payload)
                        existing_user = None
                        if existing_user_response.status_code == 200:
                            response_data = existing_user_response.json()
                            employees = response_data.get("Employees", [])
                            if employees:
                                existing_user = employees[0]

                        if existing_user:
                            # Check for role mismatch
                            role_check = check_role_mismatch_for_user_type(existing_user, user_type)
                            if role_check["has_mismatch"]:
                                df.at[index, 'status'] = 'error'
                                df.at[index, 'error'] = f"Role mismatch detected: {role_check['mismatch_details']}. Current roles: {', '.join(role_check['current_roles'])}. Expected roles: {', '.join(role_check['expected_roles'])}"
                                continue
                            else:
                                # Use existing user
                                user_uuid = existing_user.get("uuid")
                        else:
                            # Create new user based on role type
                            if user_type == "supervisor":
                                user_creation_payload = get_user_creation_payload_supervisors(request_info, row)
                            else:
                                user_creation_payload = get_user_creation_payload_staff(request_info, row)

                            user_creation_response = hrms_client.create_user(user_creation_payload)
                            user = json.loads(user_creation_response.text)
                            if user_creation_response.status_code in [200, 201, 202]:
                                user_uuid = user["Employees"][0]["uuid"]
                            else:
                                df.at[index, 'status'] = 'failed'
                                df.at[index, 'error'] = f"User Creation Error: {user_creation_response.status_code} - {user.get('Errors', [{}])[0].get('message', 'Unknown error')}"
                                continue

                        # Validate user_uuid before staff creation
                        if not user_uuid:
                            df.at[index, 'status'] = 'failed'
                            df.at[index, 'error'] = "User UUID is required for staff creation but was not obtained"
                            continue

                        # Create staff
                        staff_creation_payload = get_staff_creation_payload(request_info, user_uuid, project_id)
                        staff_creation_response = project_client.create_project_staff(staff_creation_payload)
                        if staff_creation_response.status_code in [200, 201, 202]:
                            # Validate Role column exists
                            if 'Role' not in df.columns:
                                df.at[index, 'status'] = 'failed'
                                df.at[index, 'error'] = "Role column is required for workflow state updates"
                                continue
                            # update workflow state
                            role_value = df.at[index,'Role']
                            if role_value and str(role_value).strip().lower() == 'supervisor':
                                update_workflow_state_response = project_client.update_workflow(request_info, work_stream_project_id, 'ASSIGN_FIELD_SUPERVISOR')
                            else:
                                update_workflow_state_response = project_client.update_workflow(request_info, work_stream_project_id,
                                                                                                'ASSIGN_FIELD_STAFF')
                            if update_workflow_state_response.status_code in [200, 201, 202]:
                                df.at[index,'status'] = 'success'
                                df.at[index, 'error'] = ''
                            else:
                                df.at[index, 'status'] = 'failed'
                                df.at[
                                    index, 'error'] = f"Update Workflow state Error: {update_workflow_state_response.status_code} - {update_workflow_state_response.text}"
                        else:
                            df.at[index, 'status'] = 'failed'
                            df.at[index, 'error'] = f"Staff Creation Error: {staff_creation_response.status_code} - {staff_creation_response.text}"
                    except Exception as e:
                        df.at[index, 'status'] = 'failed'
                        df.at[index, 'error'] = f"Processing Error: {str(e)}"

        # Write data to the same sheet name that was read
        with pd.ExcelWriter(output_file_path, engine='openpyxl', mode='a') as writer:
            # Remove the existing sheet if it exists
            if facility_sheet in writer.book.sheetnames:
                idx = writer.book.sheetnames.index(facility_sheet)
                writer.book.remove(writer.book.worksheets[idx])
            # Write data to the sheet
            df.to_excel(writer, sheet_name=facility_sheet, index=False)

        return FileResponse(
            path=output_file_path,
            filename=output_filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        logger.error(f"Error processing facility data: {e}")
        raise HTTPException(
            status_code=500,
            detail=f"Failed to process facility data: {str(e)}"
        )
    finally:
        if input_temp_file and os.path.exists(input_temp_file.name):
            os.unlink(input_temp_file.name)

@router.post('/projects',
             summary='Upload and process project Excel file',
             response_description='Returns processed Excel file with validations results')
async def upload_projects_excel_sheet(
        project_file: UploadFile = File(description="Excel file containing project data"),
        project_sheet_name: str = Form(default="Project Data",
                                        description="Name of the sheet containing project data"),
        request_info: str = Form(default="")
):
    input_temp_file = None
    output_temp_file = None
    request_info = request_info_from_json(request_info)
    #get_authorized_request_info(request_info)

    try:
        input_temp_file, _ = await _save_upload_to_temp_file(project_file, suffix=".xlsx")
        project_file_path = input_temp_file.name

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"project_ingestion_results_{timestamp}.xlsx"
        output_temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        output_temp_file.close()
        output_file_path = output_temp_file.name

        with open(project_file_path, 'rb') as src, open(output_file_path, 'wb') as dst:
            dst.write(src.read())

        df = pd.read_excel(project_file_path, sheet_name=project_sheet_name)

        if 'status' not in df.columns:
            df['status'] = ''
        if 'error' not in df.columns:
            df['error'] = ''
        if 'Project ID' not in df.columns:
            df['Project ID'] = ''

        if project_service_url and hrms_service_url and not df.empty:
            hrms_client = HRMSServiceClient(hrms_service_url)
            project_client = ProjectServiceClient(project_service_url)
            for index, row in df[df['status'] != 'success'].iterrows():
                try:

                    project_data_payload = create_project_payload(request_info, row)
                    response = project_client.create_project(project_data_payload)
                    response_data = response.json()

                    if df.at[index, 'Project Type'] == 'Field Plan':
                        name = df.at[index, 'Name']
                        mobile_number_raw = df.at[index, 'Mobile Number']
                        email_value = df.at[index, 'Email']
                        if pd.isna(email_value) or not email_value:
                            df.at[index, 'status'] = 'failed'
                            df.at[index, 'error'] = 'Email is required for Field Plan projects'
                            continue
                        if pd.isna(mobile_number_raw) or not mobile_number_raw:
                            df.at[index, 'status'] = 'failed'
                            df.at[index, 'error'] = 'Mobile Number is required for Field Plan projects'
                            continue


                        mobile_number = str(int(mobile_number_raw))
                        email = str(email_value).strip()

                        spoc_payload = get_installation_spoc_creation_payload(request_info, name, mobile_number, email)

                        user_response = hrms_client.search_user(spoc_payload)

                        if user_response.status_code not in [200, 201, 202]:
                            df.at[index, 'status'] = 'failed'
                            df.at[
                                index, 'error'] = f"User search failed with status: {user_response.status_code} - {user_response.text}"
                            continue

                        response_body = json.loads(user_response.text)
                        employee_list = response_body.get("Employees", [])

                        # Filter for matching email
                        matched_user = None
                        for emp in employee_list:
                            if emp["user"]["emailId"].strip() == email:
                                matched_user = emp["user"]
                                break

                        if not matched_user:
                            df.at[index, 'status'] = 'failed'
                            df.at[index, 'error'] = f"No matching user found for email: {email}"
                            continue

                    if response.status_code in [200, 201, 202] and isinstance(response_data.get('Project'), list) and response_data[
                        'Project']:
                        if df.at[index, 'Project Type'] == 'Field Plan':

                            user_uuid = matched_user.get("uuid")
                            project_id = response_data['Project'][0].get('id')

                            staff_payload = get_staff_creation_payload(request_info, user_uuid, project_id)
                            staff_response = project_client.create_project_staff(staff_payload)

                            if staff_response.status_code in [200, 201, 202]:

                                staff_search_payload = get_staff_search_payload(request_info, user_uuid)
                                staff_search_response = project_client.search_project_staff_by_id(staff_search_payload)
                                if staff_search_response.status_code in [200, 201]:
                                    logger.debug(f"Staff search response for user {user_uuid}: {staff_search_response.text}")

                                    staff_list = staff_search_response.json().get("ProjectStaff", [])
                                    logger.debug(f"Found {len(staff_list)} staff members for user {user_uuid}")

                                    if len(staff_list) == 1:
                                        sms_request = {
                                            "mobileNumber": mobile_number,
                                            "message": "Yor are assigned to the field plan",
                                            "expiryTime": None
                                        }
                                        producer = Producer()
                                        producer.send("egov.core.notification.sms", sms_request)
                                        producer.close()


                                df.at[index, 'status'] = 'success'
                                df.at[index, 'error'] = ''
                                df.at[index, 'Project ID'] = project_id
                            else:
                                df.at[index, 'status'] = 'failed'
                                df.at[index, 'error'] = (
                                    f"Staff Creation Error: {staff_response.status_code} - {staff_response.text}")
                        else:
                            project_id = response_data['Project'][0].get('id')
                            df.at[index, 'status'] = 'success'
                            df.at[index, 'error'] = ''
                            df.at[index, 'Project ID'] = project_id
                    elif response.status_code == 400:
                        error_data = response.json()
                        error_message = error_data.get('Errors', [{}])[0].get('message', 'Unknown error')
                        df.at[index, 'status'] = 'failed'
                        df.at[index, 'error'] = error_message
                    else:
                        df.at[index, 'status'] = 'failed'
                        df.at[index, 'error'] = f'{response.status_code}: {response.text}'
                except Exception as e:
                    df.at[index, 'status'] = 'failed'
                    df.at[index, 'error'] = f'Exception: {str(e)}'

        writer = ExcelDataWriter(output_file_path, output_sheet=project_sheet_name)
        writer.write_data(df)

        return FileResponse(
            path=output_file_path,
            filename=output_filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        logger.error(f"Error processing project data: {e}")
        raise HTTPException(
            status_code=500,
            detail=f"Failed to process project data: {str(e)}"
        )
    finally:
        if input_temp_file and os.path.exists(input_temp_file.name):
            os.unlink(input_temp_file.name)

@router.post('/facilitySelection',
             summary='Upload and process facility selection Excel file',
             response_description='Returns processed Excel file with validations results')
async def upload_facility_selection_excel_sheet(
        facility_selection_file: UploadFile = File(description="Excel file containing facility selection data"),
        project_id: str = Form(...),
        facility_selection_sheet_name: str = Form(default="Facility Selection Template",
                                        description="Name of the sheet containing facility selection data"),
        request_info: str = Form(default="")
):
    input_temp_file = None
    output_temp_file = None
    request_info = request_info_from_json(request_info)
    #get_authorized_request_info(request_info)

    try:
        input_temp_file, _ = await _save_upload_to_temp_file(facility_selection_file, suffix=".xlsx")
        project_file_path = input_temp_file.name

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"project_facility_ingestion_results_{timestamp}.xlsx"
        output_temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        output_temp_file.close()
        output_file_path = output_temp_file.name

        with open(project_file_path, 'rb') as src, open(output_file_path, 'wb') as dst:
            dst.write(src.read())

        df = pd.read_excel(project_file_path, sheet_name=facility_selection_sheet_name)

        if 'status' not in df.columns:
            df['status'] = ''
        if 'error' not in df.columns:
            df['error'] = ''

        if project_service_url and not df.empty:
            project_client = ProjectServiceClient(project_service_url)
            for index, row in df[(df['status'] != 'success') & (df['Selection?'] == 'Yes')].iterrows():
                try:
                    facility_id = row.get("HC ID")
                    if pd.isna(facility_id):
                        df.at[index, 'status'] = 'failed'
                        df.at[index, 'error'] = 'HC ID must not be null.'
                        continue
                    response = project_client.create_project_facility(request_info, project_id, facility_id)
                    if response.status_code in (200, 201, 202):
                        df.at[index, 'status'] = 'success'
                        df.at[index, 'error'] = ''
                    elif response.status_code == 400:
                        error_data = response.json()
                        error_message = error_data.get('Errors', [{}])[0].get('message', 'Unknown error')
                        df.at[index, 'status'] = 'failed'
                        df.at[index, 'error'] = error_message
                    else:
                        df.at[index, 'status'] = 'failed'
                        df.at[index, 'error'] = f'{response.status_code}: {response.text}'
                except Exception as e:
                    df.at[index, 'status'] = 'failed'
                    df.at[index, 'error'] = f'Exception: {str(e)}'

        writer = ExcelDataWriter(output_file_path, output_sheet=facility_selection_sheet_name)
        writer.write_data(df)

        return FileResponse(
            path=output_file_path,
            filename=output_filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        logger.error(f"Error processing facility selection data: {e}")
        raise HTTPException(
            status_code=500,
            detail=f"Failed to process facility selection data: {str(e)}"
        )
    finally:
        if input_temp_file and os.path.exists(input_temp_file.name):
            os.unlink(input_temp_file.name)


def get_hrms_employee_info(codes: List[str], db_conn) -> Dict[str, str]:
    try:
        with db_conn.cursor() as cursor:
            sql = "SELECT code, tenantid  FROM eg_hrms_employee WHERE code = ANY (%s)"
            cursor.execute(sql, (codes,))
            rows = cursor.fetchall()
            return {row[0]: row[1] for row in rows}
    except Exception as e:
        logger.error(f"Error fetching HRMS employee info: {e}")
        return {}

def get_tenant_mapping(request_info: RequestInfo, tenant_ids: List[str]) -> Dict:
    """
    Fetch tenant mapping from MDMS for PHC subtypes
    """
    all_tenant_data = {}

    for tenant_id in tenant_ids:
        try:
            search_url = f"{mdms_url}/egov-mdms-service/v1/_search"
            search_payload = {
                "RequestInfo": request_info.model_dump(by_alias=True, exclude_none=True),
                "MdmsCriteria": {
                    "tenantId": tenant_id,
                    "moduleDetails": [
                        {
                            "moduleName": "tenant",
                            "masterDetails": [
                                {
                                    "name": "tenants"
                                }
                            ]
                        }
                    ]
                }
            }
            response = requests.post(search_url, json=search_payload)
            if response.status_code == 200:
                data = response.json()
                tenants = data.get("MdmsRes", {}).get("tenant", {}).get("tenants", [])
                all_tenant_data.update({t["code"]: t for t in tenants if t.get("code") and t["code"] not in all_tenant_data})
        except Exception as e:
            logger.error(f"Error fetching tenant mapping from MDMS: {e}")

    return all_tenant_data


def get_block_mapping_from_mdms(request_info: RequestInfo, tenant_ids: List[str]) -> Dict[str, dict]:
    """
    Fetch block mapping from MDMS where moduleName is 'Incident' and masterDetails name is 'Block'.
    Returns a dictionary with 'code' from each 'data' object as the key.
    """
    block_mapping = {}

    for tenant_id in tenant_ids:
        try:
            search_url = f"{mdms_url}/egov-mdms-service/v1/_search"
            search_payload = {
                "RequestInfo": request_info.model_dump(by_alias=True, exclude_none=True),
                "MdmsCriteria": {
                    "tenantId": tenant_id,
                    "moduleDetails": [
                        {
                            "moduleName": "Incident",
                            "masterDetails": [
                                {
                                    "name": "Block"
                                }
                            ]
                        }
                    ]
                }
            }

            response = requests.post(search_url, json=search_payload)
            if response.status_code == 200:
                data = response.json()
                mdms_blocks = data.get("MdmsRes", {}).get("Incident", {}).get("Block", [])

                for block in mdms_blocks:
                    code = block.get("code")
                    if code and code not in block_mapping:
                        block_mapping[code] = block

        except Exception as e:
            logger.error(f"Error fetching block mapping from MDMS for tenant {tenant_id}: {e}")

    return block_mapping

def create_mapping_dicts(mapping_file: UploadFile, sheet_name: str):
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_file:
        temp_file.write(mapping_file.file.read())
        temp_file_path = temp_file.name

    mapping_df = pd.read_excel(temp_file_path, sheet_name=sheet_name)
    mapping_df.columns = mapping_df.columns.str.strip()
    mapping_df = mapping_df.astype(str).apply(lambda x: x.str.strip())

    os.unlink(temp_file_path)

    subtype_mapping = {
        (row['Existing Issue Type'], row['Existing Ticket Sub Type ( Saure eMitra)']):
        (row['New Issue Type'], row['New Ticket Sub type'])
        for _, row in mapping_df.iterrows()
    }

    return subtype_mapping

def get_user_info_for_mizoram(usernames: List[str], db_conn) -> Dict[str, str]:
    try:
        with db_conn.cursor() as cursor:
            sql = "SELECT username, tenantid FROM eg_user WHERE username = ANY (%s)"
            cursor.execute(sql, (usernames,))
            rows = cursor.fetchall()
            return {row[0]: row[1] for row in rows}
    except Exception as e:
        logger.error(f"Error fetching user info for Mizoram: {e}")
        return {}


@router.post("/legacy_ticket_ingestion", summary="Upload and ingest legacy tickets Excel file")
async def upload_legacy_ticket_excel_sheet(
    legacy_ticket_file: UploadFile = File(...),
    legacy_ticket_sheet_name: str = Form(default="Legacy Tickets"),
    mapping_type_subtype_file: UploadFile = File(...),
    mapping_type_subtype_sheet_name: str = Form(default="Mapping Old_New_v1.0"),
    request_info: str = Form(default="")
):
    migration_id = str(uuid.uuid4())
    request_info_obj = request_info_from_json(request_info)
    get_authorized_request_info(request_info_obj)

    subtype_mapping = create_mapping_dicts(mapping_type_subtype_file, mapping_type_subtype_sheet_name)
    tenant_creator_mapping = TENANT_CREATOR_MAPPING

    input_temp_file, _ = await _save_upload_to_temp_file(legacy_ticket_file, suffix=".xlsx")
    excel_file_path = input_temp_file.name

    df = pd.read_excel(excel_file_path, sheet_name=legacy_ticket_sheet_name)
    df.columns = df.columns.str.strip()
    df = df.reindex(columns=df.columns.tolist() + ['ticket_id', 'employee_info'], fill_value='')

    unique_states = df["State"].dropna().str.strip().unique()
    tenant_ids = [tenant_creator_mapping.get(state, {}).get("tenantId") for state in unique_states]

    tenant_mapping = get_tenant_mapping(request_info_obj, tenant_ids)
    block_mapping = get_block_mapping_from_mdms(request_info_obj, tenant_ids)

    conn = psycopg2.connect(**DB_CONFIG)
    codes = [str(row.get("NIN_HFR ID", "")).strip() for i, row in df.iterrows()
             if str(df.at[i, 'status']).strip().lower() not in ['duplicate', 'error']]
    employee_info = get_hrms_employee_info(codes, conn)

    usernames = [str(row.get("Actual User Name", "")).strip() for i, row in df.iterrows()
                 if str(row.get("State", "")).strip() == "Mizoram" and str(df.at[i, 'status']).strip().lower() not in ['duplicate', 'error']]
    user_info = get_user_info_for_mizoram(usernames, conn)

    for idx, row in df.iterrows():
        try:
            status = str(df.at[idx, 'status']).strip().lower()
            if status in ['duplicate', 'error']:
                continue

            state = str(row.get("State", "")).strip()
            if state == "Mizoram":
                identifier = str(row.get("Actual User Name", "")).strip()
                tenant_id = user_info.get(identifier)
            else:
                identifier = str(row.get("NIN_HFR ID", "")).strip()
                tenant_id = employee_info.get(identifier)

            if not tenant_id:
                df.at[idx, 'status'] = 'failed'
                df.at[idx, 'error'] = f'Employee not found for code: {identifier}'
                df.at[idx, 'employee_info'] = 'Not found'
                continue

            df.at[idx, 'employee_info'] = 'Found'

            tenant_details = tenant_mapping.get(tenant_id, {})
            if not tenant_details:
                df.at[idx, 'status'] = 'failed'
                df.at[idx, 'error'] = f'Tenant mapping not found for tenant ID: {tenant_id}'
                continue

            incident_payload = build_incident_payload(row, identifier, tenant_details, block_mapping, migration_id,
                                                      tenant_creator_mapping.get(state, {}), subtype_mapping)
            response = submit_incident_payload(incident_payload, tenant_creator_mapping.get(state, {}))
            process_response(response, df, idx, identifier)

        except Exception as e:
            df.at[idx, 'status'] = 'failed'
            df.at[idx, 'error'] = str(e)

    return write_and_return_excel(df, legacy_ticket_sheet_name)

def build_incident_payload(row, identifier, tenant_details, block_mapping, migration_id, creator_info, subtype_mapping):
    ticket_type = str(row.get("Ticket Type", "")).strip()
    ticket_subtype = str(row.get("Ticket Sub Type", "")).strip()
    system_functional = {"Yes": "FUNCTIONAL", "No": "NON_FUNCTIONAL"}.get(
        str(row.get("Is the solar system working?", "")).strip(), "")
    comments = str(row.get("Comments", "")).strip()[:256]
    mapped_pair = subtype_mapping.get((ticket_type, ticket_subtype))

    if not ticket_type or not ticket_subtype or not mapped_pair:
        raise ValueError("Missing or invalid Ticket Type/Sub Type")

    block_code = tenant_details.get("city", {}).get("blockCode", "")
    block = block_mapping.get(block_code, {}).get("name", "")

    incident_payload = {
        "incidentType": mapped_pair[0],
        "incidentSubtype": mapped_pair[1],
        "comments": comments,
        "systemFunctional": system_functional,
        "tenantId": tenant_details.get("code", ""),
        "migrationId": migration_id,
        "district": tenant_details.get("city", {}).get("districtCode", ""),
        "block": block,
        "phcType": tenant_details.get("code", ""),
        "phcSubType": tenant_details.get("centreType", ""),
        "additionalDetail": {"fileStoreId": [], "reopenreason": [], "rejectReason": [],
                              "sendBackReason": [], "sendBackSubReason": []},
        "source": "web",
        "reporter": {
            "uuid": creator_info.get("uuid"),
            "tenantId": creator_info.get("tenantId")
        }
    }

    if pd.notnull(row.get("Unique_ID")):
        incident_payload["legacyId"] = str(row.get("Unique_ID")).strip()

    reported_date = row.get("Actual_Reported_Date (mm/dd/yyyy)", None)
    if pd.notnull(reported_date):
        dt = pd.to_datetime(reported_date, format="%d/%m/%Y", errors='coerce') if isinstance(reported_date, str) else pd.to_datetime(reported_date, errors='coerce')
        if pd.notnull(dt):
            incident_payload["filedDate"] = int(dt.timestamp() * 1000)

    return incident_payload

def submit_incident_payload(payload, creator):
    profile = USER_PROFILE

    return requests.post(
        f"{im_services_url}/im-services/v2/request/_create",
        json={
            "RequestInfo": {
                "apiId": "Rainmaker",
                "authToken": "79967889-fbf5-42c6-9bd3-4adc0dbe7692",
                "userInfo": {
                    "id": creator.get("id"),
                    "uuid": creator.get("uuid"),
                    "userName": profile["userName"],
                    "name": profile["name"],
                    "mobileNumber": creator.get("mobileNumber"),
                    "emailId": None,
                    "locale": None,
                    "type": "EMPLOYEE",
                    "roles": [
                        {
                            "name": "Complainant",
                            "code": "COMPLAINANT",
                            "tenantId": creator.get("tenantId")
                        },
                        {
                            "name": "Employee",
                            "code": "EMPLOYEE",
                            "tenantId": creator.get("tenantId")
                        }
                    ],
                    "active": True,
                    "tenantId": creator.get("tenantId"),
                    "permanentCity": None
                },
                "msgId": "1744021633700|en_IN",
                "plainAccessRequest": {}
            },
            "incident": payload,
            "workflow": {"action": "APPLY", "verificationDocuments": []}
        },
        headers={"Content-Type": "application/json"}
    )

def process_response(response, df, idx, identifier):
    if response.status_code in [200, 201]:
        incident = response.json().get("IncidentWrappers", [{}])[0].get("incident")
        df.at[idx, 'status'] = 'success'
        df.at[idx, 'error'] = ''
        df.at[idx, 'ticket_id'] = incident.get("incidentId", '')
    else:
        error_msg = response.json().get('Errors', [{}])[0].get('message', response.text)
        df.at[idx, 'status'] = 'failed'
        df.at[idx, 'error'] = error_msg

def write_and_return_excel(df, sheet_name):
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = f"/tmp/legacy_ticket_ingestion_results_{timestamp}.xlsx"
    df.to_excel(output_path, sheet_name=sheet_name, index=False)
    return FileResponse(output_path, filename=os.path.basename(output_path), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@router.post("/check_duplicates")
async def check_duplicate_tickets(
        legacy_ticket_file: UploadFile = File(...),
        legacy_ticket_sheet_name: str = Form(default="Duplication Template"),
):
    input_temp_file = None
    try:
        # Save uploaded file temporarily
        input_temp_file, _ = await _save_upload_to_temp_file(legacy_ticket_file, suffix=".xlsx")
        excel_path = input_temp_file.name

        # Read Excel file
        df = pd.read_excel(excel_path, sheet_name=legacy_ticket_sheet_name)
        df.columns = df.columns.str.strip()
        df = df.reindex(columns=df.columns.tolist() + ['status', 'error'], fill_value='')

        conn = psycopg2.connect(**DB_CONFIG)
        cursor = conn.cursor()

        # Step 1: Fetch NIN_HFR ID -> tenantId mapping
        nin_hfr_ids = df["NIN_HFR ID"].dropna().astype(str).str.strip().unique().tolist()
        cursor.execute("SELECT code, tenantid FROM eg_hrms_employee WHERE code IN %s", (tuple(nin_hfr_ids),))
        code_tenant_map = dict(cursor.fetchall())

        # Step 2: Check each row for incident duplication
        for idx, row in df.iterrows():
            code = str(row.get("NIN_HFR ID", "")).strip()
            ticket_type = str(row.get("Ticket Type", "")).strip()
            ticket_subtype = str(row.get("Ticket Sub Type", "")).strip()

            tenant_id = code_tenant_map.get(code)
            if not tenant_id:
                df.at[idx, 'status'] = 'error'
                df.at[idx, 'error'] = 'Invalid NIN_HFR ID (not in eg_hrms_employee)'
                continue

            # Step 3: Check for matching incidents
            cursor.execute("""
                SELECT 1 FROM eg_incident_v2
                WHERE tenantid = %s
                AND incidenttype = %s
                AND incidentsubtype = %s
                AND applicationstatus NOT IN ('CLOSEDAFTERRESOLUTION', 'RESOLVED', 'REJECTED')
                LIMIT 1
            """, (tenant_id, ticket_type, ticket_subtype))
            exists = cursor.fetchone()

            if exists:
                df.at[idx, 'status'] = 'duplicate'

        conn.close()

        # Save updated Excel
        df.to_excel(excel_path, index=False, sheet_name=legacy_ticket_sheet_name)

        return FileResponse(
            path=excel_path,
            filename=legacy_ticket_file.filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error during duplicate check: {str(e)}")

    finally:
        if input_temp_file and os.path.exists(input_temp_file.name):
            pass


def get_request_info_to_send_back_workflow():
    return {
        "apiId": "project-api",
        "ver": "1.0",
        "ts": "",
        "action": "update",
        "did": "",
        "key": "",
        "msgId": "20240617",
        "authToken": "f6a27ba4-bead-483d-b4d8-23d46c74d153",
        "userInfo": {
            "id": 178,
            "uuid": "72743f47-9f1a-47de-ac43-b12cde70afc1",
            "userName": "dummy_manager",
            "name": "dummy_manager",
            "mobileNumber": "9911223345",
            "emailId": None,
            "locale": None,
            "type": "EMPLOYEE",
            "roles": [
                {"name": "Installation Report Part A editor", "code": "INSTALLATION_REPORT_PART_A_EDITOR",
                 "tenantId": LIVELIHOOD_TENANT_ID},
                {"name": "Installation Report Part B editor", "code": "INSTALLATION_REPORT_PART_B_EDITOR",
                 "tenantId": LIVELIHOOD_TENANT_ID},
                {"name": "Installation Report Part A reviewer", "code": "INSTALLATION_REPORT_PART_A_REVIEWER",
                 "tenantId": LIVELIHOOD_TENANT_ID},
                {"name": "Project manager", "code": "PROJECT_MANAGER", "tenantId": LIVELIHOOD_TENANT_ID},
                {"name": "Installation Report Approver QC team", "code": "INSTALLATION_REPORT_APPROVER_QC_TEAM",
                 "tenantId": LIVELIHOOD_TENANT_ID}
            ],
            "active": True,
            "tenantId": LIVELIHOOD_TENANT_ID,
            "permanentCity": None
        }
    }

@router.post("/flag_for_qc")
async def flag_for_qc(
        facility_file: UploadFile = File(...),
        facility_sheet_name: str = Form(default="Facilities"),
):
    input_temp_file = None
    try:
        # Save uploaded file temporarily
        input_temp_file, _ = await _save_upload_to_temp_file(facility_file, suffix=".xlsx")
        excel_path = input_temp_file.name

        # Read Excel file
        df = pd.read_excel(excel_path, sheet_name=facility_sheet_name)
        df.columns = df.columns.str.strip()

        # Add system columns for audit/error tracking
        if 'status' not in df.columns:
            df['status'] = ''
        if 'error' not in df.columns:
            df['error'] = ''
        if 'auditTrail' not in df.columns:
            df['auditTrail'] = ''


        for idx, row in df.iterrows():
            business_id = row.get("BusinessId")

            # Prepare request body for workflow API
            payload = {
                "RequestInfo" : get_request_info_to_send_back_workflow(),
                "projectId": business_id,
                "workflow": {
                    "action": "REMOVE_FLAG",
                }
            }

            # Call workflow API
            workflow_update = f"{project_service_url}/project/v1/project/workflow/update"
            try:
                resp = requests.post(workflow_update, json=payload, headers={"Content-Type": "application/json"})
                if resp.status_code != 200:
                    df.at[idx, "error"] = f"WF API failed: {resp.status_code} {resp.text}"
                else:
                    df.at[idx, "auditTrail"] = f"Sent back to pending approval"
            except Exception as e:
                df.at[idx, "error"] = f"WF API call error: {str(e)}"

        df.to_excel(excel_path, index=False, sheet_name=facility_sheet_name)

        return FileResponse(
            path=excel_path,
            filename=facility_file.filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error during facility status update: {str(e)}")

    finally:
        if input_temp_file and os.path.exists(input_temp_file.name):
            pass

@router.post('/incidents/dataUpdate',
             summary='Update incidents data from Excel file',
             response_description='Returns result status for each incident')
async def update_incidents_data_from_excel(
        incidents_file: UploadFile = File(..., description="Excel file containing incidents to update data"),
        incidents_sheet_name: str = Form(default="Incidents",
                                         description="Name of the sheet containing incident data"),
        request_info: str = Form(default="", description="Request info in JSON format")
):
    temp_file = None
    request_info = request_info_from_json(request_info)
    #get_authorized_request_info(request_info)

    try:
        temp_file, _ = await _save_upload_to_temp_file(incidents_file, suffix=".xlsx")

        df = pd.read_excel(temp_file.name, sheet_name=incidents_sheet_name)
        df.columns = df.columns.str.strip()


        for col in ['status', 'error']:
            if col not in df.columns:
                df[col] = ''

        incident_client = IMServiceClient(im_services_url)

        for index, row in df.iterrows():
            if pd.isna(row.get('Ticket No.')):
                df.at[index, 'status'] = 'skipped'
                df.at[index, 'error'] = 'Missing ticket_no'
                continue

            if pd.isna(row.get('Tenant ID')):
                df.at[index, 'status'] = 'skipped'
                df.at[index, 'error'] = 'Missing Tenant ID'
                continue

            incident_request_info = get_incident_request_info()

            try:
                search_response = incident_client.search_incident(
                    incident_id=row['Ticket No.'].strip(),
                    tenant_id=row['Tenant ID'].strip(),
                    request_info=incident_request_info
                )

                incident_wrappers = search_response.get("IncidentWrappers", [])
                if not incident_wrappers:
                    df.at[index, 'status'] = 'failed'
                    df.at[index, 'error'] = f"No incident found for Ticket No. {row['Ticket No.']} and Tenant ID {row['Tenant ID']}"
                    continue

                update_data = {
                    "systemFunctional": (
                        {"yes": "FUNCTIONAL", "no": "NON_FUNCTIONAL"}.get(str(row.get("Is the solar system working?", "")).strip().lower(), "")
                    )
                }

                update_payload = create_update_payload(search_response, update_data)
                update_response = incident_client.update_incident_data(update_payload)

                process_update_incident_data_response(update_response, df, index)

            except Exception as e:
                df.at[index, 'status'] = 'failed'
                df.at[index, 'error'] = str(e)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"incident_data_update_results_{timestamp}.xlsx"

        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as output_temp_file:
            df.to_excel(output_temp_file.name, sheet_name=incidents_sheet_name, index=False)

        return FileResponse(
            path=output_temp_file.name,
            filename=output_filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Failed to process incident updates: {str(e)}"
        ) from e
    finally:
        if temp_file and os.path.exists(temp_file.name):
            os.unlink(temp_file.name)

def process_update_incident_data_response(response, df, idx):
    try:
        if 'Errors' in response and response['Errors']:
            error_msg = response['Errors'][0].get('message', str(response['Errors'][0]))
            df.at[idx, 'status'] = 'failed'
            df.at[idx, 'error'] = error_msg
        else:
            df.at[idx, 'status'] = 'success'
            df.at[idx, 'error'] = ''
    except Exception as e:
        df.at[idx, 'status'] = 'failed'
        df.at[idx, 'error'] = str(e)


@router.post('/facilitiesValidateData',
             summary='Validate facility Excel file before processing',
             response_description='Returns validation report Excel with PASSED/FAILED rows')
async def validate_facilities_excel_sheet(
        background_tasks: BackgroundTasks,
        facility_file: UploadFile = File(..., description="Excel file containing facility data"),
        project_id: str = Form(description="Project ID"),
        facility_sheet_name: str = Form(default="FacilityMapping",
                                        description="Name of the sheet containing facility data"),
        boundary_sheet_name: str = Form(default="BoundaryCodes",
                                        description="Name of the sheet containing boundary data"),
        request_info: str = Form(default="")
):
    temp_input_file = None
    request_info_obj = request_info_from_json(request_info)
    mdms_client = MDMSClient(mdms_url)
    facility_client = FacilityServiceClient(facility_service_url)
    project_client = ProjectServiceClient(project_service_url)

    try:
        # Save uploaded Excel to a temp file
        temp_input_file, _ = await _save_upload_to_temp_file(facility_file, suffix=".xlsx")

        # Load workbook to preserve everything
        wb = load_workbook(temp_input_file.name)

        # ----------------- Read Boundary Sheet ----------------- #
        if boundary_sheet_name not in wb.sheetnames:
            raise HTTPException(status_code=400, detail=f"Boundary sheet '{boundary_sheet_name}' not found")

        boundary_data_df = pd.read_excel(temp_input_file.name, sheet_name=boundary_sheet_name)

        # ----------------- Validate Boundary Sheet Against Project ----------------- #
        projects = project_client.search_project(request_info_obj, project_id)
        if not projects or "Project" not in projects or len(projects["Project"]) == 0:
            raise HTTPException(status_code=400, detail=f"No project found for id {project_id}")

        project = projects["Project"][0]["project"]
        geography = project.get("additionalDetails", {}).get("geographyDetails", {})

        # Valid codes directly as a set (no loop needed)
        valid_boundary_codes = {str(block["code"]).strip() for block in geography.get("blocks", []) if
                                block.get("code")}

        # Uploaded codes directly as a set
        uploaded_codes = set(boundary_data_df["BoundaryCode"].dropna().astype(str).str.strip())

        # Equality check
        if uploaded_codes != valid_boundary_codes:
            raise HTTPException(
                status_code=400,
                detail={
                    "error": "BoundaryCode mismatch",
                    "missing": list(valid_boundary_codes - uploaded_codes),
                    "extra": list(uploaded_codes - valid_boundary_codes)
                }
            )

        # ----------------- Read Facility Sheet ----------------- #
        if facility_sheet_name not in wb.sheetnames:
            raise HTTPException(status_code=400, detail=f"Facility sheet '{facility_sheet_name}' not found")

        df = pd.read_excel(temp_input_file.name, sheet_name=facility_sheet_name)
        df.columns = [str(c).strip() for c in df.columns]
        df = df.loc[:, ~df.columns.str.startswith('Unnamed')]

        # ----------------- Read Facility Column ----------------- #
        if 'End User Id' not in df.columns:
            raise HTTPException(status_code=400, detail=f"'End User Id' column in '{facility_sheet_name}' not found")

        # Ensure status/error columns exist
        if 'status' not in df.columns:
            df['status'] = ''
        if 'error' not in df.columns:
            df['error'] = ''

        # ----------------- Run Validation ----------------- #
        validation_errors = project_facility_validation(
            df,
            mdms_client,
            request_info_obj,
            facility_client,
            boundary_data_df,
            'data-ingestion.FacilityIngestionSchema',
            localization_service_url,
        )

        # Mark rows based on validation results
        error_count = 0
        for i, errs in enumerate(validation_errors):
            if errs:
                df.at[i, 'status'] = 'FAILED'
                df.at[i, 'error'] = "; ".join(dict.fromkeys(errs))
                error_count += 1
            else:
                df.at[i, 'status'] = 'PASSED'
                df.at[i, 'error'] = ''

        # ----------------- Update Facility Sheet In-Place ----------------- #
        ws = wb[facility_sheet_name]
        header_values = [cell.value for cell in ws[1]]

        # Add status/error columns if missing
        for col_name in ["status", "error"]:
            if col_name not in header_values:
                new_col_idx = len(header_values) + 1
                cell = ws.cell(row=1, column=new_col_idx, value=col_name)
                cell.font = Font(bold=True)
                header_values.append(col_name)

                # lock header cell
                cell.protection = Protection(locked=True)

                # lock all data cells in this new column
                for r_idx in range(2, ws.max_row + 1):
                    ws.cell(row=r_idx, column=new_col_idx).protection = Protection(locked=True)

        grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        # Write data rows back (without header row)
        export_df = prepare_dataframe_for_excel_export(df)
        for r_idx, row in enumerate(dataframe_to_rows(export_df, index=False, header=False), start=2):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)

                # force lock for status/error columns
                if ws.cell(1, c_idx).value in ["status", "error"]:
                    cell.protection = Protection(locked=True)
                    cell.fill = grey_fill

        # Ensure sheet protection is ON
        ws.protection.sheet = True
        ws.protection.enable()

        # ----------------- Save to new temp file ----------------- #
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_temp_file_path = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx").name
        wb.save(output_temp_file_path)

        autofit_columns(output_temp_file_path, facility_sheet_name, auto_fit=True)

        background_tasks.add_task(cleanup_temp_file, output_temp_file_path)

        response = FileResponse(
            path=output_temp_file_path,
            filename=f"facility_validation_results_{timestamp}.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response.headers["X-Error-Count"] = str(error_count)

        return response

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Validation failed: {str(e)}")
    finally:
        if temp_input_file and os.path.exists(temp_input_file.name):
            os.unlink(temp_input_file.name)


def _state_by_boundary_code(boundary_data_df) -> dict:
    """{block boundary code: state name} from the workbook's BoundaryCodes sheet, which the
    download wrote with the same localized names it put in the State column."""
    if boundary_data_df is None:
        return {}
    columns = set(boundary_data_df.columns)
    if not {"BoundaryCode", "State"} <= columns:
        logger.warning("BoundaryCodes sheet has no BoundaryCode/State columns; cannot resolve states")
        return {}
    states = {}
    for _, boundary_row in boundary_data_df.iterrows():
        code = "" if pd.isna(boundary_row.get("BoundaryCode")) else str(boundary_row.get("BoundaryCode")).strip()
        state = "" if pd.isna(boundary_row.get("State")) else str(boundary_row.get("State")).strip()
        if code:
            states[code] = state
    return states


def _facility_states_from_sheet(df, boundary_data_df, facility_client, request_info) -> dict:
    """facility_id -> state name for the sites listed in the sheet.

    A facility has no state of its own: FacilityAddress declares state/district/block but
    facility_address has no such columns, so address.state is always null. The state lives
    only in boundary_code, so it is resolved from the workbook's BoundaryCodes sheet -- the
    same values the download used to fill the State column and to build the Solution
    dropdown. Keying off anything else would reject rows whose Solution the dropdown had
    just offered.

    The state is taken from the facility record's boundary_code rather than the row's State
    cell because the cell is editable once the sheet is unprotected. Returns {} on failure,
    which makes eligibility fail closed rather than wave rows through.
    """
    site_id_column = find_site_id_column(df)
    if not site_id_column or facility_client is None:
        return {}
    facility_ids = [
        str(v).strip() for v in df[site_id_column] if pd.notna(v) and str(v).strip()
    ]
    if not facility_ids:
        return {}
    state_by_boundary_code = _state_by_boundary_code(boundary_data_df)
    if not state_by_boundary_code:
        return {}
    try:
        result = facility_client.bulk_search_facility(
            request_info=request_info,
            tenant_ids=[LIVELIHOOD_TENANT_ID],
            facility_ids=facility_ids,
            limit=max(len(facility_ids), 50),
            send_non_paginated_response=True,
        )
    except Exception as e:
        logger.error(f"Could not resolve facility states for eligibility: {e}", exc_info=True)
        return {}

    states = {}
    for facility in (result.get("facilities") or []):
        facility_id = facility.get("facility_id")
        if not facility_id:
            continue
        boundary_code = facility.get("boundary_code") or facility.get("boundaryCode") or ""
        # A facility's code is its block's code, optionally suffixed with a facility-specific
        # segment (e.g. INDIA_ASSAM_BAKSA_BORABARI_ED/2026/0093) -- the same match the
        # download's resolve_boundary_names_for_code makes.
        states[facility_id] = next(
            (
                state for code, state in state_by_boundary_code.items()
                if boundary_code == code or boundary_code.startswith(code + "_")
            ),
            "",
        )
    return states


@router.post('/fieldPlanfacilitiesValidateData',
             summary='Validate facility Excel file before processing',
             response_description='Returns validation report Excel with PASSED/FAILED rows')
async def validate_facilities_excel_sheet(
        background_tasks: BackgroundTasks,
        facility_file: UploadFile = File(..., description="Excel file containing facility data"),
        facility_sheet_name: str = Form(default="FacilityMapping",
                                        description="Name of the sheet containing facility data"),
        boundary_sheet_name: str = Form(default="BoundaryCodes",
                                        description="Name of the sheet containing boundary data"),
        fieldplan_id: str = Form(default="",
                                 description="Field plan id; when given, its sector is used instead of the sheet's"),
        request_info: str = Form(default="")
):
    temp_input_file = None
    request_info_obj = request_info_from_json(request_info)
    mdms_client = MDMSClient(mdms_url)
    facility_client = FacilityServiceClient(facility_service_url)

    try:
        # Save uploaded Excel to a temp file
        temp_input_file, _ = await _save_upload_to_temp_file(facility_file, suffix=".xlsx")

        # Load workbook to preserve everything
        wb = load_workbook(temp_input_file.name)

        # ----------------- Read Boundary Sheet ----------------- #
        if boundary_sheet_name not in wb.sheetnames:
            raise HTTPException(status_code=400, detail=f"Boundary sheet '{boundary_sheet_name}' not found")

        boundary_data_df = pd.read_excel(temp_input_file.name, sheet_name=boundary_sheet_name)

        # ----------------- Read Facility Sheet ----------------- #
        if facility_sheet_name not in wb.sheetnames:
            raise HTTPException(status_code=400, detail=f"Facility sheet '{facility_sheet_name}' not found")

        df = pd.read_excel(temp_input_file.name, sheet_name=facility_sheet_name)
        df.columns = [str(c).strip() for c in df.columns]
        df = df.loc[:, ~df.columns.str.startswith('Unnamed')]

        # ----------------- Read Facility Column ----------------- #
        if not find_site_id_column(df):
            raise HTTPException(
                status_code=400,
                detail=f"Sheet '{facility_sheet_name}' has no site id column "
                       f"(expected one of {', '.join(SITE_ID_COLUMNS)}).",
            )

        # Ensure status/error columns exist
        if 'status' not in df.columns:
            df['status'] = ''
        if 'error' not in df.columns:
            df['error'] = ''

        # ----------------- Run Validation ----------------- #
        # Every row here is an existing site being linked to a plan, so all rows are
        # validated rather than only the id-less "new facility" rows.
        validation_errors = project_facility_validation(
            df,
            mdms_client,
            request_info_obj,
            facility_client,
            boundary_data_df,
            'data-ingestion.InstallationScopeIngestionSchema',
            validate_all_rows=True
        )

        # The plan is what makes the rest of this validation meaningful: it supplies the
        # sector and the project whose locks are enforced below. Failing to read it must
        # not be a warning -- with project_id unset the lock map comes back empty and every
        # locked row silently validates as editable, so the sheet would be reported PASSED
        # having skipped the check it most needed.
        plan_sector = None
        project_id = None
        if fieldplan_id and fieldPlan_service_url:
            try:
                field_plans = FieldPlanServiceClient(fieldPlan_service_url).search_fieldPlan(
                    request_info_obj, fieldplan_id
                ).get("FieldPlans", [])
            except Exception as e:
                logger.error(f"Error fetching field plan {fieldplan_id}: {e}", exc_info=True)
                raise HTTPException(
                    status_code=502,
                    detail=f"Could not read field plan {fieldplan_id} from field-planner, "
                           f"so lock rules cannot be enforced: {e}",
                )
            if not field_plans:
                raise HTTPException(status_code=404, detail=f"Field plan {fieldplan_id} not found")
            plan_sector = field_plans[0].get("sector")
            project_id = field_plans[0].get("projectId")

        # Sites already under installation anywhere in this project cannot be re-scoped.
        lock_map = {}
        if project_id and fieldPlan_service_url:
            lock_map = build_project_lock_map(
                FieldPlanServiceClient(fieldPlan_service_url), request_info_obj, project_id, fieldplan_id
            )

        solutions = mdms_client.fetch_installation_solutions(request_info_obj)
        linkable_rows = validate_installation_scope_solutions(
            df,
            solutions=solutions,
            sunshine_hours_by_state=fetch_state_sunshine_hours(),
            add_err=lambda i, msg: validation_errors[i].append(msg),
            plan_sector=plan_sector,
            state_by_facility_id=_facility_states_from_sheet(
                df, boundary_data_df, facility_client, request_info_obj
            ),
            lock_map=lock_map,
            solution_name_by_code=solution_names_by_code(solutions),
        )

        # A sheet that selects nothing is a mistake, not a no-op. Frozen rows carry
        # Include=Yes for display, so they must not count towards "something was selected".
        if not linkable_rows:
            raise HTTPException(
                status_code=400,
                detail="No end user sites are selected for this installation plan. "
                       "Mark at least one site as included and choose its Solution.",
            )

        # Mark rows based on validation results
        error_count = 0
        for i, errs in enumerate(validation_errors):
            if errs:
                df.at[i, 'status'] = 'FAILED'
                df.at[i, 'error'] = "; ".join(dict.fromkeys(errs))
                error_count += 1
            else:
                df.at[i, 'status'] = 'PASSED'
                df.at[i, 'error'] = ''

        # ----------------- Update Facility Sheet In-Place ----------------- #
        ws = wb[facility_sheet_name]
        header_values = [cell.value for cell in ws[1]]

        # Add status/error columns if missing
        for col_name in ["status", "error"]:
            if col_name not in header_values:
                new_col_idx = len(header_values) + 1
                cell = ws.cell(row=1, column=new_col_idx, value=col_name)
                cell.font = Font(bold=True)
                header_values.append(col_name)

                # lock header cell
                cell.protection = Protection(locked=True)

                # lock all data cells in this new column
                for r_idx in range(2, ws.max_row + 1):
                    ws.cell(row=r_idx, column=new_col_idx).protection = Protection(locked=True)

        grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        # Write data rows back (without header row)
        export_df = prepare_dataframe_for_excel_export(df)
        for r_idx, row in enumerate(dataframe_to_rows(export_df, index=False, header=False), start=2):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)

                # force lock for status/error columns
                if ws.cell(1, c_idx).value in ["status", "error"]:
                    cell.protection = Protection(locked=True)
                    cell.fill = grey_fill

        # Ensure sheet protection is ON
        ws.protection.sheet = True
        ws.protection.enable()

        # ----------------- Save to new temp file ----------------- #
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_temp_file_path = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx").name
        wb.save(output_temp_file_path)

        autofit_columns(output_temp_file_path, facility_sheet_name, auto_fit=True)

        background_tasks.add_task(cleanup_temp_file, output_temp_file_path)

        response = FileResponse(
            path=output_temp_file_path,
            filename=f"facility_validation_results_{timestamp}.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response.headers["X-Error-Count"] = str(error_count)

        return response

    except HTTPException:
        # Deliberate 4xx (e.g. "no sites selected") must reach the caller unchanged --
        # the generic handler below would otherwise relabel it as a 500.
        raise
    except Exception as e:
        logger.error(f"Unhandled error validating field plan facilities: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Validation failed: {str(e)}")
    finally:
        if temp_input_file and os.path.exists(temp_input_file.name):
            os.unlink(temp_input_file.name)


@router.post('/createFacilityAndUpdateProject',
             summary='Create passed facility in Excel file and add them to project',
             response_description='Created facilities from PASSED rows and added to the given project if selected')
async def create_facilities_and_update_project(
        background_tasks: BackgroundTasks,
        facility_file: UploadFile = File(description="Validated Excel file with PASSED/FAILED status"),
        facility_sheet_name: str = Form(default="FacilityMapping",
                                        description="Name of the sheet containing facility data"),
        project_id: str = Form(description="Project ID"),
        request_info: str = Form(default="")
):
    input_temp_file = None

    # parse
    request_info = request_info_from_json(request_info)

    try:
        # ---------- save uploaded file ----------
        input_temp_file, uploaded_size = await _save_upload_to_temp_file(facility_file, suffix=".xlsx")
        facility_file_path = input_temp_file.name
        logger.info(f"Received createFacilityAndUpdateProject file of size {uploaded_size} bytes")

        # ---------- prepare output path & load workbook ----------
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"facility_creation_and_project_update_results_{timestamp}.xlsx"
        output_temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        output_temp_file.close()
        output_file_path = output_temp_file.name

        wb = load_workbook(facility_file_path)
        if facility_sheet_name not in wb.sheetnames:
            raise HTTPException(status_code=400, detail=f"Sheet '{facility_sheet_name}' not found")
        ws = wb[facility_sheet_name]

        # ---------- read sheet into DataFrame ----------
        df = pd.read_excel(facility_file_path, sheet_name=facility_sheet_name)
        df.columns = [str(c).strip() for c in df.columns]

        # sanity checks
        # status/error are validation artifacts from previous step
        if 'status' not in df.columns or 'error' not in df.columns:
            raise HTTPException(status_code=400, detail="Missing 'status'/'error' columns. Please upload validated file.")

        # Ensure all rows are PASSED
        failed_rows = df[df['status'].str.upper() != 'PASSED']
        if not failed_rows.empty:
            raise HTTPException(
                status_code=400,
                detail="Validation failed: Some rows are not marked as PASSED. Please upload a fully validated file."
            )

        # helper to find a column by partial name (case insensitive)
        def find_col(partial):
            for c in df.columns:
                if partial.lower() in str(c).lower():
                    return c
            return None

        include_col = find_col("Include in Project")
        facility_id_col = find_col("End User Id") or "End User Id"
        status_col = find_col("status") or "status"

        # add result columns if missing
        if 'Facility Creation Status' not in df.columns:
            df['Facility Creation Status'] = ''
        if 'Project Linking Status' not in df.columns:
            df['Project Linking Status'] = ''

        project_client = ProjectServiceClient(project_service_url)

        # --- NEW: fetch already linked facilities once ---
        linked_facilities_resp = project_client.search_project_facility(request_info, project_id)
        linked_facilities = linked_facilities_resp.get("ProjectFacilities", []) if linked_facilities_resp else []
        linked_facility_ids = {pf.get("facilityId") for pf in linked_facilities if pf.get("facilityId")}

        # This template only links existing facilities (selected via the facility ingestion
        # template) to the project; it does not create new facilities. Rows without an
        # End User Id are invalid and are skipped rather than creating a new facility.
        pending_bulk_links = []
        existing_or_skipped_indexes = []
        for index, row in df.iterrows():
            include_val = ''
            if include_col:
                include_val = str(row.get(include_col, "")).strip().lower()
            else:
                include_val = str(row.get("Include in Project (Mandatory)", "")).strip().lower()
            should_link = include_val == "yes"

            facility_id_val = row.get(facility_id_col, None)
            facility_id = str(facility_id_val).strip() if pd.notna(facility_id_val) and str(facility_id_val).strip() else None
            row_status = str(row.get(status_col, "")).strip().upper()

            if facility_id:
                existing_or_skipped_indexes.append((index, row, should_link, facility_id))
            elif row_status != "PASSED":
                df.at[index, 'Facility Creation Status'] = "Skipped (Validation not PASSED)"
                df.at[index, 'Project Linking Status'] = "Not Attempted"
            else:
                df.at[index, 'Facility Creation Status'] = "Skipped (End User Id is required; new facility creation is not supported)"
                df.at[index, 'Project Linking Status'] = "Not Attempted"

        for index, row, should_link, facility_id in existing_or_skipped_indexes:
            try:
                df.at[index, 'Facility Creation Status'] = "Already Exists"
                if facility_id in linked_facility_ids:
                    if should_link:
                        df.at[index, 'Project Linking Status'] = "Already Linked"
                    else:
                        try:
                            project_facility_data = next((pf for pf in linked_facilities if pf.get("facilityId") == facility_id), None)
                            project_client.unlink_project_facility(
                                request_info=request_info,
                                project_id=project_id,
                                facility_id=facility_id,
                                project_facility_data=project_facility_data
                            )
                            df.at[index, 'Project Linking Status'] = "Unlinked"
                            linked_facility_ids.remove(facility_id)
                        except Exception as e:
                            df.at[index, 'Project Linking Status'] = f"Exception during unlink: {str(e)}"
                else:
                    if should_link:
                        pending_bulk_links.append((index, facility_id))
                    else:
                        df.at[index, 'Project Linking Status'] = "Skipped (Include in Project != Yes)"
            except Exception as e:
                df.at[index, 'Facility Creation Status'] = f"Exception: {str(e)}"
                df.at[index, 'Project Linking Status'] = "Not Attempted"
                continue

        # Bulk-link facilities to project (for include=yes rows not already linked)
        if pending_bulk_links:
            chunk_size = BULK_INGEST_CHUNK_SIZE
            for i in range(0, len(pending_bulk_links), chunk_size):
                chunk = pending_bulk_links[i:i + chunk_size]
                chunk_facility_ids = [facility_id for _, facility_id in chunk]
                try:
                    bulk_resp = project_client.create_project_facility_bulk(
                        request_info=request_info,
                        project_id=project_id,
                        facility_ids=chunk_facility_ids
                    )
                    if bulk_resp.status_code in (200, 201, 202):
                        for row_idx, facility_id in chunk:
                            df.at[row_idx, 'Project Linking Status'] = "Linked"
                            linked_facility_ids.add(facility_id)
                    else:
                        for row_idx, _ in chunk:
                            df.at[row_idx, 'Project Linking Status'] = f"Failed: {bulk_resp.status_code} {bulk_resp.text}"
                except Exception as exc:
                    for row_idx, _ in chunk:
                        df.at[row_idx, 'Project Linking Status'] = f"Exception: {str(exc)}"

        # ---------- write results back into workbook preserving formatting ----------
        # Ensure headers exist in sheet (without wiping template)
        header_values = [cell.value for cell in ws[1]]

        for col_name in ["Facility Creation Status", "Project Linking Status"]:
            if col_name not in header_values:
                cell = ws.cell(row=1, column=len(header_values) + 1, value=col_name)
                cell.font = Font(bold=True)
                header_values.append(col_name)
                if col_name not in df.columns:
                    df[col_name] = ""  # ensure column exists in dataframe

        # Delete data rows only (preserve header row and template formatting)
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)

        # Write data rows back (without header row)
        export_df = prepare_dataframe_for_excel_export(df)
        for r_idx, row in enumerate(dataframe_to_rows(export_df, index=False, header=False), start=2):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        wb.save(output_file_path)

        autofit_columns(output_file_path, facility_sheet_name , auto_fit=True)

        background_tasks.add_task(cleanup_temp_file, output_file_path)

        return FileResponse(
            path=output_file_path,
            filename=output_filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        logger.error(f"Error finalizing facility data: {e}")
        raise HTTPException(
            status_code=500,
            detail=f"Failed to finalize facility data: {str(e)}"
        )
    finally:
        if input_temp_file and os.path.exists(input_temp_file.name):
            os.unlink(input_temp_file.name)


@router.post('/createFieldPlanFacility',
             summary='Create passed facility in Excel file and add them to project',
             response_description='Created facilities from PASSED rows and added to the given project if selected')
async def create_fielplan_facilities(
        background_tasks: BackgroundTasks,
        facility_file: UploadFile = File(description="Validated Excel file with PASSED/FAILED status"),
        facility_sheet_name: str = Form(default="FacilityMapping",
                                        description="Name of the sheet containing facility data"),
        fieldplan_id: str = Form(description="FieldPlan ID"),
        request_info: str = Form(default="")
):
    input_temp_file = None

    # parse
    request_info = request_info_from_json(request_info)
    mdms_client = MDMSClient(mdms_url)

    try:
        # ---------- save uploaded file ----------
        input_temp_file, uploaded_size = await _save_upload_to_temp_file(facility_file, suffix=".xlsx")
        facility_file_path = input_temp_file.name
        logger.info(f"Received createFieldPlanFacility file of size {uploaded_size} bytes")

        # ---------- prepare output path & load workbook ----------
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"facility_fieldplan_update_results_{timestamp}.xlsx"
        output_temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        output_temp_file.close()
        output_file_path = output_temp_file.name

        wb = load_workbook(facility_file_path)
        if facility_sheet_name not in wb.sheetnames:
            raise HTTPException(status_code=400, detail=f"Sheet '{facility_sheet_name}' not found")
        ws = wb[facility_sheet_name]

        # ---------- read sheet into DataFrame ----------
        df = pd.read_excel(facility_file_path, sheet_name=facility_sheet_name)
        df.columns = [str(c).strip() for c in df.columns]

        # sanity checks
        # status/error are validation artifacts from previous step
        if 'status' not in df.columns or 'error' not in df.columns:
            raise HTTPException(status_code=400, detail="Missing 'status'/'error' columns. Please upload validated file.")

        # Ensure all rows are PASSED
        failed_rows = df[df['status'].str.upper() != 'PASSED']
        if not failed_rows.empty:
            raise HTTPException(
                status_code=400,
                detail="Validation failed: Some rows are not marked as PASSED. Please upload a fully validated file."
            )

        # helper to find a column by partial name (case insensitive)
        def find_col(partial):
            for c in df.columns:
                if partial.lower() in str(c).lower():
                    return c
            return None

        include_col = find_col("Included in Field Plan")
        facility_id_col = find_site_id_column(df)
        if not facility_id_col:
            raise HTTPException(
                status_code=400,
                detail=f"Sheet '{facility_sheet_name}' has no site id column "
                       f"(expected one of {', '.join(SITE_ID_COLUMNS)}).",
            )
        status_col = find_col("status") or "status"

        # add result columns if missing
        if 'Field Plan Linking Status' not in df.columns:
            df['Field Plan Linking Status'] = ''

        fieldplan_client = FieldPlanServiceClient(fieldPlan_service_url)
        fieldplan_activity_client = FieldPlanActivityServiceClient(fieldPlan_activity_service_url)

        # Fetch fieldplan-linked facilities if fieldplan_id is provided
        fieldplan_linked_facility_ids = set()
        if fieldplan_id:
            try:
                fieldplan_facilities_response = fieldplan_client.search_fieldplan_facility(request_info, fieldplan_id)
                fieldplan_facilities = fieldplan_facilities_response.get("FieldPlanFacilities", [])
                fieldplan_linked_facility_ids = {pf.get("facilityId") for pf in fieldplan_facilities if
                                                 pf.get("facilityId")}
                logger.info(
                    f"Found {len(fieldplan_linked_facility_ids)} facilities linked to fieldplan {fieldplan_id}")

                # Get FieldPlan status
                fieldplan_response = fieldplan_client.search_fieldPlan(request_info, fieldplan_id)
                fieldplan_data = fieldplan_response.get("FieldPlans", [])

                fieldplan_assignment_response = fieldplan_activity_client.search_fieldplan_activity_assignment(request_info, fieldplan_id)
                fieldplan_assignment_data = fieldplan_assignment_response.get("ActivitiesAssignments", [])
                role_to_ids = defaultdict(list)

                for item in fieldplan_assignment_data:
                    role = item.get("role")
                    if role:
                        code = role.get("code")
                        if code:
                            role_to_ids[code].append(item.get("assignedTo"))

                # The sheet shows Solution names; solution_id stores the MDMS code.
                solution_col = find_col("Solution")
                solution_code_by_name = solution_codes_by_name(
                    mdms_client.fetch_installation_solutions(request_info)
                )
                # Sites under installation elsewhere in this project are shown for context
                # only -- linking them here would put one site in two plans (FR-06).
                project_id = fieldplan_data[0].get("projectId") if fieldplan_data else None
                lock_map = build_project_lock_map(
                    fieldplan_client, request_info, project_id, fieldplan_id
                ) if project_id else {}

                pending_bulk_fieldplan_links = []
                # iterate all rows — handle existing facility ids (linking/unlinking)
                for index, row in df.iterrows():
                    try:
                        # normalize facility id and include flag
                        facility_id_val = row.get(facility_id_col, None)
                        facility_id = None
                        if pd.notna(facility_id_val) and str(facility_id_val).strip():
                            facility_id = str(facility_id_val).strip()

                        include_val = ''
                        if include_col:
                            include_val = str(row.get(include_col, "")).strip().lower()
                        else:
                            include_val = str(row.get("Included in Field Plan (Mandatory)", "")).strip().lower()

                        should_link = include_val == "yes"

                        lock = lock_map.get(facility_id) if facility_id else None
                        if lock is not None:
                            df.at[index, 'Field Plan Linking Status'] = (
                                "Locked (this plan)" if lock.is_this_plan
                                else f"Locked by another plan ({lock.field_plan_name or lock.field_plan_id})"
                            )
                            continue

                        solution_name = ""
                        if solution_col:
                            raw_solution = row.get(solution_col, "")
                            solution_name = "" if pd.isna(raw_solution) else str(raw_solution).strip()
                        solution_code = solution_code_by_name.get(solution_name) if solution_name else None
                        if solution_name and not solution_code:
                            df.at[index, 'Field Plan Linking Status'] = f"Unknown Solution '{solution_name}'"
                            continue

                        # ---------- CASE A: existing facility_id present -> skip creation, attempt linking if requested ----------
                        if facility_id:
                            # df.at[index, 'Facility Creation Status'] = "Already Exists"
                            # attempt linking if requested
                            if facility_id in fieldplan_linked_facility_ids:
                                if should_link:
                                    # already linked → skip API
                                    df.at[index, 'Field Plan Linking Status'] = "Already Linked"
                                else:
                                    # linked but Excel says No → unlink
                                    try:
                                        fieldPlan_facility_data = next(
                                            (pf for pf in fieldplan_facilities if pf.get("facilityId") == facility_id),
                                            None)
                                        fieldplan_client.unlink_fieldplan_facility(
                                            request_info=request_info,
                                            fieldplan_id=fieldplan_id,
                                            facility_id=facility_id,
                                            fieldplan_facility_data=fieldPlan_facility_data
                                        )

                                        facilities_activity_response = fieldplan_activity_client.search_facility_activity(
                                            request_info, fieldplan_id, facility_id)
                                        facilities_activity = facilities_activity_response.get("FacilityActivities",[])
                                        facility_activity_ids = list({fa.get("activityFacility").get("id") for fa in facilities_activity if fa.get("activityFacility").get("id")})
                                        fieldplan_activity_client.delete_facility_activity(request_info=request_info, facility_activity_id=facility_activity_ids)

                                        df.at[index, 'Field Plan Linking Status'] = "Unlinked"
                                        fieldplan_linked_facility_ids.remove(facility_id)
                                    except Exception as e:
                                        df.at[index, 'Field Plan Linking Status'] = f"Exception during unlink: {str(e)}"
                            else:
                                if should_link:
                                    pending_bulk_fieldplan_links.append((index, facility_id, solution_code))
                                else:
                                    df.at[index, 'Field Plan Linking Status'] = "Skipped (Include in Field Plan != Yes)"

                                # continue to next row
                                continue

                    except Exception as e:
                        # any unexpected error per row
                        df.at[index, 'Field Plan Linking Status'] = "Not Attempted"
                        continue

                if pending_bulk_fieldplan_links:
                    chunk_size = BULK_INGEST_CHUNK_SIZE
                    for i in range(0, len(pending_bulk_fieldplan_links), chunk_size):
                        chunk = pending_bulk_fieldplan_links[i:i + chunk_size]
                        solution_by_facility = {
                            facility_id: solution_code for _, facility_id, solution_code in chunk
                        }
                        try:
                            fieldplan_resp = fieldplan_client.create_fieldPlan_facility_bulk(
                                request_info=request_info,
                                fieldPlan_id=fieldplan_id,
                                facility_ids=list(solution_by_facility),
                                solution_id_by_facility=solution_by_facility
                            )

                            if fieldplan_resp.status_code in (200, 201, 202):
                                for row_idx, facility_id, _solution_code in chunk:
                                    df.at[row_idx, 'Field Plan Linking Status'] = "Linked"
                                    fieldplan_linked_facility_ids.add(facility_id)

                                    if fieldplan_data:
                                        fieldplan = fieldplan_data[0]
                                        if fieldplan.get("status") == 'SCHEDULED':
                                            try:
                                                facility_activity_resp = fieldplan_activity_client.create_facility_activity(
                                                    request_info=request_info,
                                                    fieldPlan=fieldplan,
                                                    roleToIds=role_to_ids,
                                                    facility_id=facility_id
                                                )
                                                logger.info(f"Facility activity created successfully for facility {facility_id}")
                                                logger.debug(f"Facility activity response: {facility_activity_resp}")
                                            except Exception as activity_exc:
                                                logger.error(f"Error creating facility activity for {facility_id}: {activity_exc}", exc_info=True)
                            else:
                                for row_idx, _facility_id, _solution_code in chunk:
                                    df.at[row_idx, 'Field Plan Linking Status'] = f"Failed: {fieldplan_resp.status_code} {fieldplan_resp.text}"
                        except Exception as bulk_exc:
                            for row_idx, _facility_id, _solution_code in chunk:
                                df.at[row_idx, 'Field Plan Linking Status'] = f"Exception: {str(bulk_exc)}"

                    # Confirm the rows actually landed before telling the Project Manager they did.
                    #
                    # "Linked" above is set on an HTTP 202, which only means the message was
                    # queued: /facility/bulk/_create pushes to a bulk topic, a consumer picks it
                    # up and republishes to save-fieldplan-facility-topic, and egov-persister
                    # finally writes the row. Nothing in that chain can report failure back here
                    # -- the consumer catches and logs, and the persister runs in another pod.
                    # This is exactly how the Installation Scope step once reported every site
                    # "Linked" while field_plan_facilities stayed empty.
                    expected_ids = {fid for _, fid, _ in pending_bulk_fieldplan_links}
                    confirmed_ids = set()
                    for attempt in range(SCOPE_LINK_CONFIRM_ATTEMPTS):
                        try:
                            stored = fieldplan_client.search_fieldplan_facility(
                                request_info, fieldplan_id).get("FieldPlanFacilities", [])
                            confirmed_ids = {
                                link.get("facilityId") for link in stored
                                if link.get("facilityId") and not link.get("isdeleted")
                            }
                        except Exception as confirm_exc:
                            logger.warning(
                                f"scope read-back attempt {attempt + 1} failed: {confirm_exc}")
                        if expected_ids <= confirmed_ids:
                            break
                        time.sleep(SCOPE_LINK_CONFIRM_INTERVAL_SECONDS)

                    unconfirmed = expected_ids - confirmed_ids
                    if unconfirmed:
                        logger.error(
                            f"{len(unconfirmed)} of {len(expected_ids)} scope row(s) were not "
                            f"visible for plan {fieldplan_id} after "
                            f"{SCOPE_LINK_CONFIRM_ATTEMPTS} attempts: {sorted(unconfirmed)}")
                        for row_idx, facility_id, _solution_code in pending_bulk_fieldplan_links:
                            # Only downgrade rows we optimistically marked Linked -- a row that
                            # already says Failed/Exception has a more specific cause.
                            if (facility_id in unconfirmed
                                    and df.at[row_idx, 'Field Plan Linking Status'] == "Linked"):
                                df.at[row_idx, 'Field Plan Linking Status'] = (
                                    "Pending: accepted but not yet saved. Re-upload this sheet "
                                    "to confirm before moving to the Template step.")
                                fieldplan_linked_facility_ids.discard(facility_id)
                    else:
                        logger.info(
                            f"confirmed {len(expected_ids)} scope row(s) persisted for plan "
                            f"{fieldplan_id}")

            except Exception as e:
                # This block does the linking, not just the fetch: swallowing it returns a
                # 200 with an empty "Field Plan Linking Status" column and nothing written,
                # which reads as "no sites matched" rather than as a failure. Keep the
                # response shape, but make the cause traceable and say so in the sheet.
                logger.error(f"Field plan linking failed for {fieldplan_id}: {e}", exc_info=True)
                df['Field Plan Linking Status'] = df['Field Plan Linking Status'].replace(
                    "", f"Not attempted: {type(e).__name__}: {e}"
                )

        # ---------- write results back into workbook preserving formatting ----------
        # Ensure headers exist in sheet (without wiping template)
        header_values = [cell.value for cell in ws[1]]

        for col_name in ["Field Plan Linking Status"]:
            if col_name not in header_values:
                cell = ws.cell(row=1, column=len(header_values) + 1, value=col_name)
                cell.font = Font(bold=True)
                header_values.append(col_name)
                if col_name not in df.columns:
                    df[col_name] = ""  # ensure column exists in dataframe

        # Delete data rows only (preserve header row and template formatting)
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)

        # Write data rows back (without header row)
        export_df = prepare_dataframe_for_excel_export(df)
        for r_idx, row in enumerate(dataframe_to_rows(export_df, index=False, header=False), start=2):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        wb.save(output_file_path)

        autofit_columns(output_file_path, facility_sheet_name , auto_fit=True)

        background_tasks.add_task(cleanup_temp_file, output_file_path)

        return FileResponse(
            path=output_file_path,
            filename=output_filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except HTTPException:
        # Deliberate 4xx (wrong sheet, unvalidated file, rows still FAILED) must reach the
        # caller as itself -- the generic handler below would relabel it a 500.
        raise
    except Exception as e:
        logger.error(f"Error finalizing facility data: {e}", exc_info=True)
        raise HTTPException(
            status_code=500,
            detail=f"Failed to finalize facility data: {str(e)}"
        )
    finally:
        if input_temp_file and os.path.exists(input_temp_file.name):
            os.unlink(input_temp_file.name)


def _load_and_parse_template(temp_path: str):
    """Open an uploaded template and pull out its two BOM sections.

    Returns (workbook, sheet, parsed). Raises HTTPException(400) when the file is not a
    template at all, which is a sheet-level problem with no row to annotate.
    """
    try:
        workbook = load_workbook(temp_path)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not read the uploaded workbook: {e}")
    try:
        sheet = first_data_sheet(workbook)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    return workbook, sheet, parse_worksheet(sheet)


def _check_template_upload(request_info, fieldplan_id: str, solution_code: str, parsed):
    """Sheet-level checks shared by validate and create.

    These have no row to annotate, so each raises a 400 with a plain message rather than
    coming back inside the workbook.
    """
    if not fieldPlan_service_url:
        raise HTTPException(status_code=500, detail="Field plan service is not configured")
    fieldplan_client = FieldPlanServiceClient(fieldPlan_service_url)

    try:
        plans = fieldplan_client.search_fieldPlan(request_info, fieldplan_id).get("FieldPlans", [])
    except Exception as e:
        logger.error(f"Could not read field plan {fieldplan_id}: {e}", exc_info=True)
        raise HTTPException(status_code=502, detail=f"Could not read installation plan {fieldplan_id}: {e}")
    if not plans:
        raise HTTPException(status_code=404, detail=f"Installation plan {fieldplan_id} not found")

    # Once a plan is published its tasks have been dispatched to vendors, so the template that
    # seeded them must not move underneath them.
    status = str(plans[0].get("status") or "").strip().upper()
    if status == PLAN_STATUS_PUBLISHED:
        raise HTTPException(
            status_code=400,
            detail=f"Installation plan {fieldplan_id} has already been published, so its IC "
                   f"Report templates can no longer be changed.")

    try:
        links = fieldplan_client.search_fieldplan_facility(
            request_info, fieldplan_id).get("FieldPlanFacilities", [])
    except Exception as e:
        logger.error(f"Could not read scope for plan {fieldplan_id}: {e}", exc_info=True)
        raise HTTPException(status_code=502, detail=f"Could not read the plan's scope: {e}")

    in_scope = any(
        str(link.get("solutionId") or "").strip() == str(solution_code).strip()
        and not link.get("isdeleted")
        for link in links
    )
    if not in_scope:
        raise HTTPException(
            status_code=400,
            detail=f"No end user site in this installation plan is assigned Solution "
                   f"{solution_code}. Assign it in the Installation Scope step first.")

    # The workbook carries its own Bundle / Item Code, which is what makes uploading the wrong
    # Solution's file detectable -- easy to do when a Plan has several similar templates open.
    if parsed.bundle_code and str(parsed.bundle_code).strip() != str(solution_code).strip():
        raise HTTPException(
            status_code=400,
            detail=f"This workbook is the template for Solution {parsed.bundle_code}, not "
                   f"{solution_code}. Please upload the file downloaded for this Solution.")


@router.post('/installationTemplateValidateData',
             summary='Validate a filled IC Report template before creating it',
             response_description='Returns the workbook with status/error columns per line item')
async def validate_installation_template(
        background_tasks: BackgroundTasks,
        template_file: UploadFile = File(..., description="The filled IC Report template"),
        fieldplan_id: str = Form(..., description="Field plan id"),
        solution_code: str = Form(..., description="MDMS Installation.Solution code"),
        request_info: str = Form(default="")
):
    """Validate only -- writes nothing.

    Errors are annotated onto the line-item rows and the workbook is handed back, so the
    Project Manager fixes the flagged rows in place and re-uploads. The frontend calls
    createInstallationTemplate once this returns X-Error-Count: 0.
    """
    request_info_obj = request_info_from_json(request_info)
    temp_file = None
    output_path = None
    try:
        temp_file, _ = await _save_upload_to_temp_file(template_file, suffix=".xlsx")
        workbook, sheet, parsed = _load_and_parse_template(temp_file.name)
        _check_template_upload(request_info_obj, fieldplan_id, solution_code, parsed)

        row_errors, sheet_errors = validate_line_items(parsed)
        if sheet_errors:
            raise HTTPException(status_code=400, detail=" ".join(sheet_errors))

        error_count = annotate_worksheet(sheet, parsed, row_errors)

        output_path = create_temp_file(suffix=".xlsx")
        workbook.save(output_path)
        background_tasks.add_task(cleanup_temp_file, output_path)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        response = FileResponse(
            path=output_path,
            filename=f"ic_report_template_validation_{timestamp}.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response.headers["X-Error-Count"] = str(error_count)
        return response

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error validating installation template: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Validation failed: {str(e)}")
    finally:
        if temp_file and os.path.exists(temp_file.name):
            os.unlink(temp_file.name)


@router.post('/createInstallationTemplate',
             summary='Store a validated IC Report template for one Solution',
             response_description='Returns the saved template summary as JSON')
async def create_installation_template(
        template_file: UploadFile = File(..., description="The filled IC Report template"),
        fieldplan_id: str = Form(..., description="Field plan id"),
        solution_code: str = Form(..., description="MDMS Installation.Solution code"),
        request_info: str = Form(default="")
):
    """Parse the filled template and store it against (field plan, Solution).

    Re-runs the same validation rather than trusting the status column the validate step
    wrote: that cell is just data in a workbook the Project Manager holds, so it can be
    edited. Validation is one call away, so checking properly costs nothing.

    The uploaded file itself is not retained -- once parsed it has no consumer, and the
    download only ever serves the blank template.
    """
    request_info_obj = request_info_from_json(request_info)
    temp_file = None
    try:
        temp_file, _ = await _save_upload_to_temp_file(template_file, suffix=".xlsx")
        _workbook, _sheet, parsed = _load_and_parse_template(temp_file.name)
        _check_template_upload(request_info_obj, fieldplan_id, solution_code, parsed)

        row_errors, sheet_errors = validate_line_items(parsed)
        if sheet_errors:
            raise HTTPException(status_code=400, detail=" ".join(sheet_errors))
        if row_errors:
            rows = ", ".join(str(r) for r in sorted(row_errors)[:10])
            raise HTTPException(
                status_code=400,
                detail=f"{len(row_errors)} line item(s) still have errors (rows {rows}). "
                       f"Run the validation step and fix the flagged rows before saving.")

        machine_section, solar_section = to_sections(parsed)

        try:
            FieldPlanServiceClient(fieldPlan_service_url).create_field_plan_template(
                request_info=request_info_obj,
                fieldplan_id=fieldplan_id,
                solution_code=solution_code,
                machine_section=machine_section,
                solar_section=solar_section,
                tender_number=parsed.tender_number,
                purchase_order_number=parsed.purchase_order_number,
            )
        except Exception as e:
            logger.error(f"Could not save field plan template: {e}", exc_info=True)
            raise HTTPException(status_code=502, detail=f"Could not save the template: {e}")

        logger.info(
            f"Stored IC Report template: fieldplan={fieldplan_id} solution={solution_code} "
            f"machines={len(machine_section)} solar={len(solar_section)}")
        return JSONResponse(content={
            "fieldPlanId": fieldplan_id,
            "solutionId": solution_code,
            "machineCount": len(machine_section),
            "solarLineItemCount": len(solar_section),
            "tenderNumber": parsed.tender_number,
            "purchaseOrderNumber": parsed.purchase_order_number,
            "message": "IC Report template saved.",
        })

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error creating installation template: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to save template: {str(e)}")
    finally:
        if temp_file and os.path.exists(temp_file.name):
            os.unlink(temp_file.name)



@router.post('/amcConfigurationValidateData',
             summary='Validate AMC configuration Excel file before processing',
             response_description='Returns validation report Excel with PASSED/FAILED rows')
async def validate_amc_configurations_excel_sheet(
        background_tasks: BackgroundTasks,
        amc_file: UploadFile = File(..., description="Excel file containing AMC configuration data"),
        amc_sheet_name: str = Form(default="amc-configurations", description="Name of the sheet containing AMC data"),
        request_info: str = Form(default="")
):
    input_temp_file = None
    output_temp_file = None
    request_info_obj = request_info_from_json(request_info)

    try:
        input_temp_file, _ = await _save_upload_to_temp_file(amc_file, suffix=".xlsx")

        # Prepare output file
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"amc_configuration_validation_results_{timestamp}.xlsx"
        output_temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        output_temp_file.close()
        output_file_path = output_temp_file.name

        # Read Excel file
        wb = load_workbook(input_temp_file.name)
        if amc_sheet_name not in wb.sheetnames:
            raise HTTPException(status_code=400, detail=f"Sheet '{amc_sheet_name}' not found")

        df = pd.read_excel(input_temp_file.name, sheet_name=amc_sheet_name)
        df.columns = [str(c).strip() for c in df.columns]

        # Status/error must be object dtype (Excel may load as float); avoids FutureWarning on df.loc writes.
        for _col in ("status", "error"):
            if _col not in df.columns:
                df[_col] = ""
            else:
                df[_col] = df[_col].map(lambda x: "" if pd.isna(x) else str(x))
            df[_col] = df[_col].astype("object")

        required_columns = [
            "Facility Id",
            "Health Facility Name",
            # "Vendor",
            "AMC-Frequency",
            "AMC-Duration"
        ]

        # Column names
        # vendor_col = "Vendor" if required_columns else "vendor"
        frequency_col = "AMC-Frequency" if required_columns else "amc-frequency"
        duration_col = "AMC-Duration" if required_columns else "amc-duration"

        # Validate each row - only check vendor, AMC frequency, and AMC duration
        error_count = 0
        for index, row in df.iterrows():
            validation_errors = []

            try:
                # Validate vendor, AMC frequency, and AMC duration
                # vendor_name = str(row.get(vendor_col, "")).strip() if not pd.isna(row.get(vendor_col)) else ""
                amc_frequency = str(row.get(frequency_col, "")).strip() if not pd.isna(row.get(frequency_col)) else ""
                amc_duration = str(row.get(duration_col, "")).strip() if not pd.isna(row.get(duration_col)) else ""

                # Check if fields are filled
                if not amc_frequency or not amc_duration:
                    validation_errors.append(
                        "Please ensure AMC frequency, and duration are selected for all listed assets before upload."
                    )

                # Set status and error
                if validation_errors:
                    df.at[index, 'status'] = 'FAILED'
                    df.at[index, 'error'] = "; ".join(validation_errors)
                    error_count += 1
                else:
                    df.at[index, 'status'] = 'PASSED'
                    df.at[index, 'error'] = ''

            except Exception as e:
                df.at[index, 'status'] = 'FAILED'
                df.at[index, 'error'] = f'Unexpected error: {str(e)}'
                error_count += 1
                logger.error(f"Error validating row {index}: {e}")

        # Write results to Excel
        ws = wb[amc_sheet_name]
        header_values = [cell.value for cell in ws[1]]

        # Add status/error columns if missing
        for col_name in ["status", "error"]:
            if col_name not in header_values:
                new_col_idx = len(header_values) + 1
                cell = ws.cell(row=1, column=new_col_idx, value=col_name)
                cell.font = Font(bold=True)
                header_values.append(col_name)

                # lock header cell
                cell.protection = Protection(locked=True)

                # lock all data cells in this new column
                for r_idx in range(2, ws.max_row + 1):
                    ws.cell(row=r_idx, column=new_col_idx).protection = Protection(locked=True)

        grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        # Write data rows back (without header row)
        export_df = prepare_dataframe_for_excel_export(df)
        for r_idx, row in enumerate(dataframe_to_rows(export_df, index=False, header=False), start=2):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)

                # force lock for status/error columns
                if ws.cell(1, c_idx).value in ["status", "error"]:
                    cell.protection = Protection(locked=True)
                    cell.fill = grey_fill

        # Ensure sheet protection is ON
        ws.protection.sheet = True
        ws.protection.enable()

        # Save to output file
        wb.save(output_file_path)

        autofit_columns(output_file_path, amc_sheet_name, auto_fit=True)

        background_tasks.add_task(cleanup_temp_file, output_file_path)
        background_tasks.add_task(cleanup_temp_file, input_temp_file.name)

        response = FileResponse(
            path=output_file_path,
            filename=output_filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response.headers["X-Error-Count"] = str(error_count)

        return response

    except HTTPException:
        if input_temp_file and os.path.exists(input_temp_file.name):
            os.unlink(input_temp_file.name)
        if output_temp_file and os.path.exists(output_temp_file.name):
            os.unlink(output_temp_file.name)
        raise
    except Exception as e:
        logger.error(f"Unhandled error in validate_amc_configurations_excel_sheet: {e}")
        if input_temp_file and os.path.exists(input_temp_file.name):
            os.unlink(input_temp_file.name)
        if output_temp_file and os.path.exists(output_temp_file.name):
            os.unlink(output_temp_file.name)
        raise HTTPException(status_code=500, detail=f"An unexpected error occurred: {str(e)}")


def get_vendor_id_for_amc_field_staff(user_info_data: List[dict]) -> str:
    # Vendor id (or name fallback) for the vendor that has a user with role AMC_FIELD_STAFF.
    role_code = "AMC_FIELD_STAFF"
    candidates: Set[str] = set()

    for vendor_mapping in user_info_data:
        vendor_id = (vendor_mapping.get("vendorId") or "").strip()
        vendor_name = (vendor_mapping.get("vendor") or "").strip()
        if not vendor_id:
            if not vendor_name:
                continue
            vendor_id = vendor_name

        users = vendor_mapping.get("users", [])
        if not users and "userId" in vendor_mapping:
            users = [{"userId": vendor_mapping.get("userId"), "userName": vendor_mapping.get("userName")}]
        if not isinstance(users, list):
            raise HTTPException(status_code=400, detail=f"users must be a list for vendorId: {vendor_id}")

        for user in users:
            for role in user.get("roles") or []:
                if role.get("code") == role_code:
                    candidates.add(vendor_id)
                    break

    if not candidates:
        raise HTTPException(
            status_code=400,
            detail="No vendor with a user having the AMC_FIELD_STAFF role was found in user_info_list",
        )
    if len(candidates) > 1:
        raise HTTPException(
            status_code=400,
            detail="Multiple vendors have the AMC_FIELD_STAFF role; user_info_list must identify a single vendor",
        )
    return next(iter(candidates))


@router.post('/amcConfigurationBulkIngest',
             summary='Bulk ingest AMC configuration template data',
             response_description="Returns processed Excel file with AMC configuration creation results")
async def bulk_ingest_amc_configurations(
        background_tasks: BackgroundTasks,
        amc_file: UploadFile = File(..., description="Excel file containing AMC configuration data"),
        amc_sheet_name: str = Form(default="amc-configurations", description="Name of the sheet containing AMC data"),
        project_id: str = Form(..., description="Project ID"),
        user_info_list: str = Form(..., description="JSON array of user info objects with vendor mapping"),
        request_info: str = Form(default="")
):
    input_temp_file = None
    output_temp_file = None
    request_info_obj = request_info_from_json(request_info)

    # Get tenant ID from request info or use default
    tenant_id = request_info_obj.user_info.tenant_id if request_info_obj.user_info and request_info_obj.user_info.tenant_id else LIVELIHOOD_TENANT_ID

    try:
        # Parse user info list
        try:
            user_info_data = json.loads(user_info_list)
            if not isinstance(user_info_data, list):
                raise HTTPException(status_code=400, detail="user_info_list must be a JSON array")
        except json.JSONDecodeError as e:
            raise HTTPException(status_code=400, detail=f"Invalid JSON in user_info_list: {str(e)}")

        amc_vendor_id = get_vendor_id_for_amc_field_staff(user_info_data)

        assignment_users = []
        vendor_mappings = []  # One entry per payload item with valid users
        for vendor_mapping in user_info_data:
            # Primary key: vendorId (UUID)
            vendor_id = vendor_mapping.get("vendorId", "").strip()
            # Secondary: vendor name (for backward compatibility and Excel lookup)
            vendor_name = vendor_mapping.get("vendor", "").strip()

            if not vendor_id:
                # Fallback: if no vendorId, use vendor name as key
                if not vendor_name:
                    logger.warning(f"Vendor mapping missing both vendorId and vendor name: {vendor_mapping}")
                    continue
                vendor_id = vendor_name  # Use name as fallback key

            # Support both old format (single user) and new format (list of users)
            users = vendor_mapping.get("users", [])
            if not users:
                # Backward compatibility: if "users" not found, check for single user fields
                if "userId" in vendor_mapping:
                    users = [{"userId": vendor_mapping.get("userId"), "userName": vendor_mapping.get("userName")}]
                else:
                    logger.warning(f"No users found for vendorId: {vendor_id}")
                    continue

            # Validate users list
            if not isinstance(users, list):
                raise HTTPException(status_code=400, detail=f"users must be a list for vendorId: {vendor_id}")

            # Process users
            for user in users:
                # Extract user ID - prefer 'id' (from full user object), then 'userId', then 'uuid'
                user_id = user.get("uuid") or user.get("userId") or user.get("id")

                if not user_id:
                    logger.warning(f"User object missing ID field: {user}")
                    continue

                # Extract user name - prefer 'name', then 'userName'
                user_name = user.get("name") or user.get("userName", "")

                # Extract tenant ID from user object if available
                user_tenant_id = user.get("tenantId")

                assignment_users.append({
                    "id": str(user_id),  # Convert to string for consistency
                    "userId": str(user_id),  # Keep for backward compatibility
                    "userName": user_name,
                    "name": user_name,
                    "tenantId": user_tenant_id,
                    "fullUser": user  # Store full user object for reference
                })

            if not assignment_users:
                logger.warning(f"No valid users found for vendorId: {vendor_id}")
                continue

            vendor_mappings.append({
                "vendorId": vendor_id,
                "vendorName": vendor_name,
                "users": assignment_users
            })

        # Save uploaded file
        input_temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        content = await amc_file.read()
        input_temp_file.write(content)
        input_temp_file.close()

        # Prepare output file
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"amc_configuration_ingestion_results_{timestamp}.xlsx"
        output_temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        output_temp_file.close()
        output_file_path = output_temp_file.name

        # Read Excel file
        wb = load_workbook(input_temp_file.name)
        if amc_sheet_name not in wb.sheetnames:
            raise HTTPException(status_code=400, detail=f"Sheet '{amc_sheet_name}' not found")

        df = pd.read_excel(input_temp_file.name, sheet_name=amc_sheet_name)
        df.columns = [str(c).strip() for c in df.columns]

        for _col in ("status", "error"):
            if _col not in df.columns:
                df[_col] = ""
            else:
                df[_col] = df[_col].map(lambda x: "" if pd.isna(x) else str(x))
            df[_col] = df[_col].astype("object")

        required_columns = ["Facility Id", "Health Facility Name", "Vendor", "AMC-Frequency", "AMC-Duration"]

        # Initialize clients
        facility_client = FacilityServiceClient(facility_service_url) if facility_service_url else None
        amc_client = AMCSchedulerServiceClient(amc_scheduler_service_url) if amc_scheduler_service_url else None

        if not amc_client:
            raise HTTPException(status_code=500, detail="AMC Scheduler Service is not configured")

        if not facility_client:
            raise HTTPException(status_code=500, detail="Facility Service is not configured")

        # Track configurations to detect duplicates (vendor-facility-project combination)
        seen_configs = set()

        facility_ids_from_file = []
        for _, row in df.iterrows():
            if pd.isna(row.get("Facility Id")) and pd.isna(row.get("Health Facility Name")):
                continue
            facility_id = str(row.get("Facility Id", "")).strip()
            if facility_id:
                facility_ids_from_file.append(facility_id)

        facility_map = {}
        if facility_ids_from_file:
            unique_facility_ids = list(dict.fromkeys(facility_ids_from_file))
            facility_batch_size = int(os.getenv("AMC_INGEST_FACILITY_ID_BATCH_SIZE", "500"))
            try:
                for batch_start in range(0, len(unique_facility_ids), facility_batch_size):
                    batch_ids = unique_facility_ids[batch_start:batch_start + facility_batch_size]
                    bulk_facility_result = facility_client.bulk_search_facility(
                        request_info=request_info_obj,
                        tenant_ids=[LIVELIHOOD_TENANT_ID],
                        facility_ids=batch_ids,
                        limit=max(len(batch_ids), 50),
                        send_non_paginated_response=True,
                    )
                    for facility in (bulk_facility_result.get("facilities", []) or []):
                        f_id = facility.get("facility_id")
                        if f_id:
                            facility_map[f_id] = facility
            except Exception as e:
                logger.error(f"Error bulk searching facilities for AMC ingest: {e}", exc_info=True)
                raise HTTPException(status_code=502, detail=f"Facility lookup failed: {str(e)}")

        asset_types_formatted = []
        asset_type_names = {
            "INVERTER": "Inverter",
            "PANEL": "Panel",
            "BATTERY": "Battery"
        }
        for asset_type in DEFAULT_AMC_ASSET_TYPES:
            asset_types_formatted.append({
                "code": asset_type,
                "name": asset_type_names.get(asset_type, asset_type.title())
            })

        assignments_template = []
        for user in assignment_users:
            assigned_user_id = user.get("id") or user.get("userId")
            assignment_tenant_id = user.get("tenantId") or tenant_id
            assignments_template.append({
                "assignedUser": str(assigned_user_id),
                "tenantId": assignment_tenant_id,
            })

        now = datetime.now()
        configuration_start_date = int(now.timestamp() * 1000)

        configs_to_create = []
        row_indexes_for_configs = []

        for index, row in df.iterrows():
            try:
                # Skip empty rows
                if pd.isna(row.get("Facility Id")) and pd.isna(row.get("Health Facility Name")):
                    df.at[index, 'status'] = 'skipped'
                    df.at[index, 'error'] = 'Empty row'
                    continue

                # Get facility by Facility ID
                facility_id = str(row.get("Facility Id", "")).strip()
                if not facility_id:
                    df.at[index, 'status'] = 'failed'
                    df.at[index, 'error'] = 'Facility Id is required'
                    continue

                if facility_id not in facility_map:
                    df.at[index, 'status'] = 'failed'
                    df.at[index, 'error'] = f'Facility not found for Facility Id: {facility_id}'
                    continue

                # Get AMC frequency and duration (already validated, just convert to months)
                frequency_col = "AMC-Frequency" if "AMC-Frequency" in df.columns else "amc-frequency"
                duration_col = "AMC-Duration" if "AMC-Duration" in df.columns else "amc-duration"
                amc_frequency = str(row.get(frequency_col, "")).strip()
                amc_duration = str(row.get(duration_col, "")).strip()

                # Convert frequency to months (format already validated in validation endpoint)
                if amc_frequency == "Every 6 Months":
                    frequency_months = 6
                elif amc_frequency == "Every 1 Year":
                    frequency_months = 12
                else:
                    df.at[index, 'status'] = 'failed'
                    df.at[index, 'error'] = f'Unexpected AMC frequency value: {amc_frequency}'
                    continue

                # Convert duration to months (format already validated in validation endpoint)
                if amc_duration == "1 Year":
                    duration_months = 12
                elif amc_duration == "3 Years":
                    duration_months = 36
                elif amc_duration == "5 Years":
                    duration_months = 60
                else:
                    df.at[index, 'status'] = 'failed'
                    df.at[index, 'error'] = f'Unexpected AMC duration value: {amc_duration}'
                    continue

                # Check for duplicate configuration (vendor-facility-project combination)
                config_key = (facility_id, project_id)
                if config_key in seen_configs:
                    df.at[index, 'status'] = 'failed'
                    df.at[index, 'error'] = 'Duplicate configuration: vendor-facility-project combination already exists'
                    continue
                seen_configs.add(config_key)

                # Create assignments array from vendor users
                assignments = []
                for user in assignment_users:
                    # Use user's id (from full user object) or userId (backward compatibility)
                    assigned_user_id = user.get("id") or user.get("userId")
                    # Prefer user's tenantId if available, otherwise use default tenant_id
                    assignment_tenant_id = user.get("tenantId") or tenant_id

                    assignment = {
                        "assignedUser": str(assigned_user_id),
                        "tenantId": assignment_tenant_id
                    }
                    assignments.append(assignment)

                # Convert asset types to API format (objects with code and name)
                asset_types_formatted = []
                asset_type_names = {
                    "INVERTER": "Inverter",
                    "PANEL": "Panel",
                    "BATTERY": "Battery"
                }
                for asset_type in DEFAULT_AMC_ASSET_TYPES:
                    asset_types_formatted.append({
                        "code": asset_type,
                        "name": asset_type_names.get(asset_type, asset_type.title())
                    })

                # Calculate configuration dates (start date = now, end date = start + duration)
                now = datetime.now()
                configuration_start_date = int(now.timestamp() * 1000)  # Convert to milliseconds
                end_date = now + timedelta(days=duration_months * 30)  # Approximate: 30 days per month
                configuration_end_date = int(end_date.timestamp() * 1000)

                configs_to_create.append({
                    "tenantId": tenant_id,
                    "vendorId": amc_vendor_id,
                    "facilityId": facility_id,
                    "projectId": project_id,
                    "durationMonths": duration_months,
                    "visitFrequencyMonths": frequency_months,
                    "status": "ACTIVE",
                    "configurationStartDate": configuration_start_date,
                    "configurationEndDate": configuration_end_date,
                    "assetTypes": asset_types_formatted,
                    "assignments": assignments
                })
                row_indexes_for_configs.append(index)
            except Exception as e:
                df.at[index, 'status'] = 'failed'
                df.at[index, 'error'] = f'Unexpected error: {str(e)}'
                logger.error(f"Error processing row {index}: {e}")

        if configs_to_create:
            chunk_size = AMC_CONFIGURATION_BULK_CHUNK_SIZE
            n_cfgs = len(configs_to_create)

            def _process_amc_chunk(
                chunk_cfgs: List[dict],
                chunk_row_indexes: List,
                http_session: requests.Session,
            ) -> None:
                try:
                    amc_client.create_amc_configurations_bulk(
                        request_info_obj,
                        chunk_cfgs,
                        session=http_session,
                    )
                    df.loc[chunk_row_indexes, "status"] = "success"
                    df.loc[chunk_row_indexes, "error"] = ""
                except Exception as exc:
                    logger.error(
                        "Bulk AMC create failed for %s rows: %s",
                        len(chunk_cfgs),
                        exc,
                        exc_info=True,
                    )
                    err = str(exc)
                    df.loc[chunk_row_indexes, "status"] = "failed"
                    df.loc[chunk_row_indexes, "error"] = err

            with requests.Session() as http_session:
                for start in range(0, n_cfgs, chunk_size):
                    _process_amc_chunk(
                        configs_to_create[start:start + chunk_size],
                        row_indexes_for_configs[start:start + chunk_size],
                        http_session,
                    )

        with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name=amc_sheet_name)

        autofit_columns(output_file_path, amc_sheet_name, auto_fit=True)

        background_tasks.add_task(cleanup_temp_file, output_file_path)
        background_tasks.add_task(cleanup_temp_file, input_temp_file.name)

        return FileResponse(
            path=output_file_path,
            filename=output_filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except HTTPException:
        if input_temp_file and os.path.exists(input_temp_file.name):
            os.unlink(input_temp_file.name)
        if output_temp_file and os.path.exists(output_temp_file.name):
            os.unlink(output_temp_file.name)
        raise
    except Exception as e:
        logger.error(f"Unhandled error in bulk_ingest_amc_configurations: {e}")
        if input_temp_file and os.path.exists(input_temp_file.name):
            os.unlink(input_temp_file.name)
        if output_temp_file and os.path.exists(output_temp_file.name):
            os.unlink(output_temp_file.name)
        raise HTTPException(status_code=500, detail=f"An unexpected error occurred: {str(e)}")