"""Parse and validate the IC Report template workbook the Project Manager fills in.

The workbook is served straight out of filestore, so its layout is the source ICC sheet's:
columns A-E are `Sl. No. | Product | Make | Capacity | Quantity`, and rows come in exactly
three kinds -- section headers, category rows, and line items. See SECTION_* below.

Everything here is position-independent. The Project Manager can insert and delete rows
freely, so nothing may key off a fixed row number; the parser walks column A and tracks
which section and category it is currently inside.
"""
from typing import Any, Dict, List, NamedTuple, Optional, Tuple

from openpyxl.worksheet.worksheet import Worksheet

from app.core.logging import AppLogger

logger = AppLogger().get_logger()

# The only three section titles that appear across all 14 source sheets.
SECTION_SOLAR = "bill of material"
SECTION_MACHINE = "associated machines"
SECTION_TECHNICIAN = "system functionality parameters"

# Column positions within the BOM sections, 1-based.
COL_SL_NO = 1
COL_PRODUCT = 2
COL_MAKE = 3
COL_CAPACITY = 4
COL_QUANTITY = 5

# Header block labels, read for the two values the Project Manager may set.
LABEL_TENDER = "tender no"
LABEL_PURCHASE_ORDER = "purchase/work order no"
LABEL_BUNDLE_CODE = "bundle"

ANNOTATION_HEADERS = ("status", "error")
STATUS_PASSED = "PASSED"
STATUS_FAILED = "FAILED"


class LineItem(NamedTuple):
    """One parsed BOM row, with the sheet row it came from so errors can be annotated."""
    row: int
    section: str          # "solar" | "machine"
    category: str
    product: str
    make: str
    capacity: str
    quantity: Optional[Any]   # kept raw; validation decides whether it is acceptable


class ParsedTemplate(NamedTuple):
    line_items: List[LineItem]
    tender_number: Optional[str]
    purchase_order_number: Optional[str]
    bundle_code: Optional[str]


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _is_section_header(label: str) -> Optional[str]:
    lowered = label.lower()
    if lowered.startswith(SECTION_SOLAR):
        return "solar"
    if lowered.startswith(SECTION_MACHINE):
        return "machine"
    if lowered.startswith(SECTION_TECHNICIAN):
        return "technician"
    return None


def parse_worksheet(sheet: Worksheet) -> ParsedTemplate:
    """Walk the sheet top to bottom and pull out the two BOM sections.

    Row classification, in order of precedence:
      * column A matches a known section title -> switch section
      * column A is an integer -> a line item
      * column A is a non-empty string and column B is empty -> a category row
      * anything else -> ignored (the header block, blank spacers, notes)

    A line item is recognised on its Sl. No. alone, deliberately not on having a Product too.
    Requiring the Product would mean a Project Manager who cleared that cell had the whole
    row silently dropped from the stored template -- 39 line items quietly becoming 38, and
    the sheet still coming back PASSED. Keying on Sl. No. turns that into the "Product is
    required" row error it should be. Verified against all 14 source sheets: inside the BOM
    sections every integer Sl. No. already carries a product, so nothing is reclassified.

    Parsing stops at the SYSTEM FUNCTIONALITY PARAMETERS section: everything below it is the
    Field Technician's on-site readings, which the Project Manager does not fill.
    """
    line_items: List[LineItem] = []
    tender_number = purchase_order_number = bundle_code = None
    section: Optional[str] = None
    category = ""

    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
        label = _text(row[COL_SL_NO - 1].value) if len(row) >= COL_SL_NO else ""
        second = row[COL_PRODUCT - 1].value if len(row) >= COL_PRODUCT else None

        # The header block sits above the first section, and its values live in column B.
        if section is None and label:
            lowered = label.lower()
            if lowered.startswith(LABEL_TENDER):
                tender_number = _text(second) or None
            elif lowered.startswith(LABEL_PURCHASE_ORDER):
                purchase_order_number = _text(second) or None
            elif LABEL_BUNDLE_CODE in lowered:
                bundle_code = _text(second) or None

        if label:
            found = _is_section_header(label)
            if found == "technician":
                break
            if found:
                section, category = found, ""
                continue

        if section is None:
            continue

        raw_sl_no = row[COL_SL_NO - 1].value
        if isinstance(raw_sl_no, int) and not isinstance(raw_sl_no, bool):
            line_items.append(LineItem(
                row=row[0].row,
                section=section,
                category=category,
                product=_text(second),
                make=_text(row[COL_MAKE - 1].value) if len(row) >= COL_MAKE else "",
                capacity=_text(row[COL_CAPACITY - 1].value) if len(row) >= COL_CAPACITY else "",
                quantity=row[COL_QUANTITY - 1].value if len(row) >= COL_QUANTITY else None,
            ))
        elif label and not _text(second):
            category = label

    logger.info(
        f"Parsed template sheet {sheet.title!r}: {sum(1 for i in line_items if i.section == 'solar')} "
        f"solar and {sum(1 for i in line_items if i.section == 'machine')} machine line items")
    return ParsedTemplate(line_items, tender_number, purchase_order_number, bundle_code)


def _validate_quantity(raw: Any) -> Optional[str]:
    """Blank is allowed by design -- the source sheet says quantities that depend on string
    configuration, cable run length or site layout are for the installing engineer to fill in
    on site. Anything present, though, has to be a positive whole number."""
    if raw is None or _text(raw) == "":
        return None
    if isinstance(raw, bool):
        return "Quantity must be a whole number"
    if isinstance(raw, int):
        return None if raw > 0 else "Quantity must be greater than zero"
    if isinstance(raw, float):
        if not raw.is_integer():
            return "Quantity must be a whole number"
        return None if raw > 0 else "Quantity must be greater than zero"
    try:
        value = float(_text(raw))
    except ValueError:
        return "Quantity must be a whole number"
    if not value.is_integer():
        return "Quantity must be a whole number"
    return None if value > 0 else "Quantity must be greater than zero"


def validate_line_items(parsed: ParsedTemplate) -> Tuple[Dict[int, List[str]], List[str]]:
    """Returns (errors keyed by sheet row, sheet-level errors).

    Sheet-level problems have no row to annotate, so the caller turns those into a 400 rather
    than a workbook.
    """
    row_errors: Dict[int, List[str]] = {}
    sheet_errors: List[str] = []

    def add(row: int, message: str) -> None:
        row_errors.setdefault(row, []).append(message)

    for item in parsed.line_items:
        if not item.product:
            add(item.row, "Product is required")
        quantity_error = _validate_quantity(item.quantity)
        if quantity_error:
            add(item.row, quantity_error)
        # A line item with no category above it means rows were inserted before the first
        # category header, which would silently lose the item's grouping.
        if item.section == "solar" and not item.category:
            add(item.row, "This line item is not under any category")

    if not any(item.section == "solar" for item in parsed.line_items):
        sheet_errors.append(
            "The Bill Of Material (For Solar System) section has no line items. "
            "Every Solution in this programme is a solar installation.")

    return row_errors, sheet_errors


def to_sections(parsed: ParsedTemplate) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    """Build the (machine_section, solar_section) arrays that get stored as JSONB.

    Sl. No. is renumbered 1..N per section rather than carried over from the sheet, so the
    stored numbering is always contiguous even after the Project Manager inserts or deletes
    rows. For the machine section this is load-bearing: Vendor Assignment turns entry N into
    the MACHINE asset with component_sequence N, so position and slNo must agree.
    """
    machine_section: List[Dict[str, Any]] = []
    solar_section: List[Dict[str, Any]] = []

    for item in parsed.line_items:
        target = machine_section if item.section == "machine" else solar_section
        entry: Dict[str, Any] = {
            "slNo": len(target) + 1,
            "product": item.product,
            "capacity": item.capacity,
        }
        # Omitted rather than stored as empty so a blank cell reads as "for the technician to
        # fill" downstream, which is what the source sheet's own note says it means.
        if item.make:
            entry["make"] = item.make
        quantity = _normalise_quantity(item.quantity)
        if quantity is not None:
            entry["quantity"] = quantity
        if item.section == "solar":
            entry["category"] = item.category
        target.append(entry)

    return machine_section, solar_section


def _normalise_quantity(raw: Any) -> Optional[int]:
    if raw is None or _text(raw) == "":
        return None
    try:
        return int(float(_text(raw)))
    except ValueError:
        return None


def _annotation_columns(sheet: Worksheet) -> Tuple[int, int]:
    """Where to put the status/error columns.

    Reuses the pair already on the sheet when there is one. The Project Manager's normal loop
    is validate, fix the flagged rows in the returned workbook, re-upload -- so without this
    every round would append another status/error pair and a sheet would end up with three or
    four of them.

    Otherwise they go after the widest column actually used, discovered per sheet rather than
    fixed at F/G: the technician's section lower down reaches column I on some sheets, and a
    hardcoded position would overwrite real content.
    """
    for column in range(1, sheet.max_column + 1):
        if _text(sheet.cell(row=1, column=column).value).lower() == ANNOTATION_HEADERS[0]:
            return column, column + 1
    status_col = sheet.max_column + 1
    return status_col, status_col + 1


def annotate_worksheet(sheet: Worksheet, parsed: ParsedTemplate,
                       row_errors: Dict[int, List[str]]) -> int:
    """Write status/error columns against each line-item row and return the failure count."""
    status_col, error_col = _annotation_columns(sheet)
    sheet.cell(row=1, column=status_col, value=ANNOTATION_HEADERS[0])
    sheet.cell(row=1, column=error_col, value=ANNOTATION_HEADERS[1])

    failed = 0
    for item in parsed.line_items:
        messages = row_errors.get(item.row)
        if messages:
            failed += 1
            sheet.cell(row=item.row, column=status_col, value=STATUS_FAILED)
            sheet.cell(row=item.row, column=error_col, value="; ".join(dict.fromkeys(messages)))
        else:
            sheet.cell(row=item.row, column=status_col, value=STATUS_PASSED)
            sheet.cell(row=item.row, column=error_col, value="")
    return failed


def first_data_sheet(workbook) -> Worksheet:
    """The template has one solution sheet, but a download also carries a read-only Sites
    sheet, and a Project Manager may have added their own. Pick the sheet that actually holds
    a BOM section rather than assuming a position or a name -- the name is the Solution's, and
    Excel truncates it at 31 characters."""
    for name in workbook.sheetnames:
        sheet = workbook[name]
        for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 60)):
            label = _text(row[0].value) if row else ""
            if label and _is_section_header(label) in ("solar", "machine"):
                return sheet
    raise ValueError(
        "No Bill Of Material section found in the uploaded workbook. "
        "Please upload the template downloaded for this Solution.")
