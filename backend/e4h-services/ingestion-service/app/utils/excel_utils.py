from typing import Dict, List, Union, Any, Optional, Tuple

import numpy as np
import pandas as pd

from app.utils.facility_validator import (
    ERR_HFR_OR_NIN_REQUIRED_WHEN_HEALTH,
    ERR_POC_USERNAME_REQUIRED_WHEN_ANGANWADI,
)
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Protection, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from app.core.logging import AppLogger

logger = AppLogger().get_logger()

FACILITY_IDENTIFIER_COLUMNS: Tuple[str, ...] = ("HFR ID", "NIN ID", "PoC Username")


def _cast_whole_float_column_to_int(df: pd.DataFrame, col: str) -> None:
    series = df[col]
    non_null = series.dropna()
    if non_null.empty:
        return
    if not np.all(np.equal(np.mod(non_null.to_numpy(), 1), 0)):
        return
    if series.isna().any():
        df[col] = series.astype("Int64")
    else:
        df[col] = series.astype("int64")


def _normalize_forced_integer_column(df: pd.DataFrame, col: str) -> None:
    """Normalize a column so numeric values are stored as integers without decimals."""
    original = df[col]
    numeric = pd.to_numeric(original, errors="coerce")
    original_as_str = original.astype(str).str.strip()
    non_empty_mask = original.notna() & original_as_str.ne("")

    if non_empty_mask.any() and numeric[non_empty_mask].notna().all():
        df[col] = np.trunc(numeric).astype("Int64")
        return

    numeric_like_mask = numeric.notna()
    if numeric_like_mask.any():
        df.loc[numeric_like_mask, col] = np.trunc(numeric[numeric_like_mask]).astype("int64")


def normalize_excel_integer_columns(
    df: pd.DataFrame,
    *,
    force_columns: Optional[Tuple[str, ...]] = None,
    convert_all_whole_float64: bool = True,
) -> pd.DataFrame:
    """
    Normalize Excel-loaded numeric columns to integers without decimals.

    - convert_all_whole_float64: convert any float64 column whose non-null values are whole numbers.
    - force_columns: always normalize these columns (e.g. HFR ID, NIN ID, PoC Username), including
      object-typed cells that look numeric.
    """
    forced = set(force_columns or ())

    if convert_all_whole_float64:
        for col in df.select_dtypes(include=["float64"]).columns:
            if col not in forced:
                _cast_whole_float_column_to_int(df, col)

    for col in force_columns or ():
        if col in df.columns:
            _normalize_forced_integer_column(df, col)

    return df


def to_excel_cell_value(value: Any) -> Any:
    """Convert pandas/numpy scalars to values openpyxl can write (pd.NA -> None)."""
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(value, np.generic):
        if isinstance(value, np.floating):
            if np.isnan(value):
                return None
            if float(value).is_integer():
                return int(value)
            return float(value)
        return value.item()
    return value


def prepare_dataframe_for_excel_export(df: pd.DataFrame) -> pd.DataFrame:
    """Return a copy safe for openpyxl / dataframe_to_rows (no pd.NA)."""
    return df.apply(lambda col: col.map(to_excel_cell_value))


"""
    Add dropdowns to Excel using hidden sheets for maximum compatibility.

    How it works:
    1. Creates a hidden sheet named "_DropdownValues" to store dropdown options
    2. Each dropdown references its values via a range formula (e.g., '_DropdownValues'!$A$1:$A$5)
    3. Uses standard Excel data validation with type="list" which automatically shows dropdown arrows

    Args:
        file_path: Path to the Excel file
        sheet_name: Name of the sheet to add dropdowns to
        dropdowns: Dictionary mapping column headers to list of dropdown options
        allow_blank_map: Optional dictionary mapping column headers to allow_blank boolean
        max_extra_rows: Maximum number of extra rows to apply validation to
"""
def add_dropdowns_to_excel(
        file_path: str,
        sheet_name: str,
        dropdowns: Dict[str, List[str]],
        allow_blank_map: Optional[Dict[str, bool]],
        max_extra_rows: int = 1000
):
    logger.trace(f"Adding dropdowns to Excel: file={file_path}, sheet={sheet_name}")
    wb = load_workbook(file_path)
    ws = wb[sheet_name]
    header_row = 1
    # Extend range and ensure that data rows always start at row 2 or below max_row
    # to avoid invalid ranges like "A2:A1" which cause "max_row must be greater than min_row" errors.
    max_row = ws.max_row + max_extra_rows
    if max_row < 2:
        max_row = 2

    dropdown_count = 0

    # Create a hidden sheet for dropdown values if it doesn't exist
    hidden_sheet_name = "_DropdownValues"
    if hidden_sheet_name not in wb.sheetnames:
        hidden_ws = wb.create_sheet(hidden_sheet_name)
        hidden_ws.sheet_state = 'hidden'  # Hide the sheet
        # Start from row 1
        current_hidden_row = 1
    else:
        hidden_ws = wb[hidden_sheet_name]
        hidden_ws.sheet_state = 'hidden'
        # Find the next available row by finding the last non-empty cell in column A
        # Check from bottom up for efficiency
        current_hidden_row = 1
        if hidden_ws.max_row > 0:
            for row in range(hidden_ws.max_row, 0, -1):
                if hidden_ws.cell(row=row, column=1).value is not None:
                    current_hidden_row = row + 2  # Leave a gap of 1 row
                    break

    for column_header, options in dropdowns.items():
        if not options:
            continue

        allow_blank = (allow_blank_map or {}).get(column_header, True)

        # Find the column for this header
        col_letter = None
        for cell in ws[header_row]:
            if cell.value == column_header:
                col_letter = cell.column_letter
                break

        if not col_letter:
            logger.warning(f"Column header '{column_header}' not found in sheet '{sheet_name}'")
            continue

        # Write dropdown values to hidden sheet starting from current_hidden_row
        start_row = current_hidden_row
        for idx, option in enumerate(options, start=start_row):
            # If a value starts with =, +, -, or @ it can be treated as a formula.
            # Prefix with an apostrophe so it is always interpreted as plain text.
            raw_value = str(option)
            if raw_value and raw_value[0] in ("=", "+", "-", "@"):
                cell_value = "'" + raw_value
            else:
                cell_value = raw_value

            hidden_ws.cell(row=idx, column=1).value = cell_value

        end_row = start_row + len(options) - 1

        # Create the formula reference to the hidden sheet
        # Format: '_DropdownValues'!$A$start_row:$A$end_row
        # Use INDIRECT for better compatibility, or direct reference
        formula = f"'{hidden_sheet_name}'!$A${start_row}:$A${end_row}"

        # Create data validation with reference to hidden sheet
        dv = DataValidation(
            type="list",
            formula1=formula,
            allow_blank=allow_blank,
            showErrorMessage=True,
            showInputMessage=True
        )
        dv.error = "Please select from the list"
        dv.errorTitle = "Invalid Entry"
        dv.prompt = "Select a value from the dropdown"
        dv.promptTitle = "Select Value"

        # Apply validation to the column (skip header row)
        dv.add(f"{col_letter}2:{col_letter}{max_row}")
        ws.add_data_validation(dv)

        # Move to next position in hidden sheet (leave a gap of 1 row)
        current_hidden_row = end_row + 2
        dropdown_count += 1
        logger.debug(f"Added dropdown to column '{column_header}' with {len(options)} options using hidden sheet (rows {start_row}-{end_row})")

    wb.save(file_path)
    logger.info(f"Added {dropdown_count} dropdowns to sheet '{sheet_name}'")


def _next_free_hidden_row(hidden_ws) -> int:
    """First usable row in the hidden dropdown-values sheet, leaving a 1-row gap."""
    if hidden_ws.max_row > 0:
        for row in range(hidden_ws.max_row, 0, -1):
            if hidden_ws.cell(row=row, column=1).value is not None:
                return row + 2
    return 1


def add_row_specific_dropdown_to_excel(
        file_path: str,
        sheet_name: str,
        column_header: str,
        options_by_row: Dict[int, List[str]],
        allow_blank: bool = True,
):
    """Add a dropdown whose option list differs per row.

    add_dropdowns_to_excel applies one option list to a whole column; this applies a
    distinct list per data row, which the Solution column needs (its valid options depend
    on the site's state). Rows sharing an identical option list share one DataValidation,
    so a sheet spanning few states stays cheap.

    options_by_row maps a 0-based data-row position (0 -> spreadsheet row 2) to that row's
    allowed values. Rows absent from the map, or mapped to an empty list, get no dropdown.
    """
    if not options_by_row:
        return

    wb = load_workbook(file_path)
    ws = wb[sheet_name]

    col_letter = None
    for cell in ws[1]:
        if cell.value == column_header:
            col_letter = cell.column_letter
            break
    if not col_letter:
        logger.warning(f"Column header '{column_header}' not found in sheet '{sheet_name}'")
        wb.save(file_path)
        return

    hidden_sheet_name = "_DropdownValues"
    if hidden_sheet_name not in wb.sheetnames:
        hidden_ws = wb.create_sheet(hidden_sheet_name)
        current_hidden_row = 1
    else:
        hidden_ws = wb[hidden_sheet_name]
        current_hidden_row = _next_free_hidden_row(hidden_ws)
    hidden_ws.sheet_state = 'hidden'

    # Group rows by their option list so each distinct list is written once.
    rows_by_options: Dict[tuple, List[int]] = {}
    for row_position, options in options_by_row.items():
        if not options:
            continue
        rows_by_options.setdefault(tuple(options), []).append(row_position)

    for options, row_positions in rows_by_options.items():
        start_row = current_hidden_row
        for offset, option in enumerate(options):
            raw_value = str(option)
            # Leading =/+/-/@ would be read as a formula; force plain text.
            cell_value = "'" + raw_value if raw_value and raw_value[0] in ("=", "+", "-", "@") else raw_value
            hidden_ws.cell(row=start_row + offset, column=1).value = cell_value
        end_row = start_row + len(options) - 1

        dv = DataValidation(
            type="list",
            formula1=f"'{hidden_sheet_name}'!$A${start_row}:$A${end_row}",
            allow_blank=allow_blank,
            showErrorMessage=True,
            showInputMessage=True,
        )
        dv.error = "Please select from the list"
        dv.errorTitle = "Invalid Entry"
        dv.prompt = "Select a value from the dropdown"
        dv.promptTitle = "Select Value"
        ws.add_data_validation(dv)
        for row_position in sorted(row_positions):
            dv.add(f"{col_letter}{row_position + 2}")

        current_hidden_row = end_row + 2

    wb.save(file_path)
    logger.info(
        f"Added row-specific '{column_header}' dropdowns to '{sheet_name}': "
        f"{len(rows_by_options)} distinct option sets across {sum(len(r) for r in rows_by_options.values())} rows"
    )


def lock_excel_columns(
        file_path: str,
        sheet_name: str,
        column_headers_to_unlock: List[Union[str, int]]
) -> None:
    """
    Locks the entire Excel sheet except the specified columns by header name or index.

    Args:
        file_path: Path to the Excel file.
        sheet_name: Name of the sheet to modify.
        column_headers_to_unlock: List of column headers (e.g., 'Selection?') or 1-based column indices to keep editable.
    """
    wb = load_workbook(file_path)
    ws = wb[sheet_name]

    # Find column indices to unlock based on header names or 1-based indices
    column_indices_to_unlock = set()
    for identifier in column_headers_to_unlock:
        if isinstance(identifier, str):
            for col_index, cell in enumerate(ws[1], 1):  # Header is assumed to be in the first row
                if cell.value and str(cell.value).strip() == identifier.strip():
                    column_indices_to_unlock.add(col_index)
                    break
        elif isinstance(identifier, int):
            column_indices_to_unlock.add(identifier)

    max_rows = ws.max_row

    # Lock all cells by default
    for row in range(1, max_rows + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).protection = Protection(locked=True)

    # Unlock only the specified columns
    for col_index in column_indices_to_unlock:
        for row in range(1, max_rows + 1):
            ws.cell(row=row, column=col_index).protection = Protection(locked=False)

    # Enable worksheet protection and allow selection of unlocked cells
    ws.protection.sheet = True
    ws.protection.formatColumns = True
    wb.save(file_path)


def lock_prefilled_rows_in_excel(
    file_path: str,
    sheet_name: str,
    editable_columns: list,
    total_rows: int,
    total_columns: int,
    always_locked_columns: list = None,  # new parameter
    extra_append_rows: int = 1000
):
    always_locked_columns = always_locked_columns or []

    wb = load_workbook(file_path)
    ws = wb[sheet_name]

    grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    no_fill = PatternFill()  # reset

    # Get header row values
    header_row = [cell.value for cell in ws[1]]

    # Editable column indices
    editable_indices = [
        i + 1 for i, col in enumerate(header_row)
        if col and any(c in col for c in editable_columns)
    ]

    # Always locked column indices
    always_locked_indices = [
        i + 1 for i, col in enumerate(header_row)
        if col and any(c in col for c in always_locked_columns)
    ]

    # Lock prefilled rows completely (grey out non-editable cells)
    # Data rows start at row 2; end at row (1 + total_rows)
    prefilled_start_row = 2
    prefilled_end_row = 1 + total_rows
    for row_idx in range(prefilled_start_row, prefilled_end_row + 1):
        for col_idx in range(1, total_columns + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            # If column is editable -> unlock
            if col_idx in editable_indices:
                cell.protection = Protection(locked=False)
                cell.fill = no_fill
            # If column is always locked -> lock + grey
            elif col_idx in always_locked_indices:
                cell.protection = Protection(locked=True)
                cell.fill = grey_fill
            else:
                # other prefilled cells -> lock + grey
                cell.protection = Protection(locked=True)
                cell.fill = grey_fill

    # Leave appendable rows mostly unlocked, but respect always_locked columns.
    # For performance, avoid applying any fill styles on extra rows – only
    # adjust protection flags so the user experience is preserved without
    # expensive formatting operations on thousands of empty cells.
    append_start_row = prefilled_end_row + 1
    append_end_row = prefilled_end_row + extra_append_rows
    for row_idx in range(append_start_row, append_end_row + 1):
        for col_idx in range(1, total_columns + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if col_idx in always_locked_indices:
                cell.protection = Protection(locked=True)
            else:
                cell.protection = Protection(locked=False)

    # Enable protection
    ws.protection.select_unlocked_cells = True
    ws.protection.formatColumns = True
    ws.protection.insertRows = True
    ws.protection.sheet = True
    ws.protection.enable()

    wb.save(file_path)


def add_validations_to_excel(file_path: str,
                             sheet_name: str,
                             validations: Dict[str, Dict[str, str]],
                             allow_blank_map: Dict[str, bool],
                             max_extra_rows: int = 1000
                             ):
    """
    Adds Excel data validation (pattern + uniqueness) to specified columns.
    :param file_path: Excel file path
    :param sheet_name: Sheet where validation needs to be added
    :param validations: Dict[column_name] = {"type": "regex"/"unique", "pattern"/"message": str}
    :param allow_blank_map: Dict[str,bool]
    :param max_extra_rows: Maximum extra number of rows to allow
    """
    wb = load_workbook(file_path)
    ws = wb[sheet_name]

    header_row = 1
    max_row = ws.max_row + max_extra_rows  # allow future rows for data entry
    # Ensure max_row is at least 2 (header row is 1, data starts at row 2)
    max_row = max(max_row, 2)
    header_cells = {cell.value.strip(): cell for cell in ws[header_row] if cell.value}

    for col_name, config in validations.items():
        header_cell = header_cells.get(col_name)
        if not header_cell:
            continue

        col_idx = header_cell.column
        col_letter = get_column_letter(col_idx)
        data_range = f"{col_letter}2:{col_letter}{max_row}"
        allow_blank = allow_blank_map.get(col_name, True)

        if config["type"] == "regex":
            pattern = config["pattern"]
            # Handle known simple patterns
            if pattern == "^\\d{10}$":  # 10-digit number validation
                formula = f'AND(ISNUMBER({col_letter}2),LEN({col_letter}2)=10)'
                dv = DataValidation(type="custom", formula1=formula,
                                    showErrorMessage=True, error=config.get("message", "Invalid value"),allow_blank=allow_blank)
                ws.add_data_validation(dv)
                dv.add(data_range)
            else:
                # For unknown/complex regex, just add a comment instead of hard validation
                for cell in ws[data_range]:
                    for c in cell:
                        c.comment = Comment(f"Validation: {config['message']}", "System")

        elif config["type"] == "unique":
            # Enforce uniqueness using COUNTIF formula
            if allow_blank:
                formula = f'OR(LEN({col_letter}2)=0, COUNTIF(${col_letter}:${col_letter},{col_letter}2)=1)'
            else:
                formula = f'COUNTIF(${col_letter}:${col_letter},{col_letter}2)=1'
            dv_unique = DataValidation(type="custom", formula1=formula,
                                       showErrorMessage=True, error=config.get("message", "Must be unique"),allow_blank=allow_blank)
            ws.add_data_validation(dv_unique)
            dv_unique.add(data_range)

    wb.save(file_path)


def autofit_columns(
    file_path: str,
    sheet_name: str,
    auto_fit: bool = True,
    default_width: int = 20,
    max_width: int = 40,
    enable_wrap_text: bool = True,
    max_rows_to_scan: Optional[int] = None,
) -> None:
    """
    Adjust column widths in a given Excel sheet and optionally apply wrap text.

    Args:
        file_path: Path to the Excel file
        sheet_name: Name of the sheet to adjust
        auto_fit: If True, width is based on longest text length in each column
        default_width: Default width if auto_fit is False
        max_width: Maximum allowed column width
        enable_wrap_text: If True, applies wrap text to all cells
    """
    wb = load_workbook(file_path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in {file_path}")

    ws = wb[sheet_name]

    # Limit how many rows we scan per column for width calculation.
    # This keeps the operation fast on very large sheets while still
    # basing widths on real data instead of a fixed default.
    if max_rows_to_scan is not None and max_rows_to_scan > 0:
        max_row = min(ws.max_row, max_rows_to_scan)
    else:
        max_row = ws.max_row

    # One pass per column: width + wrap together (faster on large sheets than two full scans).
    wrap_alignment = Alignment(wrap_text=True) if enable_wrap_text else None

    for i, col in enumerate(ws.iter_cols(min_row=1, max_row=max_row), start=1):
        col_letter = get_column_letter(i)

        if auto_fit:
            max_length = 0
            for cell in col:
                value = cell.value
                if value is not None:
                    text = str(value)
                    if len(text) > max_length:
                        max_length = len(text)
                if wrap_alignment is not None:
                    cell.alignment = wrap_alignment

            if max_length == 0:
                max_length = default_width

            ws.column_dimensions[col_letter].width = min(max_length + 2, max_width)
        else:
            ws.column_dimensions[col_letter].width = default_width
            if wrap_alignment is not None:
                for cell in col:
                    cell.alignment = wrap_alignment

    wb.save(file_path)


def add_non_blank_validations_to_file(
    file_path: str,
    sheet_name: str,
    facility_schema: List[Dict[str, Any]],
    allow_blank_map: Dict[str, bool]
) -> None:
    """
    Add non-blank validations to an Excel sheet given file path and sheet name.

    :param file_path: Path to the Excel file
    :param sheet_name: Name of the worksheet where validations should be applied
    :param facility_schema: List of dicts with column definitions
    :param allow_blank_map: Dict {header_name: bool} to override allow_blank
    """
    try:
        wb = load_workbook(file_path)
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in {file_path}")

        ws = wb[sheet_name]

        for col_idx, col in enumerate(facility_schema, start=1):
            # Match how headers are generated
            mandatory_indicator = "(Mandatory)" if col.get("required") else ""
            header_name = f"{col.get('name')} {mandatory_indicator}".strip()

            required = col.get("required", False)

            # Skip if regex/dropdown/unique present
            has_rule = any(k in col for k in ("regex", "unique", "dropdown"))

            if required and not has_rule:
                # Default allow_blank = False for required, unless overridden
                allow_blank = allow_blank_map.get(header_name, not required)
                col_letter = get_column_letter(col_idx)

                dv = DataValidation(
                    type="custom",
                    formula1=f'LEN(TRIM({col_letter}2))>0',
                    allow_blank=allow_blank,
                    showErrorMessage=True,
                    error="This field cannot be left blank"
                )
                ws.add_data_validation(dv)
                dv.add(f"{col_letter}2:{col_letter}1048576")

        wb.save(file_path)

    except Exception as e:
        raise


def _header_base_name(header: Any) -> str:
    if header is None:
        return ""
    return str(header).replace(" (Mandatory)", "").strip().lower()


def _column_letter_for_base(letter_by_header: Dict[str, str], base_target: str) -> Optional[str]:
    for h, letter in letter_by_header.items():
        if _header_base_name(h) == base_target:
            return letter
    return None


def _facility_category_column_letter(letter_by_header: Dict[str, str]) -> Optional[str]:
    for h, letter in letter_by_header.items():
        base = _header_base_name(h)
        if base in ("Category of Facility (Mandatory)", "facility category", "category"):
            return letter
    return None


def add_facility_category_conditional_validations(file_path: str, sheet_name: str) -> None:
    """
    Excel client-side rules aligned with facility_validator:
    - HEALTH: at least one of HFR ID or NIN ID (MDMS atLeastOneRequired semantics)
    - ANGANWADI: PoC Username required
    """
    try:
        wb = load_workbook(file_path)
        if sheet_name not in wb.sheetnames:
            return

        ws = wb[sheet_name]
        letter_by_header: Dict[str, str] = {}
        for cell in ws[1]:
            if cell.value is None:
                continue
            letter_by_header[str(cell.value).strip()] = cell.column_letter

        cat_letter = _facility_category_column_letter(letter_by_header)
        hfr_letter = _column_letter_for_base(letter_by_header, "hfr id")
        nin_letter = _column_letter_for_base(letter_by_header, "nin id")
        poc_letter = _column_letter_for_base(letter_by_header, "poc username")

        max_row = max(ws.max_row + 1000, 2)

        if cat_letter and hfr_letter and nin_letter:
            hfr_or_nin_formula = (
                f'=IF(UPPER(TRIM(${cat_letter}2))<>"HEALTH",TRUE,'
                f'OR(LEN(TRIM({hfr_letter}2))>0,LEN(TRIM({nin_letter}2))>0))'
            )
            dv_hfr_nin = DataValidation(
                type="custom",
                formula1=hfr_or_nin_formula,
                allow_blank=True,
                showErrorMessage=True,
                errorTitle="Validation",
                error=ERR_HFR_OR_NIN_REQUIRED_WHEN_HEALTH,
            )
            ws.add_data_validation(dv_hfr_nin)
            dv_hfr_nin.add(f"{hfr_letter}2:{hfr_letter}{max_row}")
            dv_hfr_nin.add(f"{nin_letter}2:{nin_letter}{max_row}")

        if cat_letter and poc_letter:
            poc_formula = (
                f'=IF(UPPER(TRIM(${cat_letter}2))="ANGANWADI",LEN(TRIM({poc_letter}2))>0,TRUE)'
            )
            dv_poc = DataValidation(
                type="custom",
                formula1=poc_formula,
                allow_blank=True,
                showErrorMessage=True,
                errorTitle="Validation",
                error=ERR_POC_USERNAME_REQUIRED_WHEN_ANGANWADI,
            )
            ws.add_data_validation(dv_poc)
            dv_poc.add(f"{poc_letter}2:{poc_letter}{max_row}")

        wb.save(file_path)
    except Exception as e:
        logger.warning(f"Skipping facility category conditional Excel validation: {e}")


def add_health_category_hfr_nin_validations(
    file_path: str,
    sheet_name: str,
) -> None:
    """Backward-compatible alias; applies HEALTH HFR/NIN and ANGANWADI PoC Username rules."""
    add_facility_category_conditional_validations(file_path, sheet_name)